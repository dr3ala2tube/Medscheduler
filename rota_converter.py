#!/usr/bin/env python3
"""
RotaConverter  –  Standalone GUI app
Reads a MedScheduler "Detailed" Excel export and produces the compact
Rota board (day × specialty) matching the hospital schedule format.

Usage (macOS):
    chmod +x run_converter_mac.sh && ./run_converter_mac.sh
or:
    python3 rota_converter.py
"""

from __future__ import annotations

import os
import re
import sys
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import Dict, List, Optional, Tuple

# ── openpyxl import guard ─────────────────────────────────────────────────────
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror(
        "Missing dependency",
        "openpyxl is not installed.\n\nRun:  pip install openpyxl\nthen restart the app.",
    )
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
#  Shift-code → compact column mapping
#  Columns C-M in the rota (0-indexed offset from col C = 0..10)
# ─────────────────────────────────────────────────────────────────────────────

MORNING_COL_ORDER = [
    ("T1",      ["T1"]),
    ("T2",      ["T2"]),
    ("T3",      ["T3"]),
    ("PUL",     ["PUL", "PU"]),
    ("CA/HM",   ["CA/HM", "CAHM", "CAR", "CA", "HEMO", "HM"]),
    ("NEU",     ["NEU", "NE"]),
    ("NEPH",    ["NEPH", "NP"]),
    ("GAS",     ["GAS", "GI"]),
    ("Daycare", ["DC"]),
]

ONCALL_CODES = {"DM", "DF"}
PC_CODE       = "PC"
LEAVE_CODES   = {"L", "R"}
OFF_CODES     = {"O", "_"}

DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# Column widths (approximate, in characters)
# 9 morning cols (C–K = 3–11), oncall M/F (L/M = 12/13), legend (N = 14)
COL_WIDTHS = {
    1: 5,   # A – day #
    2: 6,   # B – day name
    3: 6, 4: 6, 5: 6, 6: 6, 7: 8, 8: 6, 9: 6, 10: 6, 11: 7,  # C-K morning (9 cols)
    12: 8,  # L – Male oncall
    13: 8,  # M – Female oncall
    14: 32,  # N – legend
}

# ─────────────────────────────────────────────────────────────────────────────
#  Colour palette (matching MedScheduler export)
# ─────────────────────────────────────────────────────────────────────────────


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=9) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _border(style="thin") -> Border:
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


NAVY   = "1F3864"
WHITE  = "FFFFFF"
BLUE_H = "2E75B6"   # header row
GRAY_H = "D6E4F0"   # sub-header
YELLOW = "FFF2CC"   # weekend highlight
GREEN_L = "E2EFDA"  # legend area
ORANGE  = "FCE4D6"  # oncall columns


# ─────────────────────────────────────────────────────────────────────────────
#  1.  Parse the Detailed Excel produced by MedScheduler
# ─────────────────────────────────────────────────────────────────────────────

def parse_detailed_xlsx(path: str) -> dict:
    """
    Returns:
        {
          "title": str,
          "year":  int | None,
          "month": int | None,
          "physicians": [ {"id": int, "name": str, "gender": str} ],
          "days": [
            {
              "day_num": int,          # 1-based
              "day_name": str,         # Mon … Sun
              "shifts": { physician_id: shift_code }
            }, ...
          ]
        }

    Handles two layouts produced by MedScheduler:

    Layout A (original):
      Row 1  : title in A1
      Row 2  : "Day"  "Date"  name1  name2  …
      Row 3  : blank / team row  (skipped)
      Rows 4+: integer  day_name  shift1  shift2 …

    Layout B (current export):
      Row 1  : title in A1
      Row 2  : None  None  name1  name2  …   (physician names)
      Row 3  : "Day"  "Date"  specialty_labels…  (column header row, skipped)
      Rows 4+: "01 We"  "2026-04-01"  shift1  shift2 …

    Day column (col A) may be a plain integer ("1") or "DD Xx" ("01 We").
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    # Try "Schedule" sheet first, then first sheet
    sheet_name = "Schedule" if "Schedule" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The selected file appears to be empty.")

    # ── Title ──
    title = str(rows[0][0]).strip() if rows[0][0] else ""
    year, month = _extract_year_month(title)

    # ── Physician names (row index 1, columns 2+) ──
    if len(rows) < 2:
        raise ValueError("File has fewer than 2 rows – not a valid detailed export.")

    header_row = rows[1]
    physicians: List[dict] = []
    for col_idx, cell_val in enumerate(header_row[2:], start=0):
        name = str(cell_val).strip() if cell_val else ""
        if not name or name.lower() in ("", "none", "nan"):
            break
        # Try to infer gender from name (DM = male, DF = female) –
        # we'll refine by actual shift codes later
        physicians.append({"id": col_idx, "name": name, "gender": "?"})

    if not physicians:
        raise ValueError("No physician names found in row 2 (columns C onward).")

    n_docs = len(physicians)

    # ── Day rows: auto-detect start row, then parse ──────────────────────────
    # Skip any rows whose col A is not a day entry (e.g. "Day", "Date", blank).
    # A day entry is either a plain integer OR starts with digits ("01 We").
    _DAY_ABBREV = {
        "mo": "Mon", "tu": "Tue", "we": "Wed", "th": "Thu",
        "fr": "Fri", "sa": "Sat", "su": "Sun",
        "mon": "Mon", "tue": "Tue", "wed": "Wed", "thu": "Thu",
        "fri": "Fri", "sat": "Sat", "sun": "Sun",
    }

    def _parse_day_cell(raw) -> Optional[Tuple[int, str]]:
        """Return (day_num, day_name) or None if not a day row."""
        if raw is None:
            return None
        s = str(raw).strip()
        m = re.match(r'^(\d+)\s*([A-Za-z]*)', s)
        if not m:
            return None
        day_num = int(m.group(1))
        abbr = m.group(2).lower()
        day_name = _DAY_ABBREV.get(abbr, abbr.capitalize() if abbr else "?")
        return day_num, day_name

    days: List[dict] = []
    for row in rows[2:]:          # start from index 2 (row 3) to cover both layouts
        # Skip / stop on completely empty rows
        if not row or row[0] is None:
            if days:
                break   # we already collected some days → end of data
            continue    # still looking for first day row

        parsed_day = _parse_day_cell(row[0])
        if parsed_day is None:
            if days:
                break   # non-day row after real data → end of data
            continue    # header / label row before data starts → skip
        day_num, day_name = parsed_day

        shifts: Dict[int, str] = {}
        for i in range(n_docs):
            col_idx = 2 + i
            val = row[col_idx] if col_idx < len(row) else None
            code = str(val).strip().upper() if val else "_"
            if code in ("", "NONE", "NAN"):
                code = "_"
            shifts[i] = code

        days.append({"day_num": day_num, "day_name": day_name, "shifts": shifts})

    # ── Infer gender from DM / DF codes ──
    gender_map: Dict[int, str] = {}
    for day in days:
        for pid, code in day["shifts"].items():
            if code == "DM":
                gender_map[pid] = "M"
            elif code == "DF":
                gender_map[pid] = "F"
    for ph in physicians:
        ph["gender"] = gender_map.get(ph["id"], "?")

    wb.close()
    return {
        "title":      title,
        "year":       year,
        "month":      month,
        "physicians": physicians,
        "days":       days,
    }


def _extract_year_month(title: str) -> Tuple[Optional[int], Optional[int]]:
    MONTHS = {
        "january": 1,  "february": 2,  "march": 3,     "april": 4,
        "may": 5,      "june": 6,      "july": 7,      "august": 8,
        "september": 9, "october": 10,  "november": 11, "december": 12,
    }
    t = title.lower()
    month = None
    year  = None
    for name, num in MONTHS.items():
        if name in t:
            month = num
            break
    m = re.search(r'\b(20\d\d)\b', title)
    if m:
        year = int(m.group(1))
    return year, month


# ─────────────────────────────────────────────────────────────────────────────
#  2.  Generate 2-letter initials (collision-safe)
# ─────────────────────────────────────────────────────────────────────────────

def generate_initials(names: List[str]) -> Dict[int, str]:
    used: set = set()
    result: Dict[int, str] = {}

    for idx, raw_name in enumerate(names):
        name = re.sub(r'\bDr\.?\s*', '', raw_name, flags=re.IGNORECASE).strip()
        words = [w for w in name.split() if w]

        def try_cand(c: str) -> Optional[str]:
            c = c.upper()[:2]
            if len(c) == 2 and c not in used:
                used.add(c)
                return c
            return None

        chosen: Optional[str] = None
        if len(words) >= 2:
            chosen = try_cand(words[0][0] + words[1][0])
        if chosen is None and words and len(words[0]) >= 2:
            chosen = try_cand(words[0][:2])
        if chosen is None:
            base = words[0][0] if words else "X"
            for suffix in "23456789ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                chosen = try_cand(base + suffix)
                if chosen:
                    break
        result[idx] = chosen or "??"

    return result


# ─────────────────────────────────────────────────────────────────────────────
#  3.  Map shift code → morning column index  (0 = T1 … 10 = Daycare)
# ─────────────────────────────────────────────────────────────────────────────

def _shift_to_morning_col(code: str) -> Optional[int]:
    """Returns 0-based index into MORNING_COL_ORDER, or None."""
    cu = code.upper()
    for idx, (label, aliases) in enumerate(MORNING_COL_ORDER):
        if cu in [a.upper() for a in aliases]:
            return idx
    return None


# ─────────────────────────────────────────────────────────────────────────────
#  4.  Build the compact Rota Excel workbook
# ─────────────────────────────────────────────────────────────────────────────

def build_rota_xlsx(parsed: dict, output_path: str) -> None:
    physicians = parsed["physicians"]
    days       = parsed["days"]
    title      = parsed["title"]

    names    = [ph["name"] for ph in physicians]
    initials = generate_initials(names)

    # Build legend: sorted alphabetically by display name
    legend_entries = sorted(
        [(initials[ph["id"]], ph["name"]) for ph in physicians],
        key=lambda x: x[1].lower(),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Rota"

    # ── Column widths ──────────────────────────────────────────────────────
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16

    # ── Row 1: Title banner ────────────────────────────────────────────────
    n_morning = len(MORNING_COL_ORDER)
    oncall_m_col  = 2 + n_morning + 1   # first col after morning block
    oncall_f_col  = oncall_m_col + 1
    legend_col    = oncall_f_col + 1
    last_col_ltr  = get_column_letter(legend_col)
    ws.merge_cells(f"A1:{last_col_ltr}1")
    c = ws["A1"]
    c.value         = title or "Rota Schedule"
    c.font          = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill          = _fill(NAVY)
    c.alignment     = _center()

    # ── Row 2: Column headers ──────────────────────────────────────────────
    col_headers = (
        ["#", "Day"] +
        [label for label, _ in MORNING_COL_ORDER] +
        ["Oncall\n(M)", "Oncall\n(F)", "Legend"]
    )
    for col_idx, hdr in enumerate(col_headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=hdr)
        c.font      = _font(bold=True, color=WHITE, size=9)
        c.fill      = _fill(BLUE_H)
        c.alignment = _center(wrap=True)
        c.border    = _border()

    # ── Day rows ────────────────────────────────────────────────────────────
    for row_offset, day in enumerate(days):
        r = 3 + row_offset
        day_num  = day["day_num"]
        day_name = day["day_name"]
        shifts   = day["shifts"]   # { physician_id: shift_code }

        is_weekend = day_name[:3] in ("Sat", "Sun", "Fri")  # adjust if needed
        row_fill   = _fill(YELLOW) if is_weekend else None

        # ── Morning specialty columns ──
        morning_cells: Dict[int, List[str]] = {i: [] for i in range(len(MORNING_COL_ORDER))}

        # ── Oncall columns ──
        oncall_m: List[str] = []
        oncall_f: List[str] = []

        for pid, code in shifts.items():
            ini = initials.get(pid, "??")
            mc = _shift_to_morning_col(code)
            if mc is not None:
                morning_cells[mc].append(ini)
            elif code == "DM":
                oncall_m.append(ini)
            elif code == "DF":
                oncall_f.append(ini)
            # PC, L, R, O, _ → not shown in rota

        # ── Write day number & name ──
        for col, val in [(1, day_num), (2, day_name)]:
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(bold=(col == 1))
            c.alignment = _center()
            c.border    = _border()
            if row_fill:
                c.fill = row_fill

        # ── Write morning columns ──
        for mc_idx, ini_list in morning_cells.items():
            col = 3 + mc_idx
            val = " / ".join(ini_list) if ini_list else ""
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font()
            c.alignment = _center()
            c.border    = _border()
            if row_fill:
                c.fill = row_fill

        # ── Write oncall ──
        for col, ini_list in [(oncall_m_col, oncall_m), (oncall_f_col, oncall_f)]:
            val = " / ".join(ini_list) if ini_list else ""
            c = ws.cell(row=r, column=col, value=val)
            c.font      = _font(bold=bool(ini_list))
            c.fill      = _fill(ORANGE) if ini_list else (_fill(YELLOW) if is_weekend else PatternFill())
            c.alignment = _center()
            c.border    = _border()

        # ── Legend column ──
        legend_row_val = ""
        if row_offset < len(legend_entries):
            ini_l, name_l = legend_entries[row_offset]
            legend_row_val = f"{ini_l} – {name_l}"
        c = ws.cell(row=r, column=legend_col, value=legend_row_val)
        c.font      = _font(size=8)
        c.fill      = _fill(GREEN_L)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _border()

    # ── If legend longer than day rows, append remaining entries ──
    n_days = len(days)
    for extra_idx in range(n_days, len(legend_entries)):
        r = 3 + extra_idx
        ini_l, name_l = legend_entries[extra_idx]
        c = ws.cell(row=r, column=legend_col, value=f"{ini_l} – {name_l}")
        c.font      = _font(size=8)
        c.fill      = _fill(GREEN_L)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _border()

    # ── Freeze panes (keep header + day# visible on scroll) ──
    ws.freeze_panes = "C3"

    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
#  5.  Tkinter GUI
# ─────────────────────────────────────────────────────────────────────────────

APP_TITLE   = "RotaConverter"
APP_VERSION = "1.0"


class RotaConverterApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_TITLE} {APP_VERSION}")
        self.resizable(False, False)
        self._build_ui()
        self._center_window()

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Banner ──────────────────────────────────────────────────────────
        banner = tk.Frame(self, bg="#1F3864")
        banner.pack(fill="x")
        tk.Label(
            banner,
            text="RotaConverter",
            font=("Helvetica", 18, "bold"),
            fg="white",
            bg="#1F3864",
            pady=12,
        ).pack(side="left", padx=20)
        tk.Label(
            banner,
            text="MedScheduler Detailed Export  →  Compact Rota Board",
            font=("Helvetica", 10),
            fg="#A9C4E8",
            bg="#1F3864",
        ).pack(side="left")

        main = ttk.Frame(self, padding=16)
        main.pack(fill="both", expand=True)

        # ── Input file ──────────────────────────────────────────────────────
        in_frame = ttk.LabelFrame(main, text="Input – Detailed Excel (.xlsx)", padding=10)
        in_frame.pack(fill="x", pady=(0, 10))

        self.in_var = tk.StringVar()
        ttk.Entry(in_frame, textvariable=self.in_var, width=58).pack(side="left", padx=(0, 6))
        ttk.Button(in_frame, text="Browse…", command=self._browse_input).pack(side="left")

        # ── Detected-info strip ─────────────────────────────────────────────
        self.info_var = tk.StringVar(value="No file loaded.")
        info_lbl = ttk.Label(main, textvariable=self.info_var, foreground="#2E75B6")
        info_lbl.pack(anchor="w", pady=(0, 8))

        # ── Output file ─────────────────────────────────────────────────────
        out_frame = ttk.LabelFrame(main, text="Output – Compact Rota (.xlsx)", padding=10)
        out_frame.pack(fill="x", pady=(0, 14))

        self.out_var = tk.StringVar()
        ttk.Entry(out_frame, textvariable=self.out_var, width=58).pack(side="left", padx=(0, 6))
        ttk.Button(out_frame, text="Browse…", command=self._browse_output).pack(side="left")

        # ── Convert button ───────────────────────────────────────────────────
        btn_row = ttk.Frame(main)
        btn_row.pack(fill="x")
        self.convert_btn = ttk.Button(
            btn_row,
            text="Convert  ▶",
            command=self._convert,
            width=18,
        )
        self.convert_btn.pack(side="right")

        # ── Status bar ───────────────────────────────────────────────────────
        sep = ttk.Separator(self, orient="horizontal")
        sep.pack(fill="x", pady=(8, 0))

        self.status_var = tk.StringVar(value="Ready.")
        status_bar = ttk.Label(
            self,
            textvariable=self.status_var,
            anchor="w",
            padding=(10, 4),
        )
        status_bar.pack(fill="x")

    def _center_window(self):
        self.update_idletasks()
        w, h = self.winfo_reqwidth(), self.winfo_reqheight()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"+{(sw - w) // 2}+{(sh - h) // 2}")

    # ── File browsing ────────────────────────────────────────────────────────

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="Select Detailed Excel Export",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if not path:
            return
        self.in_var.set(path)
        # Auto-fill output path
        base, _ = os.path.splitext(path)
        self.out_var.set(base + "_Rota.xlsx")
        # Try to peek at the file
        self._peek_file(path)

    def _peek_file(self, path: str):
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            sheet_name = "Schedule" if "Schedule" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(max_row=4, values_only=True))
            wb.close()
            title = str(rows[0][0]).strip() if rows and rows[0][0] else "(untitled)"
            # Count physicians in row 2
            n_docs = 0
            if len(rows) > 1:
                for v in rows[1][2:]:
                    if v and str(v).strip() not in ("", "None", "nan"):
                        n_docs += 1
                    else:
                        break
            self.info_var.set(
                f'Detected: \u201c{title}\u201d  |  {n_docs} physician(s) found'
            )
            self.status_var.set("File loaded. Press Convert to generate the Rota.")
        except Exception as e:
            self.info_var.set(f"Could not read file: {e}")

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="Save Compact Rota As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.out_var.set(path)

    # ── Conversion ───────────────────────────────────────────────────────────

    def _convert(self):
        in_path  = self.in_var.get().strip()
        out_path = self.out_var.get().strip()

        if not in_path:
            messagebox.showwarning("No input file", "Please select the Detailed Excel file to convert.")
            return
        if not os.path.isfile(in_path):
            messagebox.showerror("File not found", f"Cannot find:\n{in_path}")
            return
        if not out_path:
            messagebox.showwarning("No output path", "Please specify where to save the Rota file.")
            return

        self.status_var.set("Parsing detailed export…")
        self.update_idletasks()

        try:
            parsed = parse_detailed_xlsx(in_path)
        except Exception as e:
            self.status_var.set(f"Parse error: {e}")
            messagebox.showerror("Parse Error", str(e))
            return

        n_docs = len(parsed["physicians"])
        n_days = len(parsed["days"])

        self.status_var.set(f"Building rota ({n_docs} physicians, {n_days} days)…")
        self.update_idletasks()

        try:
            build_rota_xlsx(parsed, out_path)
        except Exception as e:
            self.status_var.set(f"Export error: {e}")
            messagebox.showerror("Export Error", str(e))
            return

        self.status_var.set(
            f"Done! Rota saved ({n_docs} physicians, {n_days} days) → {os.path.basename(out_path)}"
        )

        if messagebox.askyesno(
            "Success",
            f"Rota exported successfully!\n\n{out_path}\n\nOpen file now?",
        ):
            self._open_file(out_path)

    @staticmethod
    def _open_file(path: str):
        try:
            if sys.platform == "darwin":
                subprocess.call(["open", path])
            elif sys.platform.startswith("win"):
                os.startfile(path)  # type: ignore[attr-defined]
            else:
                subprocess.call(["xdg-open", path])
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = RotaConverterApp()
    app.mainloop()
