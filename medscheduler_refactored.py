from __future__ import annotations

import calendar
from dataclasses import dataclass
from datetime import date, datetime
import re
import tkinter as tk
import xml.etree.ElementTree as ET
from zipfile import ZipFile
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional
import threading

try:
    from firebase_service import firebase, FirebaseAuthError, FirebaseNetworkError
    _FIREBASE_AVAILABLE = True
except ImportError:
    _FIREBASE_AVAILABLE = False
    firebase = None  # type: ignore
    FirebaseAuthError = Exception  # type: ignore
    FirebaseNetworkError = Exception  # type: ignore

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]
DN = ["Su", "Mo", "Tu", "We", "Th", "Fr", "Sa"]
TEAMS = ["T1", "T2", "T3"]
SUBS = ["CAHM", "GI", "NE", "NP", "PU"]
MORNING_K = TEAMS + SUBS + ["NENP"]   # NENP = combined Neurology+Nephrology escape
DUTY_SET = {"DM", "DF"}
OFF_SET = {"PC", "O", "L", "R"}

SHIFTS = {
    "T1": {"label": "Team 1 Morning", "short": "T1", "h": 8},
    "T2": {"label": "Team 2 Morning", "short": "T2", "h": 8},
    "T3": {"label": "Team 3 Morning", "short": "T3", "h": 8},
    "CAHM": {"label": "Cardiology / Hematology", "short": "CA/HM", "h": 8},
    "GI": {"label": "Gastroenterology", "short": "GI", "h": 8},
    "NE": {"label": "Neurology", "short": "NE", "h": 8},
    "NP": {"label": "Nephrology", "short": "NP", "h": 8},
    "NENP": {"label": "Neurology + Nephrology", "short": "NE/NP", "h": 8},
    "PU": {"label": "Pulmonology", "short": "PU", "h": 8},
    "DC": {"label": "Daycare Clinic", "short": "DC", "h": 8},
    "DM": {"label": "16hr Duty – Male Side", "short": "DM", "h": 16},
    "DF": {"label": "16hr Duty – Female Side", "short": "DF", "h": 16},
    "PC": {"label": "Post-Call Off", "short": "PC", "h": 0},
    "O": {"label": "Day Off", "short": "O", "h": 0},
    "L": {"label": "Annual Leave", "short": "L", "h": 0},
    "R": {"label": "Random Off Day", "short": "R", "h": 0},
    "_": {"label": "—", "short": "", "h": 0},
}

COLOR_MAP = {
    "T1": "DBEAFE", "T2": "CCFBF1", "T3": "EDE9FE",
    "CAHM": "FFE4E6", "GI": "FFEDD5", "NE": "F3E8FF",
    "NP": "CFFAFE", "NENP": "C7D7F9", "PU": "E0F2FE",
    "DC": "FEF3C7", "DM": "FEF9C3",
    "DF": "FCE7F3", "PC": "EDE9FE", "O": "F3F4F6",
    "L": "D1FAE5", "R": "FEE2E2", "_": "FFFFFF",
}

SPEC_OPTIONS = ["Not specified"] + [SHIFTS[k]["label"] for k in ["T1", "T2", "T3", "CAHM", "GI", "NE", "NP", "PU"]]
# All assignable duty/specialty codes available in the manual assignment picker
MANUAL_ASSIGN_CODES = ["T1", "T2", "T3", "PU", "CAHM", "NE", "NP", "GI", "DC", "DM", "DF"]
BLOCKABLE_SPECIALTIES = MORNING_K + ["DC"]

# ── Schedule grid layout constants ───────────────────────────────────────────
_NAME_W = 190   # physician name column width (pixels)
_CELL_W = 44    # day cell width (pixels)
_CELL_H = 30    # row height for each physician (pixels)
_HDR_H  = 50    # day-number header height (pixels)


@dataclass
class Doctor:
    id: int
    name: str
    spec: str
    team: str
    initials: str = ""   # user-supplied short label used in exports (auto-generated if blank)
    first_duty_day: int = 1  # day-of-month to start first duty (1 = no stagger override)


@dataclass
class LeaveBlock:
    id: int
    pid: int
    f: str
    t: str


@dataclass
class SpecialtyBlock:
    id: int
    code: str
    f: str
    t: str


@dataclass
class ManualAssignment:
    """A single user-defined duty assignment: physician on a specific day."""
    id: int
    pid: int    # physician id
    code: str   # e.g. "T1", "DM", "NE"
    day: int    # 1-based day-of-month


def p2(n: int) -> str:
    return str(n).zfill(2)


def dim(y: int, m: int) -> int:
    return calendar.monthrange(y, m + 1)[1]


def ds(y: int, m: int, d: int) -> str:
    return f"{y}-{p2(m + 1)}-{p2(d)}"


def day_of_week(y: int, m: int, d: int) -> int:
    py = date(y, m + 1, d).weekday()  # Mon=0..Sun=6
    return (py + 1) % 7  # Sun=0..Sat=6


def is_we(y: int, m: int, d: int) -> bool:
    return day_of_week(y, m, d) in (0, 6)


def is_wd(y: int, m: int, d: int) -> bool:
    return not is_we(y, m, d)


def specialty_code_from_label(label: str) -> Optional[str]:
    for code, meta in SHIFTS.items():
        if meta["label"] == label:
            return code
    return None


def auto_schedule(
    docs: List[Doctor],
    base_asgn: Dict[str, str],
    leaves: List[LeaveBlock],
    spec_blocks: List[SpecialtyBlock],
    y: int,
    m: int,
):
    """
    Rule-based scheduling engine.

    Hard constraints (strictly enforced):
      - Leave (L) and Random-off (R) are never overridden by any phase.
      - Every DM/DF assignment is followed by exactly 3 PC days; those PC
        days override any previously-placed morning specialty.
      - Specialty Lock: a physician keeps the SAME morning specialty from
        first assignment until their next 16-hour duty (DM/DF).  The lock
        is released ONLY after a DM/DF, never because of a specialty block.
        If the locked specialty is blocked on a particular day the physician
        receives a day-off (O) for that day while the lock remains intact.
      - Maximum 6 consecutive working days.
      - Monthly hours: floor 160 h, hard ceiling 168 h.
          Plain O days are converted to 8-h morning shifts (Phase 8) to
          reach the 160 h floor, unless the shortfall is caused by L
          (annual leave) or R (random off day) or user-pinned O days.
          The 168 h hard ceiling is always respected.

    Soft goals:
      - 1 DM + 1 DF per calendar day.
      - Duty count balanced across physicians (sort by fewest duties first).
      - At least one weekend day off per physician.
      - Specialty usage spread across the team.
    """
    if len(docs) < 3:
        return {"err": "Need at least 3 physicians (one per IM team)."}

    td = dim(y, m)
    a = dict(base_asgn)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def ak(pid: int, d: int) -> str:
        return f"{pid}|{y}|{m}|{d}"

    def get(pid: int, d: int) -> str:
        return a.get(ak(pid, d), "_")

    def setv(pid: int, d: int, v: str) -> None:
        a[ak(pid, d)] = v

    def day_str(d: int) -> str:
        return ds(y, m, d)

    def is_spec_blocked(code: str, d: int) -> bool:
        cur = day_str(d)
        return any(b.code == code and b.f <= cur <= b.t for b in spec_blocks)

    def calc_h(doc_id: int) -> int:
        return sum(SHIFTS[get(doc_id, d)]["h"] for d in range(1, td + 1))

    def consecutive_working_days_before(pid: int, d: int) -> int:
        """Count unbroken working days immediately before day d."""
        streak = 0
        x = d - 1
        while x >= 1 and get(pid, x) not in OFF_SET and get(pid, x) != "_":
            streak += 1
            x -= 1
        return streak

    def preplaced_working_days_after(pid: int, d: int) -> int:
        """Count consecutive pre-placed working days starting at d+1.

        DM/DF duties are counted as working days so that the streak check
        correctly blocks assigning a morning shift immediately before a duty
        (which would produce a >6-consecutive-day run).  Previously DM/DF were
        treated as streak-breakers, which allowed streaks of 6 mornings + 1 duty
        = 7 consecutive working days — a hard-rule violation."""
        streak = 0
        x = d + 1
        while x <= td:
            code = get(pid, x)
            if code in OFF_SET or code == "_":
                break   # off day or blank → streak resets here; DM/DF count as working
            streak += 1
            x += 1
        return streak

    # ------------------------------------------------------------------
    # Phase 0 – Stamp first_duty_day offsets (user-defined start days)
    # Must run before Phase 4/5 so DC/DM/DF are never placed before the
    # physician's requested first duty day.
    # ------------------------------------------------------------------
    for ph in docs:
        fdd = int(ph.first_duty_day) if ph.first_duty_day and ph.first_duty_day > 1 else 1
        fdd = min(fdd, td)
        for _d in range(1, fdd):
            if get(ph.id, _d) == "_":
                setv(ph.id, _d, "O")
            # Also add to unavail so Phase 4/5 skip these days
            # (unavail is built in Phase 2 — we'll mark them there)

    # ------------------------------------------------------------------
    # Phase 1 – Stamp leave
    # ------------------------------------------------------------------
    for b in leaves:
        for d in range(1, td + 1):
            if b.f <= day_str(d) <= b.t:
                setv(b.pid, d, "L")

    # ------------------------------------------------------------------
    # Phase 2 – Build hard-blocked day sets (L and R are immovable)
    # ------------------------------------------------------------------
    hb: Dict[int, set] = {}
    for ph in docs:
        hb[ph.id] = set()
        for d in range(1, td + 1):
            code = get(ph.id, d)
            if code in ("L", "R"):
                hb[ph.id].add(d)
            # FDD-stamped "O" days are also hard-blocked for Phase 4/5 so that
            # DC/DM/DF duties are never placed before a physician's first duty day.
            fdd = int(ph.first_duty_day) if ph.first_duty_day and ph.first_duty_day > 1 else 1
            if d < fdd:
                hb[ph.id].add(d)

    # Running counters
    unavail: Dict[int, set] = {ph.id: set(hb[ph.id]) for ph in docs}
    duty_cnt: Dict[int, int] = {ph.id: 0 for ph in docs}
    male_side_cnt: Dict[int, int] = {ph.id: 0 for ph in docs}
    dc_wk_cnt: Dict[int, int] = {ph.id: 0 for ph in docs}

    # pinned: per-physician set of days that are pre-assigned in base_asgn with
    # a non-blank, non-off code.  These are treated as immovable hard constraints:
    # Phase 4/5 will not assign DC/DM/DF if it would require PC on a pinned day,
    # and assign_duty will never overwrite a pinned day with PC.
    _pinned_codes = set(SHIFTS.keys()) - {"_", "O", "L", "R", "PC"}
    pinned: Dict[int, set] = {ph.id: set() for ph in docs}
    for ph in docs:
        for d in range(1, td + 1):
            if get(ph.id, d) in _pinned_codes:
                pinned[ph.id].add(d)
    # Add pinned days to hb and unavail so Phase 4/5 skip them for new assignments
    for ph in docs:
        hb[ph.id]     |= pinned[ph.id]
        unavail[ph.id] |= pinned[ph.id]

    # Preferred morning code per physician.
    # • Named specialty (e.g. "Cardiology") → prefer that code.
    # • "Not specified" → None, meaning no preference; the scheduler will
    #   distribute the doctor to whichever specialty has the least coverage
    #   on each cycle, balancing workload across all open specialties.
    # • Any other unrecognised value → fall back to the doctor's IM team code.
    pref_spec: Dict[int, Optional[str]] = {}
    for i, ph in enumerate(docs):
        if ph.spec == "Not specified":
            pref_spec[ph.id] = None          # no preference – use least-used
        else:
            sc = specialty_code_from_label(ph.spec)
            if sc and sc in MORNING_K:
                pref_spec[ph.id] = sc
            else:
                pref_spec[ph.id] = ph.team if ph.team in TEAMS else TEAMS[i % 3]

    # ------------------------------------------------------------------
    # assign_duty – central helper used by all duty-assignment phases
    # ------------------------------------------------------------------
    def assign_duty(pid: int, d: int, side: str) -> bool:
        """
        Assign a 16-hour duty on day d and stamp the 3 mandatory PC days
        that follow.

        PC days override any previously-placed morning specialty so that the
        hard post-call rule is never silently skipped.  Leave (L) and random-
        off (R) days are never overridden.

        Returns False (and makes no change) if the hard hour ceiling would be
        exceeded.
        """
        # Enforce 6-day consecutive limit – reject if duty would create streak > 6
        if consecutive_working_days_before(pid, d) >= 6:
            return False
        # If assigning this duty would push total hours over the soft limit
        # AND the physician worked a morning shift the day before, convert that
        # morning to O first to stay within bounds.
        if d > 1 and get(pid, d - 1) in MORNING_K and calc_h(pid) + 16 > 160:
            setv(pid, d - 1, "O")
        if calc_h(pid) + 16 > 168:  # hard ceiling – cannot assign
            return False

        # Check that none of the 3 PC days are manually pinned or pre-placed DC days.
        # A physician cannot be post-call during a Daycare week — it leaves the
        # clinic uncovered.  Refuse the duty if any PC day lands on a DC day.
        for _pc in range(1, 4):
            _pd = d + _pc
            if _pd <= td and _pd in pinned.get(pid, set()):
                return False   # pinned day — skip this candidate
            if _pd <= td and get(pid, _pd) == "DC":
                return False   # DC day must not be overwritten with PC

        setv(pid, d, side)
        duty_cnt[pid] += 1
        if side == "DM":
            male_side_cnt[pid] += 1
        unavail[pid].add(d)

        # Stamp 3 mandatory PC days.
        # Leave (L), random-off (R), Daycare (DC), other duty (DM/DF), and
        # manually PINNED days are never overridden.  Morning specialty
        # assignments are overridden unless they are pinned.
        for pc in range(1, 4):
            pd = d + pc
            if pd > td:
                break
            existing = get(pid, pd)
            if existing in ("L", "R", "DM", "DF", "DC"):
                continue   # immovable
            if pd in pinned.get(pid, set()):
                continue   # pinned manual assignment — preserve it
            setv(pid, pd, "PC")
            unavail[pid].add(pd)  # protect these days from further assignment
        return True

    # ------------------------------------------------------------------
    # Phase 3 – Build calendar weeks
    # ------------------------------------------------------------------
    weeks = []
    start_d = 1
    while start_d <= td:
        dw = day_of_week(y, m, start_d)
        days_to_sun = 0 if dw == 0 else 7 - dw
        end_d = min(start_d + days_to_sun, td)
        wdays = [x for x in range(start_d, end_d + 1) if is_wd(y, m, x)]
        wends = [x for x in range(start_d, end_d + 1) if is_we(y, m, x)]
        weeks.append({"start": start_d, "end": end_d, "wdays": wdays, "wends": wends})
        start_d = end_d + 1

    # ------------------------------------------------------------------
    # Phase 4 – Daycare rotation
    # One physician per week: Mon–Fri = DC, weekends = O.
    # Physician is selected by fewest DC weeks assigned so far (fairness),
    # then fewest total duties, then stable tie-break on id.
    # ------------------------------------------------------------------
    dc_info = []
    for wk in weeks:
        if not wk["wdays"]:
            continue
        elig = [ph for ph in docs if any(d not in hb[ph.id] for d in wk["wdays"])]
        if not elig:
            dc_info.append({"docId": None, "weekEnd": wk["end"]})
            continue
        # Sort key:
        #  1. Fewest DC weeks so far   (fairness)
        #  2. Duties in the 3 days immediately before the week starts
        #     — physicians with a recent DM/DF would need PC on days 1-3 of
        #       their DC week, leaving those weekdays without Daycare coverage.
        #     Prefer physicians with NO such overlap.
        #  3. Current consecutive-day streak (avoid creating >6-day runs)
        #  4. Fewest total duties       (fairness)
        #  5. Stable tie-break          (physician id)
        first_wd = wk["wdays"][0]
        elig.sort(key=lambda ph: (
            dc_wk_cnt[ph.id],
            sum(1 for _dp in range(max(1, first_wd - 3), first_wd)
                if get(ph.id, _dp) in ("DM", "DF")),
            consecutive_working_days_before(ph.id, first_wd),
            duty_cnt[ph.id],
            ph.id,
        ))
        doc = elig[0]
        dc_wk_cnt[doc.id] += 1
        for d in wk["wdays"]:
            # Skip days that already have a DC assignment (pinned or otherwise)
            if any(get(ph.id, d) == "DC" for ph in docs if ph.id != doc.id):
                continue  # day already has a DC physician — don't add a second
            if d not in hb[doc.id] and not is_spec_blocked("DC", d):
                # Enforce 6-day consecutive limit within DC block
                if consecutive_working_days_before(doc.id, d) >= 6:
                    setv(doc.id, d, "O")  # break streak
                else:
                    setv(doc.id, d, "DC")
                unavail[doc.id].add(d)
        for d in wk["wends"]:
            if d not in hb[doc.id] and get(doc.id, d) == "_":
                setv(doc.id, d, "O")
                unavail[doc.id].add(d)
        dc_info.append({"docId": doc.id, "weekEnd": wk["end"]})

    # Post-DC duty: one 16-hour duty on the first available weekday after
    # the DC week, followed automatically by 3 PC days (handled inside
    # assign_duty).
    for info in dc_info:
        doc_id = info["docId"]
        week_end = info["weekEnd"]
        if not doc_id:
            continue
        for d in range(week_end + 1, td + 1):
            if is_wd(y, m, d) and d not in unavail[doc_id] and get(doc_id, d) == "_":
                # Enforce 6-day consecutive limit before assigning post-DC duty
                if consecutive_working_days_before(doc_id, d) >= 6:
                    continue
                side = "DM" if male_side_cnt[doc_id] <= (duty_cnt[doc_id] - male_side_cnt[doc_id]) else "DF"
                assign_duty(doc_id, d, side)
                break

    # ------------------------------------------------------------------
    # Phase 4.5 – DC gap-fill safety net
    # Phase 4 assigns one physician per full calendar week for Daycare.
    # However, if the selected physician has a duty 1–3 days before the
    # week starts, their mandatory PC days can overlap the first days of
    # their DC week (the assign_duty guard above now prevents the
    # overwrite, but the days remain blank).  This pass fills any
    # remaining blank weekdays with the best available physician so that
    # the Daycare clinic never has zero coverage on a working day.
    # ------------------------------------------------------------------
    for _d in range(1, td + 1):
        if not is_wd(y, m, _d):
            continue
        if is_spec_blocked("DC", _d):
            continue
        if any(get(ph.id, _d) == "DC" for ph in docs):
            continue   # already covered
        # Find the best available physician for a standalone DC day
        _dc_gap = []
        for ph in docs:
            if get(ph.id, _d) != "_":
                continue   # already assigned something
            if _d in unavail[ph.id]:
                continue
            if consecutive_working_days_before(ph.id, _d) >= 6:
                continue
            if calc_h(ph.id) + 8 > 168:
                continue
            _dc_gap.append(ph)
        if _dc_gap:
            _dc_gap.sort(key=lambda ph: (dc_wk_cnt[ph.id], duty_cnt[ph.id],
                                         calc_h(ph.id), ph.id))
            _filler = _dc_gap[0]
            setv(_filler.id, _d, "DC")
            unavail[_filler.id].add(_d)
            dc_wk_cnt[_filler.id] += 1

    # ------------------------------------------------------------------
    # Phase 5 – Daily 16-hour duty coverage (DM + DF per day)
    # Candidates are sorted by fewest duties first (true fairness),
    # then fewest total hours as a tiebreaker.
    # The streak heuristic that previously overrode duty-count fairness has
    # been removed; consecutive-day enforcement is handled in Phase 7.
    # ------------------------------------------------------------------
    pairs = []
    for d in range(1, td + 1):
        exist_dm = next((ph for ph in docs if get(ph.id, d) == "DM"), None)
        exist_df = next((ph for ph in docs if get(ph.id, d) == "DF"), None)
        need_dm = exist_dm is None
        need_df = exist_df is None
        if not need_dm and not need_df:
            pairs.append({"d": d, "male": exist_dm.id, "female": exist_df.id})
            continue

        # Build candidate list
        avail = []
        for ph in docs:
            if d in unavail[ph.id] or get(ph.id, d) != "_":
                continue
            streak = consecutive_working_days_before(ph.id, d)
            if streak >= 6:
                continue  # would violate consecutive-day rule
            avail.append((ph, duty_cnt[ph.id], calc_h(ph.id)))

        # Sort strictly by fairness: fewest duties → fewest hours → stable id
        avail.sort(key=lambda x: (x[1], x[2], x[0].id))
        just_docs = [x[0] for x in avail]

        m_doc = exist_dm
        f_doc = exist_df

        if need_dm and need_df:
            if len(just_docs) >= 2:
                d1, d2 = just_docs[0], just_docs[1]
                # Give DM (male side) to the one with fewer DM assignments
                if male_side_cnt[d1.id] > (duty_cnt[d1.id] - male_side_cnt[d1.id]):
                    d1, d2 = d2, d1
                m_doc, f_doc = d1, d2
            elif len(just_docs) == 1:
                m_doc = just_docs[0]
        elif need_dm and just_docs:
            m_doc = just_docs[0]
        elif need_df and just_docs:
            f_doc = just_docs[0]

        assigned_m = assigned_f = None
        if m_doc and need_dm and assign_duty(m_doc.id, d, "DM"):
            assigned_m = m_doc.id
        if f_doc and need_df and (not assigned_m or f_doc.id != assigned_m):
            if assign_duty(f_doc.id, d, "DF"):
                assigned_f = f_doc.id

        pairs.append({
            "d": d,
            "male": assigned_m or getattr(exist_dm, "id", None),
            "female": assigned_f or getattr(exist_df, "id", None),
        })

    # ------------------------------------------------------------------
    # Phase 5.5 – DM/DF rescue pass
    # After the main Phase 5 loop some days may still lack DM, DF, or both
    # (e.g. when all candidates were in unavail).  This rescue pass looks at
    # physicians who currently have a morning specialty (not DC/PC/DM/DF/L/R)
    # and reassigns the least-loaded one to the missing side.  Their morning
    # slot is vacated (becomes blank) so Phase 7 can reassign it to another
    # physician.
    # ------------------------------------------------------------------
    for _d in range(1, td + 1):
        _has_dm = any(get(ph.id, _d) == "DM" for ph in docs)
        _has_df = any(get(ph.id, _d) == "DF" for ph in docs)
        for _side, _needed in (("DM", not _has_dm), ("DF", not _has_df)):
            if not _needed:
                continue
            _rescue = []
            for ph in docs:
                _code = get(ph.id, _d)
                # Only reassign physicians with a morning specialty or blank slot
                if _code not in MORNING_K and _code != "_":
                    continue
                if consecutive_working_days_before(ph.id, _d) >= 6:
                    continue
                if calc_h(ph.id) + 16 > 168:
                    continue
                # Refuse if PC days would land on a pinned or DC day
                _ok = True
                for _pc in range(1, 4):
                    _pd = _d + _pc
                    if _pd <= td:
                        if _pd in pinned.get(ph.id, set()):
                            _ok = False
                            break
                        if get(ph.id, _pd) == "DC":
                            _ok = False
                            break
                if _ok:
                    _rescue.append(ph)
            if _rescue:
                _rescue.sort(key=lambda ph: (duty_cnt[ph.id], calc_h(ph.id), ph.id))
                _rph = _rescue[0]
                # Clear morning specialty so Phase 7 can reassign it
                if get(_rph.id, _d) in MORNING_K:
                    setv(_rph.id, _d, "_")
                assign_duty(_rph.id, _d, _side)

    # ------------------------------------------------------------------
    # Phase 5.6 – Minimum on-call guarantee (target: ≥ 3 per physician)
    # After Phase 5 + 5.5, scan for physicians below the minimum duty
    # count and try to assign additional duties without violating hard
    # constraints (streak, hours, pinned days, DC overlap).
    # ------------------------------------------------------------------
    MIN_DUTIES = 3
    for ph in docs:
        while duty_cnt[ph.id] < MIN_DUTIES:
            placed = False
            for _d in range(1, td + 1):
                cur = get(ph.id, _d)
                # Only consider blank or morning-specialty slots
                if cur not in ("_",) and cur not in MORNING_K:
                    continue
                if consecutive_working_days_before(ph.id, _d) >= 6:
                    continue
                if calc_h(ph.id) + 16 > 168:
                    break  # no point scanning further days
                # Check PC days won't land on pinned or DC
                _ok = True
                for _pc in range(1, 4):
                    _pd = _d + _pc
                    if _pd <= td:
                        if _pd in pinned.get(ph.id, set()):
                            _ok = False
                            break
                        if get(ph.id, _pd) == "DC":
                            _ok = False
                            break
                if not _ok:
                    continue
                # Balance DM/DF sides for this physician
                _side = "DM" if male_side_cnt[ph.id] <= (duty_cnt[ph.id] - male_side_cnt[ph.id]) else "DF"
                if cur in MORNING_K:
                    setv(ph.id, _d, "_")
                if assign_duty(ph.id, _d, _side):
                    placed = True
                    break
                else:
                    if cur in MORNING_K:
                        setv(ph.id, _d, cur)  # restore on reject
            if not placed:
                break  # cannot reach minimum without violating hard constraints

    # ------------------------------------------------------------------
    # Phase 6 – Guarantee a full happy weekend (Sat + Sun both off)
    #           with a single-day fallback when a pair is not possible.
    # ------------------------------------------------------------------
    wkends = [d for d in range(1, td + 1) if is_we(y, m, d)]
    # Build Sat+Sun pairs for this month (Saturday=6, Sunday=0)
    we_pairs: List[tuple] = []
    for _d in range(1, td + 1):
        if day_of_week(y, m, _d) == 6 and _d + 1 <= td and day_of_week(y, m, _d + 1) == 0:
            we_pairs.append((_d, _d + 1))
    for ph in docs:
        # First pass: attempt to guarantee a full Sat+Sun happy weekend.
        # IMPORTANT: blank ("_") slots are NOT confirmed off — Phase 7 will fill them
        # with morning specialties.  Only count confirmed OFF_SET codes here so we
        # don't falsely conclude the physician already has a happy weekend.
        has_happy = any(
            get(ph.id, sat) in OFF_SET and get(ph.id, sun) in OFF_SET
            for sat, sun in we_pairs
        )
        if not has_happy:
            # Stamp "O" on both days of the first free Sat+Sun pair BEFORE Phase 7
            # runs.  Phase 7 will see those slots as off and assign coverage to others.
            for sat, sun in we_pairs:
                sc, uc = get(ph.id, sat), get(ph.id, sun)
                # "Free" = not locked by duty/leave/DC; blank or morning can be cleared
                sat_free = sc not in DUTY_SET and sc not in ("L", "R", "DC")
                sun_free = uc not in DUTY_SET and uc not in ("L", "R", "DC")
                if sat_free and sun_free:
                    if sc not in OFF_SET:
                        setv(ph.id, sat, "O")
                    if uc not in OFF_SET:
                        setv(ph.id, sun, "O")
                    break
        # Fallback: guarantee at least one single weekend day off.
        # Don't count "_" as off here — Phase 7 will fill it.
        has_we_off = any(get(ph.id, d) in OFF_SET for d in wkends)
        if not has_we_off:
            for w in wkends:
                if get(ph.id, w) not in DUTY_SET and get(ph.id, w) not in ("L", "R", "DC"):
                    setv(ph.id, w, "O")
                    break

    # ------------------------------------------------------------------
    # Phase 7 – DAY-FIRST morning specialty fill with guaranteed coverage
    # -----------------------------------------------------------------------
    # Processes one calendar day at a time so that all required morning
    # specialties are fully covered before any physician is given a second
    # slot on the same day.
    #
    # Coverage priority per day
    # ─────────────────────────
    #   1. All SUBS (CAHM, GI, NE, NP, PU) → each covered by exactly 1 physician.
    #   2. All TEAMS (T1, T2, T3)           → each covered by at least 1 physician.
    #   3. Remaining available physicians   → overflow into team slots only.
    #
    # Specialty lock semantics (unchanged)
    # ─────────────────────────────────────
    #   • A physician's lock is set when they are first assigned a specialty.
    #   • The lock persists through PC, O, DC, and blocked days.
    #   • The lock is released ONLY on a DM/DF duty day.
    #   • Locked physicians stay on their specialty each working day.
    #   • If a locked physician's specialty is already covered (lock conflict)
    #     or is blocked, they receive a day-off; the lock is preserved.
    # -----------------------------------------------------------------------

    # Global fairness counter – tracks total times each specialty was assigned.
    spec_usage: Dict[str, int] = {code: 0 for code in MORNING_K}
    for _d in range(1, td + 1):
        for _ph in docs:
            _code = get(_ph.id, _d)
            if _code in MORNING_K:
                spec_usage[_code] += 1

    # Per-physician lock state, persisted across the day loop.
    lock_map: Dict[int, Optional[str]] = {ph.id: None for ph in docs}

    # ── Pre-stagger: set physician start dates ───────────────────────────────
    # Phase 0 already stamped "O" on FDD-blocked days (before Phase 4/5).
    # Here we only need the random stagger for physicians without a custom FDD.
    # Physicians with first_duty_day > 1 are excluded — their start is already fixed.
    import random as _rnd
    _rng = _rnd.Random(y * 1000 + (m + 1) * 31)   # stable seed per year+month
    _default_docs = [ph for ph in docs if not (ph.first_duty_day and ph.first_duty_day > 1)]

    # Random stagger only for physicians without a custom first duty day.
    _shuffled = list(_default_docs)
    _rng.shuffle(_shuffled)
    _slots_per_day = len(SUBS) + len(TEAMS)   # 8 required morning slots
    # Keep enough physicians un-staggered on day 1 to guarantee coverage even
    # if some are already committed to DM/DF/DC.  Buffer = slots + 4 spares.
    _first_group = _slots_per_day + 4          # e.g. 12
    _stagger_max = 5                           # offset at most 5 days
    for _gi, _ph in enumerate(_shuffled):
        if _gi < _first_group:
            _offset = 0                        # first group: always starts day 1
        else:
            _offset = min((_gi - _first_group) // _slots_per_day + 1, _stagger_max)
        for _sd in range(1, _offset + 1):
            if get(_ph.id, _sd) == "_":
                setv(_ph.id, _sd, "O")

    for d in range(1, td + 1):

        # ── Step 0: update lock_map from pre-existing grid entries ──────────
        # DM/DF on today's date releases the lock; a pre-placed morning code
        # confirms (or seeds) it; all other codes leave the lock unchanged.
        for ph in docs:
            code = get(ph.id, d)
            if code in DUTY_SET:
                lock_map[ph.id] = None
            elif code in MORNING_K:
                lock_map[ph.id] = code

        # ── Step 1: gather physicians with a blank slot today ────────────────
        # Hard constraints: 6-day consecutive limit and 168-hour ceiling.
        # Soft constraint: physicians at streak=5 are given proactive rest
        # when enough other physicians are available, staggering rest days
        # across the month to avoid coverage cliffs where everyone hits the
        # hard limit on the same day.
        to_assign: List[Doctor] = []
        can_rest_early: List[Doctor] = []  # streak=5, eligible for proactive rest
        for ph in docs:
            if get(ph.id, d) != "_":
                continue
            if calc_h(ph.id) + 8 > 168:
                setv(ph.id, d, "O")   # hard hour ceiling
                continue
            backward = consecutive_working_days_before(ph.id, d)
            if backward >= 6:
                setv(ph.id, d, "O")   # hard 6-day consecutive-day limit
                continue
            # Check forward: if pre-placed DC/DM/DF days follow,
            # assigning today could create a combined streak > 6.
            forward = preplaced_working_days_after(ph.id, d)
            if backward + 1 + forward > 6:
                setv(ph.id, d, "O")   # prevent running into pre-placed block
                continue
            # Proactive staggering: at streak=5, mark as "can rest early"
            if backward >= 5:
                can_rest_early.append(ph)
            else:
                to_assign.append(ph)

        # Decide which streak-5 physicians work vs rest today.
        # Pull them in only if needed for coverage; rest the surplus.
        needed = len(SUBS) + len(TEAMS)  # 8 required morning slots
        shortfall = max(0, needed - len(to_assign))
        if shortfall > 0 and can_rest_early:
            # Sort by fewest duties first (fairest to keep working)
            can_rest_early.sort(key=lambda ph: (duty_cnt[ph.id], calc_h(ph.id), ph.id))
            to_assign.extend(can_rest_early[:shortfall])
            for ph in can_rest_early[shortfall:]:
                setv(ph.id, d, "O")  # proactive rest – lock preserved
        else:
            for ph in can_rest_early:
                setv(ph.id, d, "O")  # proactive rest – lock preserved

        # ── Step 2: separate locked vs free physicians ───────────────────────
        locked_today: List[tuple] = [
            (ph, lock_map[ph.id]) for ph in to_assign
            if lock_map[ph.id] is not None
        ]
        free_today: List[Doctor] = [
            ph for ph in to_assign if lock_map[ph.id] is None
        ]

        # Track which specialties are covered today (max 1 per SUB/TEAM).
        # Pre-seed from any physicians who already have a morning code today
        # (e.g. manually pre-assigned or stamped by an earlier phase).
        # This prevents Phase 7 from double-assigning a slot that's already filled.
        covered_today: Dict[str, int] = {}
        for _ph in docs:
            _code = get(_ph.id, d)
            if _code in MORNING_K:
                covered_today[_code] = _ph.id
                # For NENP: mark both NE and NP as covered
                if _code == "NENP":
                    covered_today["NE"] = _ph.id
                    covered_today["NP"] = _ph.id

        # ── Step 3: honor locked physicians ─────────────────────────────────
        # Locked physicians stay on their specialty.  Blocked specialties give
        # a day off (lock preserved).  Lock conflicts (specialty already taken)
        # move the physician into a "redirectable" pool so they can be used to
        # cover other uncovered required specialties in Step 4 instead of
        # wasting their availability on an unnecessary day off.
        redirectable: List[Doctor] = []
        for ph, lock in locked_today:
            if lock == "NENP":
                # Combined Neurology+Nephrology lock — handle all four cases.
                ne_blocked = is_spec_blocked("NE", d)
                np_blocked = is_spec_blocked("NP", d)
                ne_open = "NE" not in covered_today
                np_open = "NP" not in covered_today
                if ne_blocked and np_blocked:
                    setv(ph.id, d, "O")          # both blocked → rest, lock kept
                elif ne_open and np_open and not ne_blocked and not np_blocked:
                    setv(ph.id, d, "NENP")
                    covered_today["NE"] = ph.id
                    covered_today["NP"] = ph.id
                    spec_usage["NE"] += 1
                    spec_usage["NP"] += 1
                elif ne_open and not ne_blocked:
                    setv(ph.id, d, "NE")
                    lock_map[ph.id] = "NE"       # downgrade lock to single spec
                    covered_today["NE"] = ph.id
                    spec_usage["NE"] += 1
                elif np_open and not np_blocked:
                    setv(ph.id, d, "NP")
                    lock_map[ph.id] = "NP"
                    covered_today["NP"] = ph.id
                    spec_usage["NP"] += 1
                else:
                    redirectable.append(ph)      # both already covered
            elif is_spec_blocked(lock, d):
                setv(ph.id, d, "O")              # specialty blocked → rest, lock kept
            elif lock not in covered_today:
                setv(ph.id, d, lock)
                spec_usage[lock] += 1
                covered_today[lock] = ph.id
            else:
                # Lock conflict: specialty already covered → redirect to open slot.
                redirectable.append(ph)

        # ── Step 4: fill uncovered required specialties ───────────────────────
        # Pool = free physicians (no lock) + lock-conflict physicians (redirectable).
        # Priority order: SUBS first (each must have exactly 1 physician), then
        # TEAMS (each must have exactly 1 physician).
        # NENP ESCAPE: if the fill pool is smaller than the uncovered slot count,
        # and both NE and NP are uncovered, merge them into a single "NENP" slot
        # so one physician covers both.  This preserves coverage when the roster
        # is thin, without over-assigning anyone.
        required_order: List[str] = [
            s for s in (SUBS + TEAMS)
            if s not in covered_today and not is_spec_blocked(s, d)
        ]

        # Build combined fill pool, sorted by fewest duties → fewest hours → id.
        fill_pool: List[Doctor] = free_today + redirectable
        fill_pool.sort(key=lambda ph: (duty_cnt[ph.id], calc_h(ph.id), ph.id))

        # Activate NENP escape when supply < demand and both NE+NP are still open.
        if len(fill_pool) < len(required_order):
            if "NE" in required_order and "NP" in required_order:
                required_order.remove("NE")
                required_order.remove("NP")
                required_order.insert(0, "NENP")  # highest SUBS priority

        for spec in required_order:
            if not fill_pool:
                break   # no physicians left to fill this slot

            # Prefer a physician whose preferred specialty matches.
            # For NENP, accept either NE or NP preference.
            if spec == "NENP":
                best_idx = next(
                    (i for i, ph in enumerate(fill_pool)
                     if pref_spec.get(ph.id) in ("NE", "NP")),
                    None,
                )
            else:
                best_idx = next(
                    (i for i, ph in enumerate(fill_pool)
                     if pref_spec.get(ph.id) == spec),
                    None,
                )
            if best_idx is None:
                best_idx = 0   # take the fairness-best available physician

            ph = fill_pool.pop(best_idx)

            if spec == "NENP":
                # One physician covers both NE and NP.
                setv(ph.id, d, "NENP")
                lock_map[ph.id] = "NENP"
                spec_usage["NE"] += 1
                spec_usage["NP"] += 1
                covered_today["NE"] = ph.id
                covered_today["NP"] = ph.id
            else:
                setv(ph.id, d, spec)
                lock_map[ph.id] = spec      # (re-)lock physician to the assigned spec
                spec_usage[spec] += 1
                covered_today[spec] = ph.id

        # ── Step 5: overflow remaining physicians → day off ─────────────────
        # After all 8 required specialties (5 SUBS + 3 TEAMS) are covered
        # with exactly 1 physician each, remaining physicians get a rest day.
        # This prevents unnecessary overcrowding (e.g. 3 doctors in T1) and
        # ensures physicians accumulate rest days early, keeping them under
        # the 6-consecutive-day limit and 168-hour ceiling naturally rather
        # than hitting those limits all at once mid-month.
        for ph in fill_pool:
            setv(ph.id, d, "O")
            # lock_map[ph.id] is preserved — spec says lock persists through O days
            # and is only released on a DM/DF duty day.

    # ------------------------------------------------------------------
    # Phase 7.5 – Team-round rescue pass
    # After Phase 7's main loop some days may still lack T1, T2, or T3
    # coverage.  Team rounds are the highest clinical priority among
    # morning duties: they can't be substituted.  This pass finds the
    # first physician who is already working a SUBS specialty on that
    # day and redirects them to the uncovered TEAM slot.
    #
    # Rationale: redirecting from SUBS → TEAM does not change total
    # hours (both are 8 h shifts), so it can never push anyone over
    # the 168 h ceiling.  The SUBS slot is vacated (may be unassigned)
    # which is a lesser disruption than leaving a team round uncovered.
    # ------------------------------------------------------------------
    for _d in range(1, td + 1):
        for _team in ("T1", "T2", "T3"):
            if any(get(ph.id, _d) == _team for ph in docs):
                continue   # already covered
            # Prefer a physician whose pref_spec matches the team, then
            # fewest total hours (best available to spare), then id.
            _subs_pool = [
                ph for ph in docs
                if get(ph.id, _d) in SUBS          # currently on a sub-specialty
                and get(ph.id, _d) not in ("DM", "DF", "DC", "PC", "L", "R")
            ]
            if not _subs_pool:
                continue   # no SUBS physician available to redirect
            # Sort: prefer matching pref, then highest hours (they benefit
            # most from the swap since their specialty load stays the same)
            _subs_pool.sort(key=lambda ph: (
                0 if pref_spec.get(ph.id) == _team else 1,
                -calc_h(ph.id),
                ph.id,
            ))
            _rph = _subs_pool[0]
            setv(_rph.id, _d, _team)       # redirect to uncovered TEAM
            lock_map[_rph.id] = _team      # update lock for future days

    # ------------------------------------------------------------------
    # Phase 8 – Minimum hours enforcement (≥ 160 h per month)
    #
    # Hard rule: every physician must work at least 160 h per month.
    # The only legitimate reasons to fall below 160 h are:
    #   • Annual leave (L) – immovable, counts as zero working hours.
    #   • Random off day (R) – immovable, counts as zero working hours.
    #   • Manually assigned off day (O) that was pinned by the user.
    #
    # For any physician who is still below 160 h after Phase 7.5,
    # this phase converts plain "O" (day-off) slots — those not
    # hard-blocked by L/R or manually pinned — into 8-hour morning
    # shifts until the 160 h floor is reached or no further O days
    # are available.
    #
    # Constraints respected during conversion:
    #   • Hard 168 h ceiling is never exceeded.
    #   • 6-consecutive-day limit is never violated.
    #   • L, R, PC, and user-pinned days are never touched.
    # ------------------------------------------------------------------
    for ph in docs:
        if calc_h(ph.id) >= 160:
            continue   # already meets the minimum

        # Collect plain O days that are eligible for conversion:
        # not hard-blocked (L/R/FDD), not manually pinned.
        convertible = sorted(
            d for d in range(1, td + 1)
            if get(ph.id, d) == "O"
            and d not in hb[ph.id]
            and d not in pinned[ph.id]
        )

        for d in convertible:
            if calc_h(ph.id) >= 160:
                break   # floor reached

            # Respect 168 h hard ceiling
            if calc_h(ph.id) + 8 > 168:
                break

            # Respect 6-consecutive-day limit
            streak_before = consecutive_working_days_before(ph.id, d)
            streak_after  = preplaced_working_days_after(ph.id, d)
            if streak_before + 1 + streak_after > 6:
                continue   # converting this O would create a forbidden streak

            # Determine the best 8-hour code to assign.
            # Prefer the physician's locked/preferred specialty; fall back
            # to their IM team code.  If the preferred spec is blocked on
            # this day, use the team code instead.
            best_code = pref_spec.get(ph.id) or ph.team
            if best_code not in MORNING_K:
                best_code = ph.team
            if is_spec_blocked(best_code, d):
                best_code = ph.team if ph.team in MORNING_K else "T1"

            setv(ph.id, d, best_code)

    return {"a": a, "pairs": pairs}


# ─────────────────────────────────────────────────────────────────────────────
class ManualAssignDialog(tk.Toplevel):
    """Dialog for pinning specific duty/specialty assignments to specific days.

    A manual assignment is a hard pre-set: physician X does duty Y on day D.
    When Auto Schedule runs it treats these slots as already filled and works
    around them.  Before saving each entry the dialog validates:
      • The physician has no other assignment on that day.
      • Adding the duty would not create a streak > 6 consecutive working days.
      • Adding the duty would not push the physician past the 168 h ceiling.
    """

    # ── friendly label shown in the duty combobox ────────────────────────────
    _DUTY_LABELS: List[tuple] = [
        ("T1",   "T1 – Team 1 Morning"),
        ("T2",   "T2 – Team 2 Morning"),
        ("T3",   "T3 – Team 3 Morning"),
        ("PU",   "PUL – Pulmonology"),
        ("CAHM", "CA/HM – Cardiology / HM"),
        ("NE",   "NEU – Neurology"),
        ("NP",   "NEPH – Nephrology"),
        ("GI",   "GAS – Gastroenterology"),
        ("DC",   "DC – Daycare Clinic"),
        ("DM",   "DM – 16hr Duty (Male)"),
        ("DF",   "DF – 16hr Duty (Female)"),
    ]

    def __init__(self, parent: "MedSchedulerApp"):
        super().__init__(parent)
        self.parent = parent
        self.title("Manual Assignments")
        self.geometry("740x560")
        self.minsize(740, 560)
        self.transient(parent)
        self.grab_set()

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)

        # ── Add section ──────────────────────────────────────────────────────
        add_box = ttk.LabelFrame(frm, text="Add Manual Assignment", padding=10)
        add_box.pack(fill="x")

        # Row 0: labels
        for col, txt in enumerate(["Physician", "Duty / Specialty", "Day"]):
            ttk.Label(add_box, text=txt).grid(row=0, column=col, sticky="w", padx=6)

        # Row 1: widgets
        self.ph_var   = tk.StringVar()
        self.duty_var = tk.StringVar()
        self.day_var  = tk.IntVar(value=1)

        td = dim(parent.yr, parent.mo)

        ph_labels = [f"{d.initials or d.name[:4]}  –  {d.name}" for d in parent.docs]
        self.ph_cb = ttk.Combobox(add_box, textvariable=self.ph_var,
                                  values=ph_labels, state="readonly", width=28)
        if ph_labels:
            self.ph_cb.current(0)
        self.ph_cb.grid(row=1, column=0, padx=6, pady=4, sticky="w")

        duty_labels = [lbl for _, lbl in self._DUTY_LABELS]
        self.duty_cb = ttk.Combobox(add_box, textvariable=self.duty_var,
                                    values=duty_labels, state="readonly", width=30)
        self.duty_cb.current(0)
        self.duty_cb.grid(row=1, column=1, padx=6, pady=4, sticky="w")

        day_spin = ttk.Spinbox(add_box, from_=1, to=td,
                               textvariable=self.day_var, width=6)
        day_spin.grid(row=1, column=2, padx=6, pady=4, sticky="w")

        ttk.Button(add_box, text="Add Assignment",
                   command=self._add).grid(row=1, column=3, padx=12)

        # ── Warning label ─────────────────────────────────────────────────────
        self.warn_var = tk.StringVar()
        warn_lbl = ttk.Label(frm, textvariable=self.warn_var,
                             foreground="red", wraplength=700, justify="left")
        warn_lbl.pack(fill="x", pady=(4, 0))

        # ── Assignment list ──────────────────────────────────────────────────
        list_box = ttk.LabelFrame(frm, text="Current Manual Assignments", padding=8)
        list_box.pack(fill="both", expand=True, pady=(10, 0))

        cols = ("ph", "duty", "day", "del")
        self.tree = ttk.Treeview(list_box, columns=cols, show="headings", height=14)
        for c, w, h in [("ph", 200, "Physician"), ("duty", 220, "Duty/Specialty"),
                        ("day", 80, "Day"), ("del", 70, "")]:
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="center" if c in ("day", "del") else "w")
        sb = ttk.Scrollbar(list_box, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.tree.bind("<ButtonRelease-1>", self._on_click)

        # Close
        ttk.Button(frm, text="Close", command=self.destroy).pack(anchor="e", pady=(8, 0))

        self._refresh_list()

    # ── helpers ──────────────────────────────────────────────────────────────
    def _pid_from_selection(self) -> Optional[int]:
        """Return the physician id for the currently selected combobox entry."""
        idx = self.ph_cb.current()
        if idx < 0 or idx >= len(self.parent.docs):
            return None
        return self.parent.docs[idx].id

    def _code_from_selection(self) -> Optional[str]:
        idx = self.duty_cb.current()
        if idx < 0 or idx >= len(self._DUTY_LABELS):
            return None
        return self._DUTY_LABELS[idx][0]

    def _duty_label(self, code: str) -> str:
        for c, lbl in self._DUTY_LABELS:
            if c == code:
                return lbl
        return code

    def _ph_name(self, pid: int) -> str:
        d = next((p for p in self.parent.docs if p.id == pid), None)
        if d is None:
            return f"#{pid}"
        init = d.initials or d.name[:4]
        return f"{init}  –  {d.name}"

    # ── validation ────────────────────────────────────────────────────────────
    def _validate(self, pid: int, code: str, day: int) -> Optional[str]:
        """Return an error string if the assignment violates a hard rule, else None."""
        app = self.parent
        y, m = app.yr, app.mo
        td  = dim(y, m)

        # 1. Conflict: physician already has an assignment on this day
        existing = app.get(pid, day)
        if existing not in ("_", "O"):
            ph = next((p for p in app.docs if p.id == pid), None)
            name = ph.name if ph else f"#{pid}"
            return (f"{name} already has '{SHIFTS.get(existing,{}).get('short', existing)}'"
                    f" assigned on day {day}.")

        # 2. Consecutive-day streak check
        OFF = {"O", "L", "R", "PC", "_"}

        def streak_before(d: int) -> int:
            s, x = 0, d - 1
            while x >= 1:
                c = app.get(pid, x)
                if c in OFF:
                    break
                s += 1
                x -= 1
            return s

        def streak_after(d: int) -> int:
            s, x = 0, d + 1
            while x <= td:
                c = app.get(pid, x)
                if c in OFF:
                    break
                s += 1
                x += 1
            return s

        back    = streak_before(day)
        forward = streak_after(day)
        total   = back + 1 + forward
        if total > 6:
            return (f"Adding this duty creates a streak of {total} consecutive working days "
                    f"(max 6). Please add an off day nearby first.")

        # 3. Hours ceiling
        h_code  = SHIFTS.get(code, {}).get("h", 8)
        h_total = sum(SHIFTS.get(app.get(pid, d), {"h": 0})["h"] for d in range(1, td + 1))
        if h_total + h_code > 168:
            return (f"Adding this duty ({h_code}h) would push the physician's total to "
                    f"{h_total + h_code}h, exceeding the 168h monthly ceiling.")

        return None   # all clear

    # ── add / delete ──────────────────────────────────────────────────────────
    def _add(self):
        self.warn_var.set("")
        pid  = self._pid_from_selection()
        code = self._code_from_selection()
        try:
            day = int(self.day_var.get())
        except (ValueError, tk.TclError):
            self.warn_var.set("Invalid day number.")
            return

        if pid is None or code is None:
            self.warn_var.set("Select a physician and a duty.")
            return

        td = dim(self.parent.yr, self.parent.mo)
        if not 1 <= day <= td:
            self.warn_var.set(f"Day must be between 1 and {td}.")
            return

        err = self._validate(pid, code, day)
        if err:
            self.warn_var.set(err)
            return

        # Save: stamp grid + store ManualAssignment record
        self.parent.setv(pid, day, code)
        ma = ManualAssignment(
            id=self.parent.next_manual_id,
            pid=pid,
            code=code,
            day=day,
        )
        self.parent.next_manual_id += 1
        self.parent.manual_asgns.append(ma)
        self.parent.refresh_all()
        self._refresh_list()

    def _on_click(self, event):
        """Delete row when the 'Delete' column cell is clicked."""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = self.tree.identify_column(event.x)
        if col != "#4":   # 'del' is the 4th column
            return
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        ma_id = int(iid)
        self._delete(ma_id)

    def _delete(self, ma_id: int):
        ma = next((x for x in self.parent.manual_asgns if x.id == ma_id), None)
        if ma is None:
            return
        # Erase from grid only if it still matches what we placed
        current = self.parent.get(ma.pid, ma.day)
        if current == ma.code:
            self.parent.setv(ma.pid, ma.day, "O")
        self.parent.manual_asgns = [x for x in self.parent.manual_asgns if x.id != ma_id]
        self.parent.refresh_all()
        self._refresh_list()

    def _refresh_list(self):
        self.tree.delete(*self.tree.get_children())
        for ma in sorted(self.parent.manual_asgns, key=lambda x: (x.day, x.pid)):
            dow = DN[day_of_week(self.parent.yr, self.parent.mo, ma.day)]
            day_str_lbl = f"Day {ma.day} ({dow})"
            self.tree.insert("", "end", iid=str(ma.id),
                             values=(self._ph_name(ma.pid),
                                     self._duty_label(ma.code),
                                     day_str_lbl,
                                     "🗑 Delete"))


class LeaveDialog(tk.Toplevel):
    def __init__(self, parent: "MedSchedulerApp"):
        super().__init__(parent)
        self.parent = parent
        self.title("Annual Leave")
        self.geometry("700x520")
        self.minsize(700, 520)
        self.transient(parent)
        self.grab_set()

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)

        top = ttk.LabelFrame(frm, text="Add Leave Block", padding=10)
        top.pack(fill="x")
        ttk.Label(top, text="Physician").grid(row=0, column=0, sticky="w")
        self.doc_var = tk.StringVar()
        self.doc_combo = ttk.Combobox(top, textvariable=self.doc_var, state="readonly", width=35)
        self.doc_combo["values"] = [f"{d.id}: {d.name}" for d in parent.docs]
        if parent.docs:
            self.doc_combo.current(0)
        self.doc_combo.grid(row=1, column=0, padx=4, pady=4)
        ttk.Label(top, text="From (YYYY-MM-DD)").grid(row=0, column=1, sticky="w")
        self.from_var = tk.StringVar(value=ds(parent.yr, parent.mo, 1))
        ttk.Entry(top, textvariable=self.from_var, width=15).grid(row=1, column=1, padx=4, pady=4)
        ttk.Label(top, text="To (YYYY-MM-DD)").grid(row=0, column=2, sticky="w")
        self.to_var = tk.StringVar(value=ds(parent.yr, parent.mo, dim(parent.yr, parent.mo)))
        ttk.Entry(top, textvariable=self.to_var, width=15).grid(row=1, column=2, padx=4, pady=4)
        ttk.Button(top, text="Add", command=self.add_leave).grid(row=1, column=3, padx=6)

        listbox_frame = ttk.LabelFrame(frm, text="Current Leave Blocks", padding=10)
        listbox_frame.pack(fill="both", expand=True, pady=(12, 0))
        cols = ("id", "doctor", "from", "to", "delete")
        self.tree = ttk.Treeview(listbox_frame, columns=cols, show="headings", height=12)
        for c, w, h in [("id", 60, "Id"), ("doctor", 240, "Doctor"), ("from", 120, "From"), ("to", 120, "To"), ("delete", 44, "")]:
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="center", stretch=(c != "delete"))
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Button-1>", self.on_tree_click)
        btns = ttk.Frame(frm)
        btns.pack(fill="x", pady=(10, 0))
        ttk.Button(btns, text="Close", command=self.destroy).pack(side="right")
        self.refresh()

    def refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        ms = ds(self.parent.yr, self.parent.mo, 1)
        me = ds(self.parent.yr, self.parent.mo, dim(self.parent.yr, self.parent.mo))
        for b in self.parent.leaves:
            if b.f <= me and b.t >= ms:
                doctor = next((d.name for d in self.parent.docs if d.id == b.pid), "?")
                self.tree.insert("", "end", values=(b.id, doctor, b.f, b.t, "🗑"))

    def add_leave(self):
        if not self.doc_var.get():
            return
        pid = int(self.doc_var.get().split(":", 1)[0])
        f = self.from_var.get().strip()
        t = self.to_var.get().strip()
        try:
            datetime.strptime(f, "%Y-%m-%d")
            datetime.strptime(t, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Invalid date", "Use YYYY-MM-DD format.", parent=self)
            return
        if f > t:
            messagebox.showerror("Invalid range", "'From' must be on or before 'To'.", parent=self)
            return
        self.parent.add_leave(pid, f, t)
        self.refresh()

    def on_tree_click(self, event):
        item_id = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if item_id and column == "#5":
            row = self.tree.item(item_id, "values")
            self.parent.delete_leave(int(row[0]))
            self.refresh()


class SpecialtyBlockDialog(tk.Toplevel):
    def __init__(self, parent: "MedSchedulerApp"):
        super().__init__(parent)
        self.parent = parent
        self.title("Blocked Specialties")
        self.geometry("720x430")
        self.transient(parent)
        self.grab_set()

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)
        top = ttk.LabelFrame(frm, text="Block Specialty", padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="Specialty").grid(row=0, column=0, sticky="w")
        self.spec_var = tk.StringVar()
        self.spec_combo = ttk.Combobox(top, textvariable=self.spec_var, state="readonly", width=30)
        self.spec_combo["values"] = [f"{code}: {SHIFTS[code]['label']}" for code in BLOCKABLE_SPECIALTIES]
        self.spec_combo.current(0)
        self.spec_combo.grid(row=1, column=0, padx=4, pady=4)

        ttk.Label(top, text="From (YYYY-MM-DD)").grid(row=0, column=1, sticky="w")
        self.from_var = tk.StringVar(value=ds(parent.yr, parent.mo, 1))
        ttk.Entry(top, textvariable=self.from_var, width=15).grid(row=1, column=1, padx=4, pady=4)
        ttk.Label(top, text="To (YYYY-MM-DD)").grid(row=0, column=2, sticky="w")
        self.to_var = tk.StringVar(value=ds(parent.yr, parent.mo, dim(parent.yr, parent.mo)))
        ttk.Entry(top, textvariable=self.to_var, width=15).grid(row=1, column=2, padx=4, pady=4)
        ttk.Button(top, text="Add", command=self.add_block).grid(row=1, column=3, padx=6)

        treef = ttk.LabelFrame(frm, text="Current Specialty Blocks", padding=10)
        treef.pack(fill="both", expand=True, pady=(12, 0))
        cols = ("id", "specialty", "from", "to", "delete")
        self.tree = ttk.Treeview(treef, columns=cols, show="headings", height=12)
        for c, w, h in [("id", 60, "Id"), ("specialty", 260, "Specialty"), ("from", 120, "From"), ("to", 120, "To"), ("delete", 44, "")]:
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="center", stretch=(c != "delete"))
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Button-1>", self.on_tree_click)
        btns = ttk.Frame(frm)
        btns.pack(fill="x", pady=(10, 0))
        ttk.Button(btns, text="Close", command=self.destroy).pack(side="right")
        self.refresh()

    def refresh(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        ms = ds(self.parent.yr, self.parent.mo, 1)
        me = ds(self.parent.yr, self.parent.mo, dim(self.parent.yr, self.parent.mo))
        for b in self.parent.spec_blocks:
            if b.f <= me and b.t >= ms:
                self.tree.insert("", "end", values=(b.id, f"{b.code}: {SHIFTS[b.code]['label']}", b.f, b.t, "🗑"))

    def add_block(self):
        spec = self.spec_var.get().strip()
        if not spec:
            return
        code = spec.split(":", 1)[0]
        f = self.from_var.get().strip()
        t = self.to_var.get().strip()
        try:
            datetime.strptime(f, "%Y-%m-%d")
            datetime.strptime(t, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Invalid date", "Use YYYY-MM-DD format.", parent=self)
            return
        if f > t:
            messagebox.showerror("Invalid range", "'From' must be on or before 'To'.", parent=self)
            return
        self.parent.add_spec_block(code, f, t)
        self.refresh()

    def on_tree_click(self, event):
        item_id = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if item_id and column == "#5":
            row = self.tree.item(item_id, "values")
            self.parent.delete_spec_block(int(row[0]))
            self.refresh()


class QuickAssignDialog(tk.Toplevel):
    """Assign any duty to any physician on any day — all three in one dialog.

    Can be opened from:
      • Clicking a physician name or day cell in the schedule table
        (physician and day are pre-selected automatically).
      • The "Assign Duty" toolbar button (starts with physician 1, day 1).

    Workflow:
      1. Pick physician from the dropdown.
      2. Pick a day — the weekday name and current assignment update live.
      3. Pick the duty from the dropdown (auto-filled to current assignment).
      4. Click Apply — saved immediately, dialog stays open for more edits.
      5. Click Close when done.
    """

    _OPTIONS: List[tuple] = [
        ("_",    "— Blank / Clear"),
        ("O",    "O   – Day Off"),
        ("L",    "L   – Annual Leave"),
        ("R",    "R   – Random Off"),
        ("PC",   "PC  – Post-Call Off"),
        ("T1",   "T1  – Team 1 Morning"),
        ("T2",   "T2  – Team 2 Morning"),
        ("T3",   "T3  – Team 3 Morning"),
        ("CAHM", "CAHM – Cardiology / Hematology"),
        ("GI",   "GI  – Gastroenterology"),
        ("NE",   "NE  – Neurology"),
        ("NP",   "NP  – Nephrology"),
        ("NENP", "NENP – Neurology + Nephrology"),
        ("PU",   "PU  – Pulmonology"),
        ("DC",   "DC  – Daycare Clinic"),
        ("DM",   "DM  – 16hr Duty (Male)"),
        ("DF",   "DF  – 16hr Duty (Female)"),
    ]

    def __init__(self, parent: "MedSchedulerApp", pid: int = 0, day: int = 1):
        super().__init__(parent)
        self.parent = parent
        self._td    = dim(parent.yr, parent.mo)

        # Fall back to first physician if pid not given or invalid
        if not parent.docs:
            self.destroy()
            return
        ph_ids = [ph.id for ph in parent.docs]
        if pid not in ph_ids:
            pid = parent.docs[0].id
        self._pid = pid

        self.title(f"Assign Duty — {MONTHS[parent.mo]} {parent.yr}")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        outer = ttk.Frame(self, padding=20)
        outer.pack(fill="both", expand=True)
        outer.columnconfigure(1, weight=1)

        # ── Title ────────────────────────────────────────────────────────
        ttk.Label(outer, text="Assign Duty",
                  font=("TkDefaultFont", 12, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 14))

        # ── Physician picker ─────────────────────────────────────────────
        ttk.Label(outer, text="Physician:").grid(
            row=1, column=0, sticky="w", pady=4)

        self._ph_var = tk.StringVar()
        ph_labels    = [f"{ph.name}  ({ph.team})" for ph in parent.docs]
        self._ph_cb  = ttk.Combobox(outer, textvariable=self._ph_var,
                                    values=ph_labels, state="readonly", width=36)
        # Pre-select the clicked physician
        ph_idx = next((i for i, ph in enumerate(parent.docs) if ph.id == pid), 0)
        self._ph_cb.current(ph_idx)
        self._ph_cb.grid(row=1, column=1, sticky="ew", padx=(8, 0), pady=4)
        self._ph_cb.bind("<<ComboboxSelected>>", lambda e: self._on_ph_change())

        # ── Day picker ───────────────────────────────────────────────────
        ttk.Label(outer, text="Day:").grid(
            row=2, column=0, sticky="w", pady=4)

        day_frame = ttk.Frame(outer)
        day_frame.grid(row=2, column=1, sticky="w", padx=(8, 0), pady=4)

        self._day_var = tk.IntVar(value=max(1, min(day, self._td)))
        self._day_spin = ttk.Spinbox(day_frame, from_=1, to=self._td,
                                     textvariable=self._day_var, width=5,
                                     command=self._on_day_change)
        self._day_spin.pack(side="left")
        self._day_spin.bind("<Return>", lambda e: self._on_day_change())
        self._day_spin.bind("<FocusOut>", lambda e: self._on_day_change())

        self._dow_lbl = ttk.Label(day_frame, text="", foreground="gray", width=16)
        self._dow_lbl.pack(side="left", padx=(10, 0))

        # ── Current assignment (live) ────────────────────────────────────
        ttk.Label(outer, text="Current:").grid(
            row=3, column=0, sticky="w", pady=(2, 6))
        self._cur_lbl = ttk.Label(outer, text="", foreground="#1a56c4",
                                  anchor="w", width=40)
        self._cur_lbl.grid(row=3, column=1, sticky="w", padx=(8, 0), pady=(2, 6))

        # ── Duty picker ──────────────────────────────────────────────────
        ttk.Separator(outer, orient="horizontal").grid(
            row=4, column=0, columnspan=2, sticky="ew", pady=(4, 10))

        ttk.Label(outer, text="Assign duty:").grid(
            row=5, column=0, sticky="w", pady=4)

        self._duty_var = tk.StringVar()
        self._duty_cb  = ttk.Combobox(outer, textvariable=self._duty_var,
                                      values=[lbl for _, lbl in self._OPTIONS],
                                      state="readonly", width=36)
        self._duty_cb.current(0)
        self._duty_cb.grid(row=5, column=1, sticky="ew", padx=(8, 0), pady=4)

        # ── Status line (shows result of last Apply) ─────────────────────
        self._status_var = tk.StringVar(value="")
        ttk.Label(outer, textvariable=self._status_var,
                  foreground="green", anchor="w").grid(
            row=6, column=0, columnspan=2, sticky="w", pady=(6, 0))

        # ── Buttons ──────────────────────────────────────────────────────
        ttk.Separator(outer, orient="horizontal").grid(
            row=7, column=0, columnspan=2, sticky="ew", pady=(10, 8))

        btn_row = ttk.Frame(outer)
        btn_row.grid(row=8, column=0, columnspan=2, sticky="e")
        ttk.Button(btn_row, text="Close",
                   command=self.destroy).pack(side="right", padx=(8, 0))
        ttk.Button(btn_row, text="Apply",
                   command=self._apply).pack(side="right")

        # Initialise live labels
        self._on_day_change()

        # Centre on parent
        self.update_idletasks()
        px = parent.winfo_rootx() + parent.winfo_width()  // 2 - self.winfo_width()  // 2
        py = parent.winfo_rooty() + parent.winfo_height() // 2 - self.winfo_height() // 2
        self.geometry(f"+{max(0, px)}+{max(0, py)}")

    # ── internal helpers ─────────────────────────────────────────────────
    def _selected_pid(self) -> int:
        idx = self._ph_cb.current()
        if 0 <= idx < len(self.parent.docs):
            return self.parent.docs[idx].id
        return self._pid

    def _current_day(self) -> int:
        try:
            d = int(self._day_var.get())
            return max(1, min(d, self._td))
        except (ValueError, tk.TclError):
            return 1

    def _on_ph_change(self):
        """When physician selection changes, refresh the current-assignment display."""
        self._pid = self._selected_pid()
        self._on_day_change()

    def _on_day_change(self):
        """Refresh weekday label and current-assignment label whenever day changes."""
        d    = self._current_day()
        pid  = self._selected_pid()
        dow  = DN[day_of_week(self.parent.yr, self.parent.mo, d)]
        code = self.parent.get(pid, d)
        lbl  = SHIFTS.get(code, {}).get("label", "—") if code != "_" else "Blank"
        self._dow_lbl.config(text=f"({dow})")
        self._cur_lbl.config(text=f"{code}  —  {lbl}")
        # Auto-select the duty combobox to match current assignment
        idx = next((i for i, (c, _) in enumerate(self._OPTIONS) if c == code), 0)
        self._duty_cb.current(idx)
        self._status_var.set("")

    def _apply(self):
        duty_idx = self._duty_cb.current()
        if duty_idx < 0:
            return
        code = self._OPTIONS[duty_idx][0]
        day  = self._current_day()
        pid  = self._selected_pid()
        ph   = next((p for p in self.parent.docs if p.id == pid), None)
        self.parent.setv(pid, day, code)
        self.parent.refresh_all()
        # Refresh live labels and show confirmation
        self._on_day_change()
        ph_name = ph.name if ph else f"#{pid}"
        dow     = DN[day_of_week(self.parent.yr, self.parent.mo, day)]
        self._status_var.config(foreground="green") if hasattr(self._status_var, "config") else None
        self._status_var.set(f"✓  {ph_name}  ·  Day {day} ({dow})  →  {code}")


# ── Firebase dialogs ─────────────────────────────────────────────────────────

class LoginDialog(tk.Toplevel):
    """Sign-in / sign-up dialog for Firebase Authentication."""

    def __init__(self, parent: "MedSchedulerApp"):
        super().__init__(parent)
        self.parent   = parent
        self.title("MedScheduler Cloud — Sign In")
        self.geometry("400x320")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        frm = ttk.Frame(self, padding=20)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="☁  MedScheduler Cloud",
                  font=("Segoe UI", 13, "bold")).pack(pady=(0, 14))

        # Email
        ttk.Label(frm, text="Email address").pack(anchor="w")
        self._email_var = tk.StringVar()
        email_e = ttk.Entry(frm, textvariable=self._email_var, width=38)
        email_e.pack(fill="x", pady=(2, 8))
        email_e.focus_set()

        # Password
        ttk.Label(frm, text="Password").pack(anchor="w")
        self._pw_var = tk.StringVar()
        ttk.Entry(frm, textvariable=self._pw_var, show="•", width=38).pack(fill="x", pady=(2, 12))

        # Error label
        self._err_var = tk.StringVar()
        self._err_lbl = ttk.Label(frm, textvariable=self._err_var,
                                  foreground="red", wraplength=360)
        self._err_lbl.pack(pady=(0, 8))

        # Buttons
        btn_row = ttk.Frame(frm)
        btn_row.pack(fill="x")
        ttk.Button(btn_row, text="Sign In",
                   command=lambda: self._auth(sign_up=False)).pack(side="left", padx=(0, 6))
        ttk.Button(btn_row, text="Create Account",
                   command=lambda: self._auth(sign_up=True)).pack(side="left")
        ttk.Button(btn_row, text="Cancel",
                   command=self.destroy).pack(side="right")

        self.bind("<Return>", lambda e: self._auth(sign_up=False))

    def _auth(self, sign_up: bool) -> None:
        email    = self._email_var.get().strip()
        password = self._pw_var.get()
        if not email or not password:
            self._err_var.set("Please enter both email and password.")
            return
        self._err_var.set("Connecting…")
        self.update_idletasks()
        try:
            if sign_up:
                firebase.sign_up(email, password)
            else:
                firebase.sign_in(email, password)
            self.parent._update_firebase_label()
            self.destroy()
        except FirebaseAuthError as exc:
            self._err_var.set(str(exc))
        except Exception as exc:
            self._err_var.set(f"Error: {exc}")


class CloudFilesDialog(tk.Toplevel):
    """Browse, upload and download exported files stored in Firebase Storage."""

    def __init__(self, parent: "MedSchedulerApp"):
        super().__init__(parent)
        self.parent = parent
        self.title("Cloud Files")
        self.geometry("600x420")
        self.minsize(500, 360)
        self.transient(parent)
        self.grab_set()

        frm = ttk.Frame(self, padding=14)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="📁  Cloud Exported Files",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 8))

        # File list treeview
        cols = ("name", "size", "date")
        self._tree = ttk.Treeview(frm, columns=cols, show="headings", height=12)
        self._tree.heading("name", text="File name")
        self._tree.heading("size", text="Size (KB)")
        self._tree.heading("date", text="Uploaded")
        self._tree.column("name", width=300, anchor="w")
        self._tree.column("size", width=80,  anchor="center")
        self._tree.column("date", width=100, anchor="center")
        sb = ttk.Scrollbar(frm, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="left", fill="y")

        # Status / action panel
        side = ttk.Frame(frm, padding=(10, 0, 0, 0))
        side.pack(side="left", fill="y")

        self._status_var = tk.StringVar(value="")
        ttk.Label(side, textvariable=self._status_var,
                  wraplength=140, justify="left").pack(anchor="w", pady=(0, 10))

        ttk.Button(side, text="⬇  Download",
                   command=self._download, width=16).pack(fill="x", pady=2)
        ttk.Button(side, text="⬆  Upload File",
                   command=self._upload, width=16).pack(fill="x", pady=2)
        ttk.Button(side, text="🔄  Refresh",
                   command=self._refresh, width=16).pack(fill="x", pady=2)
        ttk.Button(side, text="Close",
                   command=self.destroy, width=16).pack(fill="x", pady=(14, 0))

        self._refresh()

    def _refresh(self) -> None:
        self._status_var.set("Loading…")
        self.update_idletasks()
        try:
            files = firebase.list_files()
        except Exception as exc:
            self._status_var.set(f"Error: {exc}")
            return
        for row in self._tree.get_children():
            self._tree.delete(row)
        for f in files:
            self._tree.insert("", "end",
                              values=(f["name"], f["size_kb"], f["updated"]))
        self._status_var.set(f"{len(files)} file(s)" if files else "No files yet.")

    def _selected_name(self) -> Optional[str]:
        sel = self._tree.selection()
        if not sel:
            return None
        return self._tree.item(sel[0])["values"][0]

    def _download(self) -> None:
        name = self._selected_name()
        if not name:
            messagebox.showinfo("Select File", "Please select a file to download.",
                                parent=self)
            return
        path = filedialog.asksaveasfilename(
            title="Save file as", initialfile=name,
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            parent=self,
        )
        if not path:
            return
        self._status_var.set("Downloading…")
        self.update_idletasks()

        def _do():
            try:
                firebase.download_file(name, path)
                self.after(0, lambda: self._status_var.set(f"Downloaded ✓\n{name}"))
            except Exception as exc:
                self.after(0, lambda: self._status_var.set(f"Error: {exc}"))

        threading.Thread(target=_do, daemon=True).start()

    def _upload(self) -> None:
        path = filedialog.askopenfilename(
            title="Select file to upload",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            parent=self,
        )
        if not path:
            return
        self._status_var.set("Uploading…")
        self.update_idletasks()

        def _do():
            try:
                firebase.upload_file(path)
                self.after(0, lambda: (
                    self._status_var.set("Uploaded ✓"),
                    self._refresh(),
                ))
            except Exception as exc:
                self.after(0, lambda: self._status_var.set(f"Upload failed:\n{exc}"))

        threading.Thread(target=_do, daemon=True).start()


# ── Main application ─────────────────────────────────────────────────────────

class MedSchedulerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("MedScheduler - Python Desktop Edition")
        self.geometry("1520x860")
        self.minsize(1120, 680)
        today = date.today()
        self.yr = today.year
        self.mo = today.month - 1
        self.next_doc_id = 7
        self.next_leave_id = 1
        self.next_spec_block_id = 1
        self.next_manual_id = 1

        self.docs: List[Doctor] = [
            Doctor(1, "Dr. Ahmed Al-Rashid", "Internal Medicine", "T1"),
            Doctor(2, "Dr. Sara Hassan", "Internal Medicine", "T1"),
            Doctor(3, "Dr. Khalid Al-Mansour", "Internal Medicine", "T2"),
            Doctor(4, "Dr. Layla Mahmoud", "Internal Medicine", "T2"),
            Doctor(5, "Dr. Omar Al-Farsi", "Internal Medicine", "T3"),
            Doctor(6, "Dr. Nour Al-Sayed", "Internal Medicine", "T3"),
        ]
        self.asgn: Dict[str, str] = {}
        self.leaves: List[LeaveBlock] = []
        self.spec_blocks: List[SpecialtyBlock] = []
        self.manual_asgns: List[ManualAssignment] = []   # user-pinned assignments
        self.status_var = tk.StringVar(value="Ready")
        self.build_ui()
        self.refresh_all()
        self._update_firebase_label()

    def ak(self, pid: int, d: int) -> str:
        return f"{pid}|{self.yr}|{self.mo}|{d}"

    def get(self, pid: int, d: int) -> str:
        return self.asgn.get(self.ak(pid, d), "_")

    def setv(self, pid: int, d: int, value: str) -> None:
        self.asgn[self.ak(pid, d)] = value

    def build_ui(self):
        # Status bar — packed first so it anchors to the bottom correctly
        ttk.Label(self, textvariable=self.status_var,
                  relief="sunken", anchor="w",
                  padding=(8, 3)).pack(fill="x", side="bottom")

        # ── Toolbar row 1: title · month nav · cloud panel ───────────────
        tb1 = ttk.Frame(self, padding=(10, 5, 10, 2))
        tb1.pack(fill="x", side="top")

        # Firebase cloud panel — packed RIGHT first so it stays right-aligned
        fb_frame = ttk.Frame(tb1)
        fb_frame.pack(side="right")
        self._fb_user_var = tk.StringVar(value="☁  Not signed in")
        self._fb_user_lbl = ttk.Label(
            fb_frame, textvariable=self._fb_user_var,
            foreground="#777", cursor="hand2",
        )
        self._fb_user_lbl.pack(side="left", padx=(0, 6))
        self._fb_user_lbl.bind("<Button-1>", lambda e: self._firebase_login())
        ttk.Button(fb_frame, text="💾 Save",  command=self._firebase_save,  width=8).pack(side="left", padx=1)
        ttk.Button(fb_frame, text="⬇ Load",   command=self._firebase_load,  width=8).pack(side="left", padx=1)
        ttk.Button(fb_frame, text="📁 Files",  command=self._firebase_files, width=8).pack(side="left", padx=1)

        # App title
        ttk.Label(tb1, text="MedScheduler",
                  font=("Segoe UI", 14, "bold")).pack(side="left", padx=(0, 14))

        # Month navigation
        nav = ttk.Frame(tb1)
        nav.pack(side="left")
        ttk.Button(nav, text="◀", width=3,
                   command=lambda: self.nav_month(-1)).pack(side="left")
        self.month_lbl = ttk.Label(nav, text="", width=14, anchor="center",
                                   font=("Segoe UI", 11, "bold"))
        self.month_lbl.pack(side="left", padx=6)
        ttk.Button(nav, text="▶", width=3,
                   command=lambda: self.nav_month(1)).pack(side="left")

        ttk.Separator(self, orient="horizontal").pack(fill="x", side="top")

        # ── Toolbar row 2: all action buttons ────────────────────────────
        tb2 = ttk.Frame(self, padding=(10, 3, 10, 3))
        tb2.pack(fill="x", side="top")

        for txt, cmd in [("Auto-Schedule",   self.schedule),
                         ("Assign Duty",     self.open_quick_assign_dialog),
                         ("Annual Leave",    self.open_leave_dialog),
                         ("Block Specialty", self.open_spec_block_dialog),
                         ("Manual Assign",   self.open_manual_assign_dialog)]:
            ttk.Button(tb2, text=txt, command=cmd).pack(side="left", padx=2)

        ttk.Separator(tb2, orient="vertical").pack(
            side="left", fill="y", padx=8, pady=3)

        ttk.Button(tb2, text="Export Rota",
                   command=self.export_simplified_xlsx).pack(side="left", padx=2)
        ttk.Button(tb2, text="Export Full",
                   command=self.export_xlsx).pack(side="left", padx=2)

        ttk.Separator(tb2, orient="vertical").pack(
            side="left", fill="y", padx=8, pady=3)

        ttk.Button(tb2, text="Clear Month",
                   command=self.clear_month).pack(side="left", padx=2)

        ttk.Separator(self, orient="horizontal").pack(fill="x", side="top")

        # ── Body: sidebar + content ──────────────────────────────────────
        body = ttk.PanedWindow(self, orient="horizontal")
        body.pack(fill="both", expand=True, side="top")

        sidebar_host = ttk.Frame(body)
        body.add(sidebar_host, weight=0)
        self._build_sidebar(sidebar_host)

        content_host = ttk.Frame(body)
        body.add(content_host, weight=1)

        self.notebook = ttk.Notebook(content_host)
        self.notebook.pack(fill="both", expand=True)
        self.schedule_tab = ttk.Frame(self.notebook)
        self.summary_tab  = ttk.Frame(self.notebook)
        self.notebook.add(self.schedule_tab, text="  Schedule  ")
        self.notebook.add(self.summary_tab,  text="  Summary  ")
        self.build_schedule_tab()
        self.build_summary_tab()

    def _build_sidebar(self, parent):
        """Left sidebar: physician add form + scrollable list + specialty changer."""
        parent.configure(width=260)
        parent.pack_propagate(False)

        frm = ttk.Frame(parent, padding=(8, 8, 4, 8))
        frm.pack(fill="both", expand=True)

        # ── Add physician form ────────────────────────────────────────
        add_box = ttk.LabelFrame(frm, text="Add Physician", padding=8)
        add_box.pack(fill="x")

        self.name_var     = tk.StringVar()
        self.initials_var = tk.StringVar()
        self.spec_var     = tk.StringVar(value="Not specified")

        ttk.Label(add_box, text="Full name").pack(anchor="w")
        name_e = ttk.Entry(add_box, textvariable=self.name_var)
        name_e.pack(fill="x", pady=(1, 5))
        name_e.bind("<Return>", lambda e: self.add_doc())

        inits_row = ttk.Frame(add_box)
        inits_row.pack(fill="x", pady=(0, 5))
        ttk.Label(inits_row, text="Initials", width=8, anchor="w").pack(side="left")
        ttk.Entry(inits_row, textvariable=self.initials_var, width=7).pack(side="left")

        ttk.Label(add_box, text="Specialty").pack(anchor="w")
        self.spec_combo = ttk.Combobox(add_box, textvariable=self.spec_var,
                                       values=SPEC_OPTIONS, state="readonly")
        self.spec_combo.pack(fill="x", pady=(1, 6))

        b_row = ttk.Frame(add_box)
        b_row.pack(fill="x")
        ttk.Button(b_row, text="Add Physician",
                   command=self.add_doc).pack(side="left", fill="x",
                                              expand=True, padx=(0, 3))
        ttk.Button(b_row, text="Import",
                   command=self.import_doctors).pack(side="left")

        # ── Physician list ────────────────────────────────────────────
        ttk.Label(frm, text="Physicians",
                  font=("TkDefaultFont", 9, "bold")).pack(anchor="w", pady=(10, 2))

        list_wrap = ttk.Frame(frm)
        list_wrap.pack(fill="both", expand=True)

        self._doc_list_canvas = tk.Canvas(list_wrap, highlightthickness=0)
        _sb = ttk.Scrollbar(list_wrap, orient="vertical",
                            command=self._doc_list_canvas.yview)
        self.doc_list_frame = ttk.Frame(self._doc_list_canvas)
        self.doc_list_frame.bind(
            "<Configure>",
            lambda e: self._doc_list_canvas.configure(
                scrollregion=self._doc_list_canvas.bbox("all")))
        self._doc_list_canvas.create_window(
            (0, 0), window=self.doc_list_frame, anchor="nw")
        self._doc_list_canvas.configure(yscrollcommand=_sb.set)
        self._doc_list_canvas.pack(side="left", fill="both", expand=True)
        _sb.pack(side="right", fill="y")
        self._doc_list_canvas.bind(
            "<MouseWheel>",
            lambda e: self._doc_list_canvas.yview_scroll(
                int(-1 * (e.delta / 120)), "units"))
        self._doc_list_canvas.bind(
            "<Map>",
            lambda e: self._doc_list_canvas.configure(
                bg=ttk.Style().lookup("TFrame", "background")))

        self._doc_check_vars:    dict = {}
        self._doc_initials_vars: dict = {}
        self._doc_fdd_vars:      dict = {}

        # ── Bulk actions ──────────────────────────────────────────────
        a_frm = ttk.Frame(frm)
        a_frm.pack(fill="x", pady=(4, 0))
        ttk.Button(a_frm, text="Remove Selected",
                   command=self.remove_selected_docs).pack(fill="x", pady=(0, 2))
        sel_row = ttk.Frame(a_frm)
        sel_row.pack(fill="x")
        ttk.Button(sel_row, text="Select All",
                   command=self._select_all_docs).pack(side="left", fill="x",
                                                       expand=True, padx=(0, 2))
        ttk.Button(sel_row, text="Deselect All",
                   command=self._deselect_all_docs).pack(side="left", fill="x",
                                                         expand=True)

        # ── Specialty changer ─────────────────────────────────────────
        spec_box = ttk.LabelFrame(frm, text="Change Specialty", padding=8)
        spec_box.pack(fill="x", pady=(8, 0))

        self.assign_doc_var  = tk.StringVar()
        self.assign_spec_var = tk.StringVar(value="Not specified")

        self.assign_doc_combo = ttk.Combobox(spec_box,
                                             textvariable=self.assign_doc_var,
                                             state="readonly")
        self.assign_doc_combo.pack(fill="x", pady=(0, 4))

        self.assign_spec_combo = ttk.Combobox(spec_box,
                                              textvariable=self.assign_spec_var,
                                              values=SPEC_OPTIONS,
                                              state="readonly")
        self.assign_spec_combo.pack(fill="x", pady=(0, 6))

        ttk.Button(spec_box, text="Apply Specialty",
                   command=self.assign_specialty_to_doc).pack(fill="x")

    def build_schedule_tab(self):
        """Canvas-based colour-coded schedule grid.

        Layout
        ──────
          ┌──[_corner_cv]──┬──[_hdr_cv scrolls H]──┐
          │   PHYSICIAN    │  01 Mo │ 02 Tu │ …      │
          ├──[_name_cv  ]──┼──[_data_cv scrolls H+V]┤
          │  Dr. Ahmed     │  DM    │  PC   │ …      │
          │  Dr. Sara      │  T1    │  T1   │ …      │
          └────────────────┴────────────────────────-┘

        Clicking any cell or name pre-fills the persistent Quick Assign bar above.
        """
        container = ttk.Frame(self.schedule_tab)
        container.pack(fill="both", expand=True)

        # ── Quick Assign Bar ──────────────────────────────────────────────────
        _QA_OPTIONS: List[tuple] = [
            ("_",    "— Blank / Clear"),
            ("O",    "O   – Day Off"),
            ("L",    "L   – Annual Leave"),
            ("R",    "R   – Random Off"),
            ("PC",   "PC  – Post-Call Off"),
            ("T1",   "T1  – Team 1 Morning"),
            ("T2",   "T2  – Team 2 Morning"),
            ("T3",   "T3  – Team 3 Morning"),
            ("CAHM", "CAHM – Cardiology / Hematology"),
            ("GI",   "GI  – Gastroenterology"),
            ("NE",   "NE  – Neurology"),
            ("NP",   "NP  – Nephrology"),
            ("NENP", "NENP – Neurology + Nephrology"),
            ("PU",   "PU  – Pulmonology"),
            ("DC",   "DC  – Daycare Clinic"),
            ("DM",   "DM  – 16hr Duty (Male)"),
            ("DF",   "DF  – 16hr Duty (Female)"),
        ]
        self._qa_codes = [c for c, _ in _QA_OPTIONS]
        _qa_duty_labels = [lbl for _, lbl in _QA_OPTIONS]

        qa_bar = ttk.LabelFrame(container, text="Quick Assign", padding=(8, 4))
        qa_bar.pack(fill="x", padx=4, pady=(4, 4))

        row1 = ttk.Frame(qa_bar)
        row1.pack(fill="x")

        # Physician combobox
        ttk.Label(row1, text="Physician:").pack(side="left", padx=(0, 4))
        self._qa_ph_var = tk.StringVar()
        _qa_ph_labels = [f"{ph.name}  ({ph.team})" for ph in self.docs]
        self._qa_ph_cb = ttk.Combobox(row1, textvariable=self._qa_ph_var,
                                      values=_qa_ph_labels, state="readonly", width=28)
        if _qa_ph_labels:
            self._qa_ph_cb.current(0)
        self._qa_ph_cb.pack(side="left", padx=(0, 14))
        self._qa_ph_cb.bind("<<ComboboxSelected>>", self._qa_on_change)

        # Day spinbox
        ttk.Label(row1, text="Day:").pack(side="left", padx=(0, 4))
        self._qa_day_var = tk.IntVar(value=1)
        _qa_td = dim(self.yr, self.mo)
        _qa_spin = ttk.Spinbox(row1, from_=1, to=_qa_td,
                               textvariable=self._qa_day_var, width=4,
                               command=self._qa_on_change)
        _qa_spin.pack(side="left")
        _qa_spin.bind("<Return>",   self._qa_on_change)
        _qa_spin.bind("<FocusOut>", self._qa_on_change)

        self._qa_dow_lbl = ttk.Label(row1, text="", foreground="gray", width=5)
        self._qa_dow_lbl.pack(side="left", padx=(4, 14))

        # Current assignment display
        ttk.Label(row1, text="Current:").pack(side="left", padx=(0, 4))
        self._qa_cur_lbl = ttk.Label(row1, text="—", foreground="#1a56c4",
                                     width=8, font=("TkDefaultFont", 9, "bold"))
        self._qa_cur_lbl.pack(side="left", padx=(0, 14))

        # Duty combobox
        ttk.Label(row1, text="Assign as:").pack(side="left", padx=(0, 4))
        self._qa_duty_var = tk.StringVar()
        self._qa_duty_cb = ttk.Combobox(row1, textvariable=self._qa_duty_var,
                                        values=_qa_duty_labels, state="readonly", width=28)
        self._qa_duty_cb.current(0)
        self._qa_duty_cb.pack(side="left", padx=(0, 8))

        # Assign button + status
        ttk.Button(row1, text="✓  Assign", command=self._qa_assign,
                   style="Accent.TButton").pack(side="left", padx=(0, 12))
        self._qa_status_lbl = ttk.Label(row1, text="  ← click any cell to pre-fill",
                                        foreground="gray")
        self._qa_status_lbl.pack(side="left")

        # Trigger initial state
        self.after_idle(self._qa_on_change)

        grid_outer = ttk.Frame(container)
        grid_outer.pack(fill="both", expand=True, padx=4, pady=(0, 4))

        # ── Left column: corner + name canvas ────────────────────────
        left_col = ttk.Frame(grid_outer)
        left_col.pack(side="left", fill="y")

        self._corner_cv = tk.Canvas(left_col, width=_NAME_W, height=_HDR_H,
                                    highlightthickness=0)
        self._corner_cv.pack()

        self._name_cv = tk.Canvas(left_col, width=_NAME_W,
                                  highlightthickness=1,
                                  highlightbackground="#cccccc")
        self._name_cv.pack(fill="y", expand=True)
        self._name_cv.bind("<Button-1>", self._on_name_click)

        # ── Right column: day header + data grid ──────────────────────
        right_col = ttk.Frame(grid_outer)
        right_col.pack(side="left", fill="both", expand=True)

        self._hdr_cv = tk.Canvas(right_col, height=_HDR_H, highlightthickness=0)
        self._hdr_cv.pack(fill="x")

        data_wrap = ttk.Frame(right_col)
        data_wrap.pack(fill="both", expand=True)
        data_wrap.rowconfigure(0, weight=1)
        data_wrap.columnconfigure(0, weight=1)

        self._data_cv = tk.Canvas(data_wrap, highlightthickness=0)
        xsb = ttk.Scrollbar(data_wrap, orient="horizontal",
                            command=self._grid_xscroll)
        ysb = ttk.Scrollbar(data_wrap, orient="vertical",
                            command=self._grid_yscroll)
        self._data_cv.configure(xscrollcommand=xsb.set,
                                yscrollcommand=ysb.set)
        self._data_cv.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")

        self._data_cv.bind("<Button-1>", self._on_grid_click)

        for w in (self._data_cv, self._name_cv):
            w.bind("<MouseWheel>",
                   lambda e: self._grid_yscroll(
                       "scroll", int(-1 * (e.delta / 120)), "units"))

    # ── Grid scroll helpers ───────────────────────────────────────────────
    def _grid_xscroll(self, *args):
        self._hdr_cv.xview(*args)
        self._data_cv.xview(*args)

    def _grid_yscroll(self, *args):
        self._data_cv.yview(*args)
        self._name_cv.yview(*args)

    # ── Canvas drawing ────────────────────────────────────────────────────
    def _draw_grid(self):
        """Redraw the entire canvas-based schedule grid."""
        td   = dim(self.yr, self.mo)
        docs = self.docs
        nd   = len(docs)

        for cv in (self._corner_cv, self._name_cv, self._hdr_cv, self._data_cv):
            cv.delete("all")

        # Palette
        C_CORNER    = "#2C3E50"
        C_WD_HDR    = "#2980B9"
        C_WE_HDR    = "#C0392B"
        C_NAME_ODD  = "#EBF5FB"
        C_NAME_EVEN = "#F8F9FA"
        C_BORDER    = "#D5D8DC"
        C_WHITE     = "#FFFFFF"
        TEAM_COLORS = {"T1": "#2980B9", "T2": "#27AE60", "T3": "#8E44AD"}

        # ── Corner ────────────────────────────────────────────────────
        self._corner_cv.create_rectangle(
            0, 0, _NAME_W, _HDR_H, fill=C_CORNER, outline="")
        self._corner_cv.create_text(
            _NAME_W // 2, _HDR_H // 2, text="PHYSICIAN",
            fill=C_WHITE, font=("TkDefaultFont", 9, "bold"))

        # ── Name column ───────────────────────────────────────────────
        total_name_h = max(nd * _CELL_H, 1)
        self._name_cv.configure(
            scrollregion=(0, 0, _NAME_W, total_name_h))

        for ri, ph in enumerate(docs):
            y0 = ri * _CELL_H
            y1 = y0 + _CELL_H
            bg = C_NAME_ODD if ri % 2 == 0 else C_NAME_EVEN
            self._name_cv.create_rectangle(
                0, y0, _NAME_W, y1, fill=bg, outline=C_BORDER)
            tc = TEAM_COLORS.get(ph.team, "#7F8C8D")
            # Team colour bar on the left edge
            self._name_cv.create_rectangle(
                0, y0, 5, y1, fill=tc, outline="")
            # Physician name, truncated to fit
            name = ph.name
            self._name_cv.create_text(
                10, (y0 + y1) // 2, text=name, anchor="w",
                fill="#1A252F", font=("TkDefaultFont", 9))

        # ── Day header ────────────────────────────────────────────────
        total_w = td * _CELL_W
        self._hdr_cv.configure(scrollregion=(0, 0, total_w, _HDR_H))

        for d in range(1, td + 1):
            x0  = (d - 1) * _CELL_W
            x1  = x0 + _CELL_W
            dow = day_of_week(self.yr, self.mo, d)
            bg  = C_WE_HDR if is_we(self.yr, self.mo, d) else C_WD_HDR
            self._hdr_cv.create_rectangle(
                x0, 0, x1, _HDR_H, fill=bg, outline=C_BORDER)
            self._hdr_cv.create_text(
                (x0 + x1) // 2, 16, text=str(d),
                fill=C_WHITE, font=("TkDefaultFont", 9, "bold"))
            self._hdr_cv.create_text(
                (x0 + x1) // 2, 34, text=DN[dow],
                fill=C_WHITE, font=("TkDefaultFont", 8))

        # ── Data cells ────────────────────────────────────────────────
        self._data_cv.configure(
            scrollregion=(0, 0, total_w, nd * _CELL_H))

        for ri, ph in enumerate(docs):
            y0 = ri * _CELL_H
            y1 = y0 + _CELL_H
            for d in range(1, td + 1):
                x0   = (d - 1) * _CELL_W
                x1   = x0 + _CELL_W
                code = self.get(ph.id, d)
                hex_bg = COLOR_MAP.get(code, "FFFFFF")
                # Blank cells get a subtle grid pattern
                if code == "_":
                    hex_bg = ("F2F3F4" if is_we(self.yr, self.mo, d)
                              else ("EBF5FB" if ri % 2 == 0 else "FDFEFE"))
                self._data_cv.create_rectangle(
                    x0, y0, x1, y1, fill=f"#{hex_bg}", outline=C_BORDER)
                short = SHIFTS[code]["short"]
                if short:
                    bold = ("bold" if code in DUTY_SET or code == "DC"
                            else "normal")
                    self._data_cv.create_text(
                        (x0 + x1) // 2, (y0 + y1) // 2,
                        text=short, fill="#1A252F",
                        font=("TkDefaultFont", 8, bold))

    # ── Canvas click handlers ─────────────────────────────────────────────
    def _on_grid_click(self, event):
        """Click on a day cell → pre-fill the Quick Assign bar."""
        cx  = self._data_cv.canvasx(event.x)
        cy  = self._data_cv.canvasy(event.y)
        col = int(cx // _CELL_W)
        row = int(cy // _CELL_H)
        td  = dim(self.yr, self.mo)
        day = col + 1
        if not (1 <= day <= td) or not (0 <= row < len(self.docs)):
            return
        self._qa_ph_cb.current(row)
        self._qa_day_var.set(day)
        self._qa_on_change()

    def _on_name_click(self, event):
        """Click on physician name → pre-fill the Quick Assign bar (day stays unchanged)."""
        cy  = self._name_cv.canvasy(event.y)
        row = int(cy // _CELL_H)
        if 0 <= row < len(self.docs):
            self._qa_ph_cb.current(row)
            self._qa_on_change()

    # ── Quick Assign bar helpers ──────────────────────────────────────────
    def _qa_selected_pid(self) -> Optional[int]:
        """Return the Doctor.id of the physician selected in the Quick Assign bar."""
        idx = self._qa_ph_cb.current()
        if idx < 0 or idx >= len(self.docs):
            return None
        return self.docs[idx].id

    def _qa_on_change(self, *_):
        """Update weekday label, current-assignment label and auto-fill duty dropdown."""
        if not hasattr(self, "_qa_ph_cb"):
            return
        pid = self._qa_selected_pid()
        if pid is None:
            return
        try:
            day = int(self._qa_day_var.get())
        except (ValueError, tk.TclError):
            return
        td = dim(self.yr, self.mo)
        day = max(1, min(day, td))
        self._qa_day_var.set(day)

        # Weekday label
        dow = date(self.yr, self.mo + 1, day).weekday()
        self._qa_dow_lbl.config(text=["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][dow])

        # Current assignment
        key = f"{pid}|{self.yr}|{self.mo}|{day}"
        cur = self.asgn.get(key, "_")
        short = SHIFTS.get(cur, {}).get("short", cur)
        self._qa_cur_lbl.config(text=short if short else "—")

        # Auto-fill duty dropdown to current assignment
        try:
            self._qa_duty_cb.current(self._qa_codes.index(cur))
        except ValueError:
            self._qa_duty_cb.current(0)

        # Clear old status
        self._qa_status_lbl.config(text="", foreground="gray")

    def _qa_assign(self):
        """Read the Quick Assign bar and save the selection to the schedule."""
        pid = self._qa_selected_pid()
        if pid is None:
            self._qa_status_lbl.config(text="No physician selected.", foreground="#dc2626")
            return
        try:
            day = int(self._qa_day_var.get())
        except (ValueError, tk.TclError):
            self._qa_status_lbl.config(text="Invalid day.", foreground="#dc2626")
            return
        td = dim(self.yr, self.mo)
        if not (1 <= day <= td):
            self._qa_status_lbl.config(text=f"Day must be 1–{td}.", foreground="#dc2626")
            return

        idx = self._qa_duty_cb.current()
        if idx < 0 or idx >= len(self._qa_codes):
            self._qa_status_lbl.config(text="No duty selected.", foreground="#dc2626")
            return
        code = self._qa_codes[idx]

        self.setv(pid, day, code)
        self.refresh_schedule()
        self.refresh_summary()

        # Refresh current-assignment label
        short = SHIFTS.get(code, {}).get("short", code) if code != "_" else "—"
        self._qa_cur_lbl.config(text=short if short else "—")

        ph = next((d for d in self.docs if d.id == pid), None)
        ph_name = ph.name if ph else "?"
        dow = date(self.yr, self.mo + 1, day).weekday()
        dow_str = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][dow]
        self._qa_status_lbl.config(
            text=f"✓  {ph_name} — Day {day} ({dow_str}) → {code}",
            foreground="#16a34a"
        )

    def build_summary_tab(self):
        container = ttk.Frame(self.summary_tab)
        container.pack(fill="both", expand=True)
        cols = ["name", "team", "h8", "h16", "total", "calls", "daycare", "postcall", "off", "leave", "blocked", "random", "weekend_off"]
        self.summary_tree = ttk.Treeview(container, columns=cols, show="headings")
        widths = {"name": 220, "team": 60, "h8": 70, "h16": 70, "total": 80, "calls": 70, "daycare": 70, "postcall": 80, "off": 60, "leave": 60, "blocked": 70, "random": 70, "weekend_off": 90}
        for c in cols:
            self.summary_tree.heading(c, text=c.replace("_", " ").title())
            self.summary_tree.column(c, width=widths[c], anchor="center")
        self.summary_tree.pack(fill="both", expand=True)

    def nav_month(self, delta: int):
        m = self.mo + delta
        y = self.yr
        if m < 0:
            m = 11
            y -= 1
        elif m > 11:
            m = 0
            y += 1
        self.mo, self.yr = m, y
        self.refresh_all()

    def refresh_all(self):
        self.month_lbl.config(text=f"{MONTHS[self.mo]} {self.yr}")
        self.refresh_doctor_selector()
        self.refresh_schedule()
        self.refresh_summary()

    def parse_doctor_names(self, path: str) -> List[tuple]:
        """
        Parse a doctor list file and return a list of (name, initials) tuples.

        Supported file formats: .txt and .docx

        Supported line formats (all are backwards-compatible):
          - Name only:            DR. ALAA
          - Name | Initials:      DR. ALAA | ALA
          - Name TAB Initials:    DR. ALAA<TAB>ALA

        Lines may optionally start with bullets, numbers, or dashes:
          - 1. DR. ALAA | ALA
          - • DR. ALAA | ALA
        """
        lower = path.lower()
        text  = ""
        if lower.endswith(".txt"):
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
        elif lower.endswith(".docx"):
            with ZipFile(path) as zf:
                xml_bytes = zf.read("word/document.xml")
            root = ET.fromstring(xml_bytes)
            ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            paras = []
            for p in root.findall('.//w:p', ns):
                runs = [t.text or "" for t in p.findall('.//w:t', ns)]
                line = "".join(runs).strip()
                if line:
                    paras.append(line)
            text = "\n".join(paras)
        else:
            raise ValueError("Supported formats are .txt and .docx")

        results: List[tuple] = []
        seen: set = set()
        for raw in text.splitlines():
            # Strip leading bullets / numbering
            line = re.sub(r"^[\-•*\d\.)\s]+", "", raw.strip()).strip()
            if not line:
                continue
            # Split on pipe or tab to get optional initials
            if "|" in line:
                parts    = line.split("|", 1)
                name     = parts[0].strip()
                initials = parts[1].strip()
            elif "\t" in line:
                parts    = line.split("\t", 1)
                name     = parts[0].strip()
                initials = parts[1].strip()
            else:
                name     = line
                initials = ""
            if name and name.casefold() not in seen:
                seen.add(name.casefold())
                results.append((name, initials))
        return results

    def import_doctors(self):
        path = filedialog.askopenfilename(
            title="Import doctor list",
            filetypes=[("Supported files", "*.txt *.docx"),
                       ("Text files", "*.txt"),
                       ("Word documents", "*.docx")],
        )
        if not path:
            return
        try:
            entries = self.parse_doctor_names(path)
        except Exception as exc:
            messagebox.showerror("Import doctors", str(exc))
            return

        existing_map = {d.name.casefold(): d for d in self.docs}
        added = skipped = updated = 0

        for name, initials in entries:
            key = name.casefold()
            if key in existing_map:
                # Doctor already exists — update initials if the file provided them
                if initials:
                    existing_map[key].initials = initials
                    updated += 1
                else:
                    skipped += 1
            else:
                team = TEAMS[len(self.docs) % 3]
                self.docs.append(
                    Doctor(self.next_doc_id, name, "Not specified", team,
                           initials=initials)
                )
                self.next_doc_id += 1
                existing_map[key] = self.docs[-1]
                added += 1

        self.refresh_all()
        parts = []
        if added:   parts.append(f"{added} added")
        if updated: parts.append(f"{updated} initials updated")
        if skipped: parts.append(f"{skipped} skipped (already exist)")
        self.status_var.set("Import complete: " + ", ".join(parts) + ".")

    def refresh_doctor_selector(self):
        # Rebuild the scrollable physician list with checkboxes
        self.refresh_doctor_list()

        # Values for the Assign Specialty combo: id + name + current specialty
        assign_values = [f"{d.id}: {d.name}  [{d.spec}]" for d in self.docs]
        self.assign_doc_combo["values"] = assign_values
        if assign_values and self.assign_doc_var.get() not in assign_values:
            self.assign_doc_combo.current(0)
        if not assign_values:
            self.assign_doc_var.set("")

        # Keep the Quick Assign bar physician combobox in sync
        if hasattr(self, "_qa_ph_cb"):
            ph_labels = [f"{ph.name}  ({ph.team})" for ph in self.docs]
            prev_idx  = self._qa_ph_cb.current()
            self._qa_ph_cb["values"] = ph_labels
            if ph_labels:
                new_idx = prev_idx if 0 <= prev_idx < len(ph_labels) else 0
                self._qa_ph_cb.current(new_idx)
            else:
                self._qa_ph_var.set("")

    def refresh_doctor_list(self):
        """Rebuild the scrollable physician list with checkboxes and initials entries.

        All widgets – header labels and per-physician controls – are placed
        directly into doc_list_frame using grid so every column is
        pixel-perfect aligned regardless of widget type.
        """
        for w in self.doc_list_frame.winfo_children():
            w.destroy()
        self._doc_check_vars    = {}
        self._doc_initials_vars = {}
        self._doc_fdd_vars      = {}

        if not self.docs:
            ttk.Label(self.doc_list_frame,
                      text="No physicians added yet.").grid(
                          row=0, column=0, padx=6, pady=4, sticky="w")
            return

        # ------------------------------------------------------------------
        # Column layout (shared between header and every data row)
        #   col 0 : checkbox          (~26 px)
        #   col 1 : Name              (~165 px)
        #   col 2 : Initials entry    (~52 px)
        #   col 3 : 1st Day spinbox   (~62 px)
        #   col 4 : Specialty label   (~195 px)
        #   col 5 : Team label        (~52 px)
        # ------------------------------------------------------------------
        _col_min = [26, 165, 52, 62, 195, 52]
        for ci, mw in enumerate(_col_min):
            self.doc_list_frame.columnconfigure(ci, minsize=mw)

        # Header row (grid row 0)
        _hdr_font = ("TkDefaultFont", 9, "bold")
        for ci, txt in enumerate(["", "Name", "Initials", "1st Day", "Specialty", "Team"]):
            ttk.Label(self.doc_list_frame, text=txt, anchor="w",
                      font=_hdr_font).grid(row=0, column=ci,
                                           sticky="w", padx=3, pady=(3, 1))

        # Separator (grid row 1, spans all columns)
        ttk.Separator(self.doc_list_frame, orient="horizontal").grid(
            row=1, column=0, columnspan=6, sticky="ew", pady=2)

        td_cur = dim(self.yr, self.mo)

        for ri, ph in enumerate(self.docs):
            gr = ri + 2   # grid row (0=header, 1=sep, 2+ = data)

            # col 0 – Checkbox
            chk_var = tk.IntVar(value=0)
            self._doc_check_vars[ph.id] = chk_var
            ttk.Checkbutton(self.doc_list_frame,
                            variable=chk_var).grid(row=gr, column=0,
                                                   sticky="w", padx=3)

            # col 1 – Name
            ttk.Label(self.doc_list_frame, text=ph.name,
                      anchor="w").grid(row=gr, column=1,
                                       sticky="w", padx=3)

            # col 2 – Initials entry
            init_var = tk.StringVar(value=ph.initials)
            self._doc_initials_vars[ph.id] = init_var
            init_entry = ttk.Entry(self.doc_list_frame,
                                   textvariable=init_var, width=6)
            init_entry.grid(row=gr, column=2, sticky="w", padx=3)
            init_entry.bind("<FocusOut>",
                            lambda e, pid=ph.id, sv=init_var:
                                self._save_initials(pid, sv))
            init_entry.bind("<Return>",
                            lambda e, pid=ph.id, sv=init_var:
                                self._save_initials(pid, sv))

            # col 3 – First Duty Day spinbox  (1 = default / auto)
            fdd_var = tk.IntVar(value=max(1, int(ph.first_duty_day)))
            self._doc_fdd_vars[ph.id] = fdd_var
            fdd_spin = ttk.Spinbox(self.doc_list_frame,
                                   from_=1, to=td_cur,
                                   textvariable=fdd_var, width=5)
            fdd_spin.grid(row=gr, column=3, sticky="w", padx=3)
            for _ev in ("<FocusOut>", "<Return>", "<<Increment>>", "<<Decrement>>"):
                fdd_spin.bind(_ev,
                              lambda e, pid=ph.id, sv=fdd_var:
                                  self._save_fdd(pid, sv))

            # col 4 – Specialty label
            ttk.Label(self.doc_list_frame, text=ph.spec,
                      anchor="w").grid(row=gr, column=4,
                                       sticky="w", padx=3)

            # col 5 – Team label
            ttk.Label(self.doc_list_frame, text=ph.team,
                      anchor="w").grid(row=gr, column=5,
                                       sticky="w", padx=3)

    def _save_initials(self, pid: int, sv: tk.StringVar):
        """Persist the initials entry value back to the Doctor object."""
        doc = next((d for d in self.docs if d.id == pid), None)
        if doc is not None:
            doc.initials = sv.get().strip().upper()

    def _save_fdd(self, pid: int, sv: tk.IntVar):
        """Persist the first_duty_day spinbox value back to the Doctor object."""
        doc = next((d for d in self.docs if d.id == pid), None)
        if doc is not None:
            try:
                val = int(sv.get())
                td_cur = dim(self.yr, self.mo)
                doc.first_duty_day = max(1, min(val, td_cur))
            except (ValueError, tk.TclError):
                doc.first_duty_day = 1

    def _flush_fdd_vars(self):
        """Force-save all first_duty_day spinbox values (called before scheduling)."""
        for pid, sv in self._doc_fdd_vars.items():
            self._save_fdd(pid, sv)

    def _select_all_docs(self):
        for var in self._doc_check_vars.values():
            var.set(1)

    def _deselect_all_docs(self):
        for var in self._doc_check_vars.values():
            var.set(0)

    def add_doc(self):
        name = self.name_var.get().strip()
        if not name:
            return
        spec     = self.spec_var.get().strip() or "Internal Medicine"
        initials = self.initials_var.get().strip().upper()
        team     = TEAMS[len(self.docs) % 3]
        self.docs.append(Doctor(self.next_doc_id, name, spec, team, initials))
        self.next_doc_id += 1
        self.name_var.set("")
        self.initials_var.set("")
        self.spec_var.set("Internal Medicine")
        self.refresh_all()

    def remove_selected_doc(self):
        """Single-doctor remove — kept for internal compatibility."""
        self.remove_selected_docs()

    def remove_selected_docs(self):
        """Bulk-remove all physicians whose checkbox is ticked."""
        # Flush any pending initials / first-duty-day edits first
        for pid, sv in self._doc_initials_vars.items():
            self._save_initials(pid, sv)
        self._flush_fdd_vars()

        to_remove = {pid for pid, var in self._doc_check_vars.items() if var.get() == 1}
        if not to_remove:
            messagebox.showinfo("Remove", "No physicians selected.\nTick the checkbox next to a physician's name to select them.")
            return
        names = [d.name for d in self.docs if d.id in to_remove]
        msg = (f"Remove {len(to_remove)} physician(s)?\n"
               + "\n".join(f"  • {n}" for n in names)
               + "\n\nThis will also delete their assignments and leave blocks.")
        if not messagebox.askyesno("Remove physicians", msg):
            return
        self.docs        = [d for d in self.docs if d.id not in to_remove]
        self.leaves      = [b for b in self.leaves if b.pid not in to_remove]
        self.manual_asgns = [ma for ma in self.manual_asgns if ma.pid not in to_remove]
        self.asgn        = {k: v for k, v in self.asgn.items()
                            if not any(k.startswith(f"{pid}|") for pid in to_remove)}
        self.refresh_all()
        self.status_var.set(f"Removed {len(to_remove)} physician(s).")

    def assign_specialty_to_doc(self):
        """Assign (or change) the specialty of an existing physician.

        Reads the selection from the 'Assign Specialty' row in the Physician
        Management panel and updates the Doctor record in-place.  The change
        takes effect on the next Auto-Schedule run; any already-generated
        schedule for the current month is left untouched.
        """
        choice = self.assign_doc_var.get().strip()
        if not choice:
            return
        pid = int(choice.split(":", 1)[0])
        doctor = next((d for d in self.docs if d.id == pid), None)
        if doctor is None:
            return
        new_spec = self.assign_spec_var.get().strip()
        if not new_spec:
            return
        if new_spec == doctor.spec:
            self.status_var.set(f"{doctor.name} is already assigned to '{new_spec}' – no change.")
            return
        old_spec = doctor.spec
        doctor.spec = new_spec
        # If the new specialty is one of the team morning slots, sync ph.team too
        # so the schedule grid and summary columns stay consistent.
        sc = specialty_code_from_label(new_spec)
        if sc and sc in TEAMS:
            doctor.team = sc
        self.refresh_all()          # refreshes assign combo so it shows the new specialty
        self.status_var.set(f"{doctor.name}: specialty changed from '{old_spec}'  →  '{new_spec}'.")

    def open_quick_assign_dialog(self):
        """Switch to the Schedule tab and focus the Quick Assign bar."""
        if not self.docs:
            messagebox.showinfo("No physicians", "Add at least one physician first.")
            return
        self.notebook.select(self.schedule_tab)
        if hasattr(self, "_qa_ph_cb"):
            self._qa_ph_cb.focus_set()

    def open_leave_dialog(self):
        LeaveDialog(self)

    def open_spec_block_dialog(self):
        SpecialtyBlockDialog(self)

    def open_manual_assign_dialog(self):
        ManualAssignDialog(self)

    def add_leave(self, pid: int, f: str, t: str):
        self.leaves.append(LeaveBlock(self.next_leave_id, pid, f, t))
        self.next_leave_id += 1
        for d in range(1, dim(self.yr, self.mo) + 1):
            cur = ds(self.yr, self.mo, d)
            if f <= cur <= t:
                self.setv(pid, d, "L")
        self.refresh_all()

    def delete_leave(self, leave_id: int):
        block = next((x for x in self.leaves if x.id == leave_id), None)
        if not block:
            return
        self.leaves = [x for x in self.leaves if x.id != leave_id]
        for d in range(1, dim(self.yr, self.mo) + 1):
            cur = ds(self.yr, self.mo, d)
            if block.f <= cur <= block.t and self.asgn.get(self.ak(block.pid, d)) == "L":
                self.asgn.pop(self.ak(block.pid, d), None)
        self.refresh_all()

    def add_spec_block(self, code: str, f: str, t: str):
        self.spec_blocks.append(SpecialtyBlock(self.next_spec_block_id, code, f, t))
        self.next_spec_block_id += 1
        self.status_var.set(f"Blocked {SHIFTS[code]['label']} from {f} to {t}.")

    def delete_spec_block(self, block_id: int):
        self.spec_blocks = [b for b in self.spec_blocks if b.id != block_id]
        self.status_var.set("Specialty block removed.")

    def clear_month(self):
        if not messagebox.askyesno("Clear month",
                                   "Remove all assignments for the current month?\n"
                                   "(Manual assignments for this month will also be cleared.)"):
            return
        marker = f"|{self.yr}|{self.mo}|"
        self.asgn = {k: v for k, v in self.asgn.items() if marker not in k}
        # Clear manual assignments for this month
        self.manual_asgns = [ma for ma in self.manual_asgns
                             if not (ma.pid in {d.id for d in self.docs})]
        # Simpler: just clear all manual assignments (they are month-specific)
        self.manual_asgns = []
        self.refresh_all()
        self.status_var.set("Current month cleared (including manual assignments).")

    def schedule(self):
        if len(self.docs) < 3:
            messagebox.showerror("Not enough physicians", "Add at least 3 physicians.")
            return
        # Flush any unsaved initials / first-duty-day edits before scheduling
        for pid, sv in self._doc_initials_vars.items():
            self._save_initials(pid, sv)
        self._flush_fdd_vars()

        # Stamp manual assignments as hard pre-sets into the base grid.
        # These are treated as immovable by auto_schedule (non-blank slots are
        # skipped by Phase 7 Step 1 and are excluded from Phase 4/5 via hb).
        for ma in self.manual_asgns:
            self.setv(ma.pid, ma.day, ma.code)

        base = dict(self.asgn)
        res = auto_schedule(self.docs, base, self.leaves, self.spec_blocks, self.yr, self.mo)
        if "err" in res:
            messagebox.showerror("Auto-schedule", res["err"])
            return
        self.asgn = res["a"]

        # Re-stamp manual assignments so auto_schedule can't accidentally
        # overwrite them (e.g. via PC stamping from a nearby DM/DF).
        for ma in self.manual_asgns:
            self.setv(ma.pid, ma.day, ma.code)

        pairs = res["pairs"]
        td = dim(self.yr, self.mo)
        full  = sum(1 for p in pairs if p["male"] and p["female"])
        part  = sum(1 for p in pairs if bool(p["male"]) ^ bool(p["female"]))
        uncov = sum(1 for p in pairs if not p["male"] and not p["female"])
        self.refresh_all()
        n_manual = len(self.manual_asgns)
        manual_note = f"  ({n_manual} manual pre-set{'s' if n_manual != 1 else ''} honoured)" if n_manual else ""
        self.status_var.set(
            f"Auto-schedule complete — {full}/{td} full duty coverage; "
            f"{part} partial; {uncov} uncovered.{manual_note}"
        )

    def compute_summary(self):
        td = dim(self.yr, self.mo)
        rows = []
        for ph in self.docs:
            h8 = h16 = oc = dc_d = pc_d = o_d = l_d = r_d = we_off = 0
            for d in range(1, td + 1):
                t = self.get(ph.id, d)
                if t in MORNING_K:
                    h8 += 8
                if t == "DC":
                    h8 += 8
                    dc_d += 1
                if t in DUTY_SET:
                    h16 += 16
                    oc += 1
                if t == "PC":
                    pc_d += 1
                if t == "O":
                    o_d += 1
                if t == "L":
                    l_d += 1
                if t == "R":
                    r_d += 1
                if is_we(self.yr, self.mo, d) and (t in OFF_SET or t == "_"):
                    we_off += 1
            ms = ds(self.yr, self.mo, 1)
            me = ds(self.yr, self.mo, td)
            blocked = 0
            for b in self.leaves:
                if b.pid != ph.id:
                    continue
                ef = max(b.f, ms)
                et = min(b.t, me)
                if ef <= et:
                    d1 = datetime.strptime(ef, "%Y-%m-%d").date()
                    d2 = datetime.strptime(et, "%Y-%m-%d").date()
                    blocked += (d2 - d1).days + 1
            rows.append({
                "name": ph.name,
                "team": ph.team,
                "h8": h8,
                "h16": h16,
                "total": h8 + h16,
                "calls": oc,
                "daycare": dc_d,
                "postcall": pc_d,
                "off": o_d,
                "leave": l_d,
                "blocked": blocked,
                "random": r_d,
                "weekend_off": we_off,
            })
        return rows

    def refresh_schedule(self):
        """Redraw the canvas-based schedule grid."""
        self._draw_grid()

    def refresh_summary(self):
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        for row in self.compute_summary():
            self.summary_tree.insert(
                "", "end",
                values=[row[k] for k in self.summary_tree["columns"]])

    # ------------------------------------------------------------------
    # Simplified "Rota Board" export helpers
    # ------------------------------------------------------------------

    def _generate_initials(self) -> Dict[int, str]:
        """Return a unique 2-letter uppercase abbreviation for every physician.

        Strategy (in priority order):
          1. First letter of first name  +  first letter of last name.
          2. First two letters of first name.
          3. First letter  +  a digit/letter suffix to break ties.
        """
        used: set = set()
        result: Dict[int, str] = {}

        for ph in self.docs:
            # Use the physician's own initials if they supplied one.
            if ph.initials:
                cand = ph.initials.upper()[:4]
                if cand not in used:
                    used.add(cand)
                    result[ph.id] = cand
                    continue
                # Collision: fall through to auto-generation below.

            # Strip leading title ("Dr.", "Dr")
            name = re.sub(r'\bDr\.?\s*', '', ph.name, flags=re.IGNORECASE).strip()
            words = [w for w in name.split() if w]

            def try_cand(c: str) -> Optional[str]:
                c = c.upper()[:2]
                if c not in used:
                    used.add(c)
                    return c
                return None

            chosen: Optional[str] = None

            # Strategy 1 – first letters of first two name-words
            if len(words) >= 2:
                chosen = try_cand(words[0][0] + words[1][0])

            # Strategy 2 – first two letters of the first word
            if chosen is None and words and len(words[0]) >= 2:
                chosen = try_cand(words[0][:2])

            # Strategy 3 – first letter + suffix to resolve collision
            if chosen is None:
                base = words[0][0] if words else "X"
                for suffix in "23456789ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                    chosen = try_cand(base + suffix)
                    if chosen:
                        break

            result[ph.id] = chosen or "??"

        return result

    def export_simplified_xlsx(self):
        """Export a compact rota-board style schedule — one row per day,
        one column per specialty slot, cells filled with physician initials.

        Layout mirrors the hospital-board format:
          Col A  : day number
          Col B  : day name
          Cols C+ : morning specialty slots  (T1 … Daycare)
          Next 2  : On-call (Male) / (Female)
          Last    : Initials legend  (AB – Dr. Ahmed …)
        """
        td = dim(self.yr, self.mo)

        path = filedialog.asksaveasfilename(
            title="Export Rota Schedule",
            defaultextension=".xlsx",
            initialfile=f"Rota_{MONTHS[self.mo]}_{self.yr}.xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return

        initials = self._generate_initials()   # {pid: "AW"}

        # Morning specialty columns: (internal code, display label)
        MORNING_COLS: List[tuple] = [
            ("T1",   "T1"),
            ("T2",   "T2"),
            ("T3",   "T3"),
            ("PU",   "PUL"),
            ("CAHM", "CA/HM"),
            ("NE",   "NEU"),
            ("NP",   "NEPH"),
            ("GI",   "GAS"),
            ("DC",   "Daycare"),
        ]
        N_MORN      = len(MORNING_COLS)
        DATA_COL    = 3                      # first morning specialty column (1-indexed)
        DM_COL      = DATA_COL + N_MORN      # Male on-call
        DF_COL      = DM_COL + 1             # Female on-call
        LEG_COL     = DF_COL + 1             # Legend
        TOTAL_COLS  = LEG_COL

        # ── Build inverted schedule (day → slot → list of initials) ─────────
        SLOT_CODES = [code for code, _ in MORNING_COLS] + ["DM", "DF"]
        day_slots: Dict[int, Dict[str, List[str]]] = {}
        for d in range(1, td + 1):
            row: Dict[str, List[str]] = {s: [] for s in SLOT_CODES}
            for ph in self.docs:
                code = self.get(ph.id, d)
                if code in row:
                    row[code].append(initials[ph.id])
            day_slots[d] = row

        # ── Colour palette ───────────────────────────────────────────────────
        C_NAVY      = "1F3864"   # title background
        C_BLUE      = "2E75B6"   # morning section header
        C_BROWN     = "843C00"   # on-call section header
        C_PURPLE    = "7030A0"   # legend header
        C_AMBER     = "F4B942"   # morning column sub-headers
        C_GREEN_H   = "70AD47"   # on-call column sub-headers
        C_WKND_ROW  = "FCE4D6"   # weekend data rows
        C_WKND_LBL  = "E05C00"   # weekend day label fill
        C_OC_CELL   = "E2EFDA"   # on-call data cell (filled)
        C_DC_CELL   = "FFF2CC"   # daycare cell
        C_ALT       = "F2F2F2"   # alternating row shade
        C_LEG_CELL  = "EAD1DC"   # legend entry cells
        WHITE       = "FFFFFF"

        def F(hex_col: str) -> PatternFill:
            return PatternFill(fill_type="solid", fgColor=hex_col)

        def BF(color: str = "000000", sz: int = 9, bold: bool = True) -> Font:
            return Font(bold=bold, color=color, size=sz)

        CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LFT = Alignment(horizontal="left",   vertical="center", wrap_text=False)

        # ── Workbook ─────────────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Rota"

        # Row 1 – Title (merged across all columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
        tc = ws.cell(row=1, column=1,
                     value=f"First On Call Internal Medicine  ROTA  /  {MONTHS[self.mo].upper()}  {self.yr}")
        tc.font      = Font(bold=True, color=WHITE, size=13)
        tc.fill      = F(C_NAVY)
        tc.alignment = CTR
        ws.row_dimensions[1].height = 22

        # Row 2 – Section group headers
        # Month/Year (cols A-B, rows 2-3 merged)
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        mc_m = ws.cell(row=2, column=1, value=f"{MONTHS[self.mo]}\n{self.yr}")
        mc_m.font = BF(WHITE, 10)
        mc_m.fill = F(C_NAVY)
        mc_m.alignment = CTR

        ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
        mc_b = ws.cell(row=2, column=2, value="")
        mc_b.fill = F(C_NAVY)

        # "MORNING ROUND 7:30 – 15:30"
        ws.merge_cells(start_row=2, start_column=DATA_COL, end_row=2, end_column=DATA_COL + N_MORN - 1)
        mr = ws.cell(row=2, column=DATA_COL, value="MORNING ROUND  7:30 – 15:30")
        mr.font = BF(WHITE, 11)
        mr.fill = F(C_BLUE)
        mr.alignment = CTR

        # "ONCALL 15:30 – 07:30"
        ws.merge_cells(start_row=2, start_column=DM_COL, end_row=2, end_column=DF_COL)
        oc = ws.cell(row=2, column=DM_COL, value="ONCALL\n15:30 – 07:30")
        oc.font = BF(WHITE, 10)
        oc.fill = F(C_BROWN)
        oc.alignment = CTR

        # Legend section header (merged rows 2-3)
        ws.merge_cells(start_row=2, start_column=LEG_COL, end_row=3, end_column=LEG_COL)
        lh = ws.cell(row=2, column=LEG_COL, value="INITIALS  →  PHYSICIAN")
        lh.font = BF(WHITE, 9)
        lh.fill = F(C_PURPLE)
        lh.alignment = CTR

        ws.row_dimensions[2].height = 30

        # Row 3 – Column sub-headers
        for i, (code, label) in enumerate(MORNING_COLS):
            c = ws.cell(row=3, column=DATA_COL + i, value=label)
            c.font = BF(sz=9)
            c.fill = F(C_AMBER)
            c.alignment = CTR

        dm_h = ws.cell(row=3, column=DM_COL, value="(Male)")
        df_h = ws.cell(row=3, column=DF_COL, value="(Female)")
        for h in (dm_h, df_h):
            h.font = BF(sz=9)
            h.fill = F(C_GREEN_H)
            h.alignment = CTR

        ws.row_dimensions[3].height = 18

        # Rows 4+ – Data
        DAY_NAMES = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

        # Legend list (sorted alphabetically by initials)
        legend_entries = sorted(
            [f"{initials[ph.id]}  –  {ph.name}" for ph in self.docs]
        )

        for d in range(1, td + 1):
            r = 3 + d
            dw       = day_of_week(self.yr, self.mo, d)
            weekend  = is_we(self.yr, self.mo, d)
            slots    = day_slots[d]
            row_bg   = C_WKND_ROW if weekend else (C_ALT if d % 2 == 0 else WHITE)

            # Day number
            dn = ws.cell(row=r, column=1, value=d)
            dn.font = BF(WHITE if weekend else "000000", 9)
            dn.fill = F(C_WKND_LBL if weekend else "D9D9D9")
            dn.alignment = CTR

            # Day name
            dd = ws.cell(row=r, column=2, value=DAY_NAMES[dw].upper())
            dd.font = BF(WHITE if weekend else "000000", 9)
            dd.fill = F(C_WKND_LBL if weekend else "D9D9D9")
            dd.alignment = CTR

            # Morning specialty cells
            for i, (code, _) in enumerate(MORNING_COLS):
                val = "/".join(slots[code])
                c = ws.cell(row=r, column=DATA_COL + i, value=val or "")
                c.alignment = CTR
                c.font = Font(bold=bool(val), size=9,
                              color="1F3864" if val else "AAAAAA")
                if code == "DC" and val:
                    c.fill = F(C_DC_CELL)
                else:
                    c.fill = F(C_WKND_ROW if weekend else row_bg)

            # On-call cells
            for col, key in ((DM_COL, "DM"), (DF_COL, "DF")):
                val = "/".join(slots[key])
                c = ws.cell(row=r, column=col, value=val or "")
                c.alignment = CTR
                c.font = Font(bold=bool(val), size=9,
                              color="1F3864" if val else "BBBBBB")
                c.fill = F(C_OC_CELL if val else (C_WKND_ROW if weekend else row_bg))

            # Legend entry (first N rows; one per physician)
            leg_idx = d - 1
            if leg_idx < len(legend_entries):
                lc = ws.cell(row=r, column=LEG_COL, value=legend_entries[leg_idx])
                lc.font = Font(size=8)
                lc.fill = F(C_LEG_CELL)
                lc.alignment = LFT
            else:
                ws.cell(row=r, column=LEG_COL, value="")

            ws.row_dimensions[r].height = 15

        # ── Column widths ────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 6
        for i in range(N_MORN):
            ws.column_dimensions[get_column_letter(DATA_COL + i)].width = 7
        ws.column_dimensions[get_column_letter(DM_COL)].width = 8
        ws.column_dimensions[get_column_letter(DF_COL)].width = 8
        ws.column_dimensions[get_column_letter(LEG_COL)].width = 30

        # Freeze header rows and day-label columns
        ws.freeze_panes = f"{get_column_letter(DATA_COL)}4"

        wb.save(path)
        self.status_var.set(f"Rota schedule exported → {path}")

    def export_xlsx(self):
        td = dim(self.yr, self.mo)
        path = filedialog.asksaveasfilename(title="Export Excel Workbook", defaultextension=".xlsx", initialfile=f"Schedule_{MONTHS[self.mo]}_{self.yr}.xlsx", filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        ws["A1"] = f"Physician Schedule — {MONTHS[self.mo]} {self.yr}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.cell(row=3, column=1, value="Day").font = Font(bold=True)
        ws.cell(row=3, column=2, value="Date").font = Font(bold=True)
        for col in (1, 2):
            ws.cell(row=3, column=col).fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
            ws.cell(row=3, column=col).alignment = Alignment(horizontal="center", vertical="center")
        doctor_start_col = 3
        for col_idx, ph in enumerate(self.docs, start=doctor_start_col):
            # Use custom initials as the column header when supplied; fall back to full name
            col_label = ph.initials if ph.initials else ph.name
            ws.cell(row=2, column=col_idx, value=col_label).font = Font(bold=True)
            ws.cell(row=2, column=col_idx).fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
            ws.cell(row=3, column=col_idx, value=f"{ph.spec} • {ph.team}").font = Font(bold=True)
            ws.cell(row=3, column=col_idx).fill = PatternFill(fill_type="solid", fgColor="EEF5FB")
            ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=3, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for d in range(1, td + 1):
            row_idx = 3 + d
            ws.cell(row=row_idx, column=1, value=f"{p2(d)} {DN[day_of_week(self.yr, self.mo, d)]}")
            ws.cell(row=row_idx, column=2, value=ds(self.yr, self.mo, d))
            for col in (1, 2):
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")
            for col_idx, ph in enumerate(self.docs, start=doctor_start_col):
                code = self.get(ph.id, d)
                cell = ws.cell(row=row_idx, column=col_idx, value=SHIFTS[code]["short"] or "")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_MAP.get(code, "FFFFFF"))
        summary_start = td + 6
        ws.cell(row=summary_start, column=1, value="Summary").font = Font(bold=True, size=12)
        metrics = [("8hr Hrs", "h8"), ("16hr Hrs", "h16"), ("Total Hrs", "total"), ("On-Calls", "calls"), ("Daycare Days", "daycare"), ("Post-Call", "postcall"), ("Days Off", "off"), ("Annual Leave", "leave"), ("Blocked", "blocked"), ("Random", "random"), ("Weekend Off", "weekend_off")]
        for row_offset, (label, _) in enumerate(metrics, start=1):
            ws.cell(row=summary_start + row_offset, column=1, value=label).font = Font(bold=True)
            ws.cell(row=summary_start + row_offset, column=1).fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        stats_by_name = {row["name"]: row for row in self.compute_summary()}
        for col_idx, ph in enumerate(self.docs, start=doctor_start_col):
            stats = stats_by_name.get(ph.name, {})
            for row_offset, (_, key) in enumerate(metrics, start=1):
                ws.cell(row=summary_start + row_offset, column=col_idx, value=stats.get(key, 0)).alignment = Alignment(horizontal="center", vertical="center")
        legend_col = doctor_start_col + max(len(self.docs), 1) + 2
        ws.cell(row=2, column=legend_col, value="Legend").font = Font(bold=True, size=12)
        legend_row = 3
        for key, value in SHIFTS.items():
            if key == "_":
                continue
            ws.cell(row=legend_row, column=legend_col, value=value["short"])
            ws.cell(row=legend_row, column=legend_col + 1, value=value["label"])
            ws.cell(row=legend_row, column=legend_col + 2, value=f'{value["h"]}h')
            ws.cell(row=legend_row, column=legend_col).fill = PatternFill(fill_type="solid", fgColor=COLOR_MAP.get(key, "FFFFFF"))
            legend_row += 1
        block_ws = wb.create_sheet("Blocked Specialties")
        block_headers = ["Specialty", "From", "To"]
        for c, h in enumerate(block_headers, start=1):
            block_ws.cell(row=1, column=c, value=h).font = Font(bold=True)
            block_ws.cell(row=1, column=c).fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        for r, b in enumerate(self.spec_blocks, start=2):
            block_ws.cell(row=r, column=1, value=f"{b.code} - {SHIFTS[b.code]['label']}")
            block_ws.cell(row=r, column=2, value=b.f)
            block_ws.cell(row=r, column=3, value=b.t)
        summary_ws = wb.create_sheet("Summary")
        headers = ["Physician", "Team", "8hr Hrs", "16hr Hrs", "Total Hrs", "On-Calls", "Daycare", "Post-Call", "Off", "Leave", "Random", "Weekend Off"]
        for c, h in enumerate(headers, start=1):
            summary_ws.cell(row=1, column=c, value=h).font = Font(bold=True)
            summary_ws.cell(row=1, column=c).fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        for r, row in enumerate(self.compute_summary(), start=2):
            values = [row["name"], row["team"], row["h8"], row["h16"], row["total"], row["calls"], row["daycare"], row["postcall"], row["off"], row["leave"], row["random"], row["weekend_off"]]
            for c, value in enumerate(values, start=1):
                summary_ws.cell(row=r, column=c, value=value)
        for sheet in (ws, summary_ws, block_ws):
            for col_idx, column_cells in enumerate(sheet.columns, start=1):
                max_len = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 28)
        ws.freeze_panes = "C4"
        summary_ws.freeze_panes = "A2"
        block_ws.freeze_panes = "A2"
        wb.save(path)
        self.status_var.set(f"Excel workbook exported to {path}")

    # ── Firebase integration ─────────────────────────────────────────────────

    def _update_firebase_label(self) -> None:
        """Refresh the toolbar cloud label to reflect current auth state."""
        if not _FIREBASE_AVAILABLE or firebase is None:
            self._fb_user_var.set("☁  Cloud unavailable")
            self._fb_user_lbl.configure(foreground="#aaa")
            return
        if firebase.is_signed_in:
            short = (firebase.email or "")[:28]
            self._fb_user_var.set(f"☁  {short}  [sign out]")
            self._fb_user_lbl.configure(foreground="#1a7f4b")
        else:
            self._fb_user_var.set("☁  Not signed in  [click to sign in]")
            self._fb_user_lbl.configure(foreground="#777")

    def _firebase_login(self) -> None:
        """Open sign-in dialog, or sign out if already signed in."""
        if not _FIREBASE_AVAILABLE or firebase is None:
            messagebox.showinfo("Unavailable",
                                "Firebase is not available.\n"
                                "Run:  pip install requests", parent=self)
            return
        if firebase.is_signed_in:
            if messagebox.askyesno("Sign Out",
                                   f"Signed in as {firebase.email}.\n\nSign out?",
                                   parent=self):
                firebase.sign_out()
                self._update_firebase_label()
                self.status_var.set("Signed out of MedScheduler Cloud.")
        else:
            LoginDialog(self)

    def _firebase_save(self) -> None:
        """Save all app data to Firestore (runs in background thread)."""
        if not _FIREBASE_AVAILABLE or firebase is None:
            messagebox.showinfo("Unavailable",
                                "Firebase is not available.\n"
                                "Run:  pip install requests", parent=self)
            return
        if not firebase.is_signed_in:
            messagebox.showinfo("Not Signed In",
                                "Please sign in to your MedScheduler Cloud account first.",
                                parent=self)
            return
        self.status_var.set("Saving to cloud…")
        data = self._serialize_app()

        def _do():
            try:
                firebase.save_app_data(data)
                self.after(0, lambda: self.status_var.set("✓  Data saved to MedScheduler Cloud."))
            except Exception as exc:
                self.after(0, lambda: (
                    messagebox.showerror("Cloud Save Failed", str(exc), parent=self),
                    self.status_var.set("Cloud save failed."),
                ))

        threading.Thread(target=_do, daemon=True).start()

    def _firebase_load(self) -> None:
        """Load app data from Firestore (runs in background thread)."""
        if not _FIREBASE_AVAILABLE or firebase is None:
            messagebox.showinfo("Unavailable",
                                "Firebase is not available.\n"
                                "Run:  pip install requests", parent=self)
            return
        if not firebase.is_signed_in:
            messagebox.showinfo("Not Signed In",
                                "Please sign in to your MedScheduler Cloud account first.",
                                parent=self)
            return
        if not messagebox.askyesno(
            "Load from Cloud",
            "This will REPLACE the current schedule with your last cloud save.\n\n"
            "Any unsaved local changes will be lost. Continue?",
            parent=self,
        ):
            return
        self.status_var.set("Loading from cloud…")

        def _do():
            try:
                data = firebase.load_app_data()
                if data is None:
                    self.after(0, lambda: (
                        messagebox.showinfo("No Cloud Data",
                                            "No cloud save found for this account.\n"
                                            "Use '💾 Save' to create your first cloud backup.",
                                            parent=self),
                        self.status_var.set("No cloud data found."),
                    ))
                    return
                self.after(0, lambda: self._apply_cloud_data(data))
            except Exception as exc:
                self.after(0, lambda: (
                    messagebox.showerror("Cloud Load Failed", str(exc), parent=self),
                    self.status_var.set("Cloud load failed."),
                ))

        threading.Thread(target=_do, daemon=True).start()

    def _apply_cloud_data(self, data: Dict) -> None:
        """Apply loaded cloud data to app state (must run on main thread)."""
        self._deserialize_app(data)
        self.status_var.set("✓  Schedule loaded from MedScheduler Cloud.")

    def _firebase_files(self) -> None:
        """Open the Cloud Files dialog."""
        if not _FIREBASE_AVAILABLE or firebase is None:
            messagebox.showinfo("Unavailable",
                                "Firebase is not available.\n"
                                "Run:  pip install requests", parent=self)
            return
        if not firebase.is_signed_in:
            messagebox.showinfo("Not Signed In",
                                "Please sign in to your MedScheduler Cloud account first.",
                                parent=self)
            return
        CloudFilesDialog(self)

    # ── Serialization ────────────────────────────────────────────────────────

    def _serialize_app(self) -> Dict:
        """Convert full app state to a plain Python dict for cloud storage."""
        self._flush_fdd_vars()
        return {
            "yr":                  self.yr,
            "mo":                  self.mo,
            "next_doc_id":         self.next_doc_id,
            "next_leave_id":       self.next_leave_id,
            "next_spec_block_id":  self.next_spec_block_id,
            "next_manual_id":      self.next_manual_id,
            "docs": [
                {
                    "id":             d.id,
                    "name":           d.name,
                    "spec":           d.spec,
                    "team":           d.team,
                    "initials":       d.initials,
                    "first_duty_day": d.first_duty_day,
                }
                for d in self.docs
            ],
            "asgn": self.asgn,
            "leaves": [
                {"id": lb.id, "pid": lb.pid, "f": lb.f, "t": lb.t}
                for lb in self.leaves
            ],
            "spec_blocks": [
                {"id": sb.id, "code": sb.code, "f": sb.f, "t": sb.t}
                for sb in self.spec_blocks
            ],
            "manual_asgns": [
                {"id": ma.id, "pid": ma.pid, "code": ma.code, "day": ma.day}
                for ma in self.manual_asgns
            ],
        }

    def _deserialize_app(self, data: Dict) -> None:
        """Restore app state from a plain Python dict (from cloud storage)."""
        self.yr                 = int(data.get("yr",               self.yr))
        self.mo                 = int(data.get("mo",               self.mo))
        self.next_doc_id        = int(data.get("next_doc_id",        self.next_doc_id))
        self.next_leave_id      = int(data.get("next_leave_id",      self.next_leave_id))
        self.next_spec_block_id = int(data.get("next_spec_block_id", self.next_spec_block_id))
        self.next_manual_id     = int(data.get("next_manual_id",     self.next_manual_id))

        self.docs = [
            Doctor(
                id            = int(d["id"]),
                name          = d["name"],
                spec          = d["spec"],
                team          = d["team"],
                initials      = d.get("initials", ""),
                first_duty_day= int(d.get("first_duty_day", 1)),
            )
            for d in data.get("docs", [])
        ]

        self.asgn = data.get("asgn", {})

        self.leaves = [
            LeaveBlock(id=int(lb["id"]), pid=int(lb["pid"]),
                       f=lb["f"], t=lb["t"])
            for lb in data.get("leaves", [])
        ]

        self.spec_blocks = [
            SpecialtyBlock(id=int(sb["id"]), code=sb["code"],
                           f=sb["f"], t=sb["t"])
            for sb in data.get("spec_blocks", [])
        ]

        self.manual_asgns = [
            ManualAssignment(id=int(ma["id"]), pid=int(ma["pid"]),
                             code=ma["code"], day=int(ma["day"]))
            for ma in data.get("manual_asgns", [])
        ]

        self.refresh_all()


def main():
    app = MedSchedulerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
