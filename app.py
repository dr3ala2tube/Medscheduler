"""
MedScheduler Web – Flask backend
Exposes a REST API consumed by the single-page frontend.
Data is stored in Firebase Firestore (shared document for the whole team).
Authentication uses Firebase ID tokens verified via the REST API.
"""
from __future__ import annotations

import io
import json
import os
import urllib.request
import urllib.parse
import urllib.error
from functools import wraps
from typing import Any, Dict, List, Optional

from flask import Flask, jsonify, request, send_file, render_template, abort

from scheduler import (
    auto_schedule, compute_summary, dim, ds, day_of_week, is_we,
    Doctor, LeaveBlock, SpecialtyBlock, ManualAssignment,
    SHIFTS, COLOR_MAP, MONTHS, DN, TEAMS, SUBS, MORNING_K,
    DUTY_SET, OFF_SET, SPEC_OPTIONS, MANUAL_ASSIGN_CODES, BLOCKABLE_SPECIALTIES,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
FIREBASE_API_KEY  = "AIzaSyC_d-HgnEnLAWW1f3dSKjuuAz4eplcVWz8"
PROJECT_ID        = "medscheduler-e0853"
FIRESTORE_BASE    = (f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}"
                     f"/databases/(default)/documents")
SHARED_DOC        = f"{FIRESTORE_BASE}/shared/schedule"

app = Flask(__name__)
app.config["JSON_SORT_KEYS"] = False


# ── Firestore helpers (urllib, no SDK) ────────────────────────────────────────

def _http_json(method: str, url: str, body: Optional[bytes] = None,
               headers: Optional[Dict] = None, timeout: int = 15) -> Dict:
    req = urllib.request.Request(url, data=body, method=method)
    req.add_header("Content-Type", "application/json")
    if headers:
        for k, v in headers.items():
            req.add_header(k, v)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as exc:
        raw = exc.read()
        try:
            return json.loads(raw)
        except Exception:
            return {"error": {"code": exc.code, "message": raw.decode(errors="replace")[:300]}}
    except urllib.error.URLError as exc:
        return {"error": {"code": 0, "message": str(exc.reason)}}


def _py_to_fs(value: Any) -> Dict:
    if value is None:           return {"nullValue": None}
    if isinstance(value, bool): return {"booleanValue": value}
    if isinstance(value, int):  return {"integerValue": str(value)}
    if isinstance(value, float):return {"doubleValue": value}
    if isinstance(value, str):  return {"stringValue": value}
    if isinstance(value, list): return {"arrayValue": {"values": [_py_to_fs(v) for v in value]}}
    if isinstance(value, dict): return {"mapValue": {"fields": {k: _py_to_fs(v) for k, v in value.items()}}}
    return {"stringValue": str(value)}


def _fs_to_py(value: Dict) -> Any:
    if "nullValue"    in value: return None
    if "booleanValue" in value: return value["booleanValue"]
    if "integerValue" in value: return int(value["integerValue"])
    if "doubleValue"  in value: return float(value["doubleValue"])
    if "stringValue"  in value: return value["stringValue"]
    if "arrayValue"   in value: return [_fs_to_py(v) for v in value["arrayValue"].get("values", [])]
    if "mapValue"     in value: return {k: _fs_to_py(v) for k, v in value["mapValue"].get("fields", {}).items()}
    return None


def fs_load(id_token: str) -> Optional[Dict]:
    data = _http_json("GET", SHARED_DOC,
                      headers={"Authorization": f"Bearer {id_token}"})
    if "error" in data:
        code = data["error"].get("code", 0)
        if code == 404:
            return None
        raise Exception(data["error"].get("message", "Firestore read error"))
    if "fields" not in data:
        return None
    return _fs_to_py({"mapValue": {"fields": data["fields"]}})


def fs_save(id_token: str, payload: Dict) -> None:
    body = json.dumps({"fields": _py_to_fs(payload)["mapValue"]["fields"]}).encode()
    data = _http_json("PATCH", SHARED_DOC, body=body,
                      headers={"Authorization": f"Bearer {id_token}"}, timeout=20)
    if "error" in data:
        raise Exception(data["error"].get("message", "Firestore write error"))


# ── Firebase token verification ───────────────────────────────────────────────

def verify_token(id_token: str) -> Optional[Dict]:
    """Verify a Firebase ID token and return the decoded payload, or None."""
    url = (f"https://identitytoolkit.googleapis.com/v1/accounts:lookup"
           f"?key={FIREBASE_API_KEY}")
    body = json.dumps({"idToken": id_token}).encode()
    data = _http_json("POST", url, body=body, timeout=10)
    if "error" in data or "users" not in data:
        return None
    return data["users"][0]


def require_auth(f):
    """Decorator: extracts Bearer token, verifies it, injects (token, user) into handler."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"error": "Missing auth token"}), 401
        token = auth[7:]
        user  = verify_token(token)
        if user is None:
            return jsonify({"error": "Invalid or expired token"}), 401
        return f(token, user, *args, **kwargs)
    return wrapper


# ── Data deserialization helpers ──────────────────────────────────────────────

def _docs_from(raw: List) -> List[Doctor]:
    return [Doctor(id=int(d["id"]), name=d["name"], spec=d["spec"], team=d["team"],
                   initials=d.get("initials",""), first_duty_day=int(d.get("first_duty_day",1)))
            for d in raw]

def _leaves_from(raw: List) -> List[LeaveBlock]:
    return [LeaveBlock(id=int(x["id"]),pid=int(x["pid"]),f=x["f"],t=x["t"]) for x in raw]

def _spec_blocks_from(raw: List) -> List[SpecialtyBlock]:
    return [SpecialtyBlock(id=int(x["id"]),code=x["code"],f=x["f"],t=x["t"]) for x in raw]

def _manual_from(raw: List) -> List[ManualAssignment]:
    return [ManualAssignment(id=int(x["id"]),pid=int(x["pid"]),code=x["code"],day=int(x["day"]))
            for x in raw]


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/constants")
def api_constants():
    """Return all static constants the frontend needs."""
    return jsonify({
        "shifts":       SHIFTS,
        "color_map":    COLOR_MAP,
        "months":       MONTHS,
        "dn":           DN,
        "teams":        TEAMS,
        "subs":         SUBS,
        "morning_k":    MORNING_K,
        "duty_set":     list(DUTY_SET),
        "off_set":      list(OFF_SET),
        "spec_options": SPEC_OPTIONS,
        "manual_codes": MANUAL_ASSIGN_CODES,
        "blockable":    BLOCKABLE_SPECIALTIES,
    })


@app.route("/api/data", methods=["GET"])
@require_auth
def api_load(token, user):
    try:
        data = fs_load(token)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    return jsonify({"data": data})


@app.route("/api/data", methods=["POST"])
@require_auth
def api_save(token, user):
    payload = request.get_json(force=True)
    try:
        fs_save(token, payload)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    return jsonify({"ok": True})


@app.route("/api/schedule", methods=["POST"])
@require_auth
def api_schedule(token, user):
    body  = request.get_json(force=True)
    docs  = _docs_from(body.get("docs", []))
    asgn  = body.get("asgn", {})
    yr    = int(body["yr"])
    mo    = int(body["mo"])
    leaves       = _leaves_from(body.get("leaves", []))
    spec_blocks  = _spec_blocks_from(body.get("spec_blocks", []))
    manual_asgns = _manual_from(body.get("manual_asgns", []))

    # Stamp manual assignments into base asgn before scheduling
    for ma in manual_asgns:
        asgn[f"{ma.pid}|{yr}|{mo}|{ma.day}"] = ma.code

    result = auto_schedule(docs, asgn, leaves, spec_blocks, yr, mo)
    if "err" in result:
        return jsonify({"error": result["err"]}), 400

    # Re-stamp manual assignments after scheduling (prevent overwrite)
    for ma in manual_asgns:
        result["a"][f"{ma.pid}|{yr}|{mo}|{ma.day}"] = ma.code

    return jsonify({"asgn": result["a"], "pairs": result["pairs"]})


@app.route("/api/summary", methods=["POST"])
@require_auth
def api_summary(token, user):
    body = request.get_json(force=True)
    docs = _docs_from(body.get("docs", []))
    asgn = body.get("asgn", {})
    yr   = int(body["yr"])
    mo   = int(body["mo"])
    rows = compute_summary(docs, asgn, yr, mo)
    return jsonify({"rows": rows})


@app.route("/api/export/rota", methods=["POST"])
@require_auth
def api_export_rota(token, user):
    body = request.get_json(force=True)
    docs = _docs_from(body.get("docs", []))
    asgn = body.get("asgn", {})
    yr   = int(body["yr"])
    mo   = int(body["mo"])

    def get(pid, d): return asgn.get(f"{pid}|{yr}|{mo}|{d}", "_")
    def p2(n): return str(n).zfill(2)
    def gen_initials(docs):
        used, out = set(), {}
        for ph in docs:
            if ph.initials:
                out[ph.id] = ph.initials; used.add(ph.initials.upper()); continue
            words = ph.name.upper().split()
            words = [w for w in words if w not in ("DR","DR.")]
            cands = []
            if len(words) >= 2: cands.append(words[0][0]+words[1][0])
            if words: cands += [words[0][:i] for i in (3,4,2)]
            init = next((c for c in cands if c not in used), words[0][:3] if words else "?")
            used.add(init); out[ph.id] = init
        return out

    td = dim(yr, mo)
    initials_map = gen_initials(docs)

    ROTA_COLS = [
        ("T1","T1"),("T2","T2"),("T3","T3"),
        ("PU","PUL"),("CAHM","CA/HM"),("NE","NEU"),
        ("NP","NEPH"),("GI","GAS"),("NENP","NE/NP"),("DC","Daycare"),
        ("DM","On-Call(M)"),("DF","On-Call(F)"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Rota"

    title = f"First On Call Internal Medicine ROTA / {MONTHS[mo].upper()} {yr}"
    ws.merge_cells(f"A1:{get_column_letter(3+len(ROTA_COLS))}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, (_, lbl) in enumerate(ROTA_COLS, start=3):
        cell = ws.cell(row=2, column=ci, value=lbl)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=2,column=1,value="Day").font = Font(bold=True)
    ws.cell(row=2,column=2,value="Date").font = Font(bold=True)
    ws.cell(row=2,column=1).fill = PatternFill("solid",fgColor="D9EAF7")
    ws.cell(row=2,column=2).fill = PatternFill("solid",fgColor="D9EAF7")
    ws.row_dimensions[2].height = 32

    AMBER = "FFF3CD"
    for d in range(1, td+1):
        row = d + 2
        dw  = day_of_week(yr, mo, d)
        is_weekend = is_we(yr, mo, d)
        day_lbl = f"{p2(d)} {DN[dw]}"
        date_lbl = f"{yr}-{p2(mo+1)}-{p2(d)}"
        ws.cell(row=row, column=1, value=day_lbl).alignment  = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=date_lbl).alignment = Alignment(horizontal="center")
        if is_weekend:
            for c in range(1, 3+len(ROTA_COLS)+1):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=AMBER)

        for ci, (code, _) in enumerate(ROTA_COLS, start=3):
            ph = next((p for p in docs if get(p.id, d) == code), None)
            if code in ("DM","DF") and ph is None:
                ph = next((p for p in docs if get(p.id, d)==("DM" if code=="DM" else "DF")), None)
            val = initials_map.get(ph.id, "?") if ph else ""
            cell = ws.cell(row=row, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if not is_weekend and val:
                cell.fill = PatternFill("solid", fgColor=COLOR_MAP.get(code,"FFFFFF"))

    # Legend
    leg_col = 3 + len(ROTA_COLS) + 1
    ws.cell(row=2, column=leg_col, value="Legend").font = Font(bold=True)
    for pi, ph in enumerate(docs, start=3):
        ws.cell(row=pi, column=leg_col,   value=initials_map.get(ph.id,""))
        ws.cell(row=pi, column=leg_col+1, value=ph.name)

    for ci, _ in enumerate(ROTA_COLS, start=3):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions[get_column_letter(leg_col)].width = 8
    ws.column_dimensions[get_column_letter(leg_col+1)].width = 22
    ws.freeze_panes = f"C3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Rota_{MONTHS[mo]}_{yr}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/export/full", methods=["POST"])
@require_auth
def api_export_full(token, user):
    body = request.get_json(force=True)
    docs = _docs_from(body.get("docs", []))
    asgn = body.get("asgn", {})
    yr   = int(body["yr"])
    mo   = int(body["mo"])

    def get(pid, d): return asgn.get(f"{pid}|{yr}|{mo}|{d}", "_")
    def p2(n): return str(n).zfill(2)

    td   = dim(yr, mo)
    wb   = Workbook()
    ws   = wb.active
    ws.title = "Schedule"

    ws["A1"] = f"Physician Schedule — {MONTHS[mo]} {yr}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.cell(row=3,column=1,value="Day").font  = Font(bold=True)
    ws.cell(row=3,column=2,value="Date").font = Font(bold=True)
    for col in (1,2):
        ws.cell(row=3,column=col).fill = PatternFill("solid",fgColor="D9EAF7")
        ws.cell(row=3,column=col).alignment = Alignment(horizontal="center",vertical="center")

    doc_start = 3
    for ci, ph in enumerate(docs, start=doc_start):
        lbl = ph.initials if ph.initials else ph.name
        ws.cell(row=2,column=ci,value=lbl).font = Font(bold=True)
        ws.cell(row=2,column=ci).fill = PatternFill("solid",fgColor="D9EAF7")
        ws.cell(row=3,column=ci,value=f"{ph.spec} • {ph.team}").font = Font(bold=True)
        ws.cell(row=3,column=ci).fill = PatternFill("solid",fgColor="EEF5FB")
        for r in (2,3):
            ws.cell(row=r,column=ci).alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    for d in range(1, td+1):
        row = 3 + d
        dw  = day_of_week(yr, mo, d)
        ws.cell(row=row,column=1,value=f"{p2(d)} {DN[dw]}").alignment = Alignment(horizontal="center",vertical="center")
        ws.cell(row=row,column=2,value=f"{yr}-{p2(mo+1)}-{p2(d)}").alignment = Alignment(horizontal="center",vertical="center")
        for ci, ph in enumerate(docs, start=doc_start):
            code = get(ph.id, d)
            cell = ws.cell(row=row,column=ci,value=SHIFTS[code]["short"] or "")
            cell.alignment = Alignment(horizontal="center",vertical="center")
            cell.fill = PatternFill("solid",fgColor=COLOR_MAP.get(code,"FFFFFF"))

    # Summary rows
    summary_rows = compute_summary(docs, asgn, yr, mo)
    sum_start = td + 6
    ws.cell(row=sum_start,column=1,value="Summary").font = Font(bold=True,size=12)
    metrics = [("Total Hrs","total"),("On-Calls","calls"),("Daycare","daycare"),
               ("Post-Call","postcall"),("Off","off"),("Leave","leave"),("Random","random")]
    for ro,(lbl,_) in enumerate(metrics,start=1):
        c = ws.cell(row=sum_start+ro,column=1,value=lbl)
        c.font = Font(bold=True); c.fill = PatternFill("solid",fgColor="D9EAF7")
    stats_map = {r["name"]: r for r in summary_rows}
    for ci, ph in enumerate(docs, start=doc_start):
        stats = stats_map.get(ph.name, {})
        for ro,(_,key) in enumerate(metrics,start=1):
            ws.cell(row=sum_start+ro,column=ci,value=stats.get(key,0)).alignment = Alignment(horizontal="center",vertical="center")

    for ci_,col_cells in enumerate(ws.columns,start=1):
        maxl = max((len(str(c.value)) for c in col_cells if c.value),default=0)
        ws.column_dimensions[get_column_letter(ci_)].width = min(max(maxl+2,8),26)
    ws.freeze_panes="C4"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"Schedule_{MONTHS[mo]}_{yr}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
