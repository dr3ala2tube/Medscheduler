"""
MedScheduler Web – Flask backend
Exposes a REST API consumed by the single-page frontend.
Data is stored in Firebase Firestore, one private workspace per user
(workspaces/{uid}) with email-based member sharing enforced by Firestore
security rules. Authentication uses Firebase ID tokens verified via the
REST API; all Firestore calls are made with the caller's own token, so
access control is enforced by the rules, not by this server.
"""
from __future__ import annotations

import io
import json
import os
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timezone
from functools import wraps
from typing import Any, Dict, List, Optional

from flask import Flask, jsonify, request, send_file, render_template, abort, send_from_directory

from scheduler import (
    auto_schedule, compute_summary, dim, ds, day_of_week, is_we,
    Doctor, LeaveBlock, SpecialtyBlock, ManualAssignment,
    ScheduleRules, DEFAULT_RULES,
    ShiftEntry, ShiftConfig, DEFAULT_SHIFT_CONFIG,
    SHIFTS, COLOR_MAP, MONTHS, DN, TEAMS, SUBS, MORNING_K,
    DUTY_SET, OFF_SET, SPEC_OPTIONS, MANUAL_ASSIGN_CODES, BLOCKABLE_SPECIALTIES,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
FIREBASE_API_KEY  = "AIzaSyC_d-HgnEnLAWW1f3dSKjuuAz4eplcVWz8"
PROJECT_ID        = "medscheduler-e0853"
FIRESTORE_BASE    = (f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}"
                     f"/databases/(default)/documents")

# Accounts created at/after this moment (epoch ms, 2026-06-10T00:00:00Z) must
# have a verified email; accounts created before it are grandfathered.
# MUST match VERIFICATION_CUTOFF_MS in templates/index.html.
VERIFICATION_CUTOFF_MS = 1781049600000

def ws_meta_url(ws_id: str) -> str:
    """Workspace meta/ACL document (owner_uid, owner_email, members)."""
    return f"{FIRESTORE_BASE}/workspaces/{ws_id}"

def ws_data_url(ws_id: str) -> str:
    """Workspace schedule payload document."""
    return f"{FIRESTORE_BASE}/workspaces/{ws_id}/data/schedule"

app = Flask(__name__)
app.config["JSON_SORT_KEYS"] = False


# ── Firestore helpers (urllib, no SDK) ────────────────────────────────────────

def _http_json(method: str, url: str, body: Optional[bytes] = None,
               headers: Optional[Dict] = None, timeout: int = 15) -> Dict:
    req = urllib.request.Request(url, data=body, method=method)
    req.add_header("Content-Type", "application/json")
    if headers:
        for k, v in headers.items():
            req.add_header(k, v)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as exc:
        raw = exc.read()
        try:
            return json.loads(raw)
        except Exception:
            return {"error": {"code": exc.code, "message": raw.decode(errors="replace")[:300]}}
    except urllib.error.URLError as exc:
        return {"error": {"code": 0, "message": str(exc.reason)}}


def _py_to_fs(value: Any) -> Dict:
    if value is None:           return {"nullValue": None}
    if isinstance(value, bool): return {"booleanValue": value}
    if isinstance(value, int):  return {"integerValue": str(value)}
    if isinstance(value, float):return {"doubleValue": value}
    if isinstance(value, str):  return {"stringValue": value}
    if isinstance(value, list): return {"arrayValue": {"values": [_py_to_fs(v) for v in value]}}
    if isinstance(value, dict): return {"mapValue": {"fields": {k: _py_to_fs(v) for k, v in value.items()}}}
    return {"stringValue": str(value)}


def _fs_to_py(value: Dict) -> Any:
    if "nullValue"    in value: return None
    if "booleanValue" in value: return value["booleanValue"]
    if "integerValue" in value: return int(value["integerValue"])
    if "doubleValue"  in value: return float(value["doubleValue"])
    if "stringValue"  in value: return value["stringValue"]
    if "arrayValue"   in value: return [_fs_to_py(v) for v in value["arrayValue"].get("values", [])]
    if "mapValue"     in value: return {k: _fs_to_py(v) for k, v in value["mapValue"].get("fields", {}).items()}
    return None


class FsError(Exception):
    """Firestore REST error carrying the upstream HTTP code."""
    def __init__(self, code: int, message: str):
        super().__init__(message)
        self.code = code

    def http_status(self) -> int:
        # Surface permission denials as 403; everything else is a server error.
        return 403 if self.code == 403 else 500


def fs_load(id_token: str, url: str) -> Optional[Dict]:
    data = _http_json("GET", url,
                      headers={"Authorization": f"Bearer {id_token}"})
    if "error" in data:
        code = data["error"].get("code", 0)
        if code == 404:
            return None
        raise FsError(code, data["error"].get("message", "Firestore read error"))
    if "fields" not in data:
        return None
    return _fs_to_py({"mapValue": {"fields": data["fields"]}})


def fs_save(id_token: str, url: str, payload: Dict) -> None:
    body = json.dumps({"fields": _py_to_fs(payload)["mapValue"]["fields"]}).encode()
    data = _http_json("PATCH", url, body=body,
                      headers={"Authorization": f"Bearer {id_token}"}, timeout=20)
    if "error" in data:
        code = data["error"].get("code", 0)
        raise FsError(code, data["error"].get("message", "Firestore write error"))


def fs_create(id_token: str, parent_url: str, payload: Dict) -> str:
    """Create a document with an auto-generated id; returns the new id."""
    body = json.dumps({"fields": _py_to_fs(payload)["mapValue"]["fields"]}).encode()
    data = _http_json("POST", parent_url, body=body,
                      headers={"Authorization": f"Bearer {id_token}"}, timeout=20)
    if "error" in data:
        code = data["error"].get("code", 0)
        raise FsError(code, data["error"].get("message", "Firestore create error"))
    return data.get("name", "").rsplit("/", 1)[-1]


def fs_patch_fields(id_token: str, doc_url: str, fields: Dict) -> None:
    """PATCH only the given fields of an existing document (updateMask +
    exists precondition, so a bad id can never create a stray document)."""
    mask = "&".join("updateMask.fieldPaths=" + urllib.parse.quote(k) for k in fields)
    url  = f"{doc_url}?{mask}&currentDocument.exists=true"
    body = json.dumps({"fields": {k: _py_to_fs(v) for k, v in fields.items()}}).encode()
    data = _http_json("PATCH", url, body=body,
                      headers={"Authorization": f"Bearer {id_token}"}, timeout=20)
    if "error" in data:
        code = data["error"].get("code", 0)
        raise FsError(code, data["error"].get("message", "Firestore patch error"))


def fs_query_shared_workspaces(id_token: str, email: str,
                               field: str = "members") -> List[Dict]:
    """Return workspaces whose `field` array (members or invites) contains
    `email`.

    Uses Firestore runQuery; the security rules allow this list operation
    only when the query is constrained to the caller's own email.
    """
    body = json.dumps({
        "structuredQuery": {
            "from": [{"collectionId": "workspaces"}],
            "where": {
                "fieldFilter": {
                    "field": {"fieldPath": field},
                    "op":    "ARRAY_CONTAINS",
                    "value": {"stringValue": email},
                }
            },
        }
    }).encode()
    data = _http_json("POST", f"{FIRESTORE_BASE}:runQuery", body=body,
                      headers={"Authorization": f"Bearer {id_token}"}, timeout=20)
    if isinstance(data, dict) and "error" in data:
        code = data["error"].get("code", 0)
        raise FsError(code, data["error"].get("message", "Firestore query error"))
    results = []
    for item in (data if isinstance(data, list) else []):
        doc = item.get("document")
        if not doc:
            continue
        fields = _fs_to_py({"mapValue": {"fields": doc.get("fields", {})}})
        fields["_id"] = doc["name"].rsplit("/", 1)[-1]
        results.append(fields)
    return results


def fs_query_notifications(id_token: str, uid: str, limit: int = 50) -> List[Dict]:
    """Return the user's notifications, newest first."""
    body = json.dumps({
        "structuredQuery": {
            "from": [{"collectionId": "items"}],
            "orderBy": [{"field": {"fieldPath": "created"},
                         "direction": "DESCENDING"}],
            "limit": limit,
        }
    }).encode()
    data = _http_json("POST", f"{FIRESTORE_BASE}/notifications/{uid}:runQuery",
                      body=body,
                      headers={"Authorization": f"Bearer {id_token}"}, timeout=20)
    if isinstance(data, dict) and "error" in data:
        code = data["error"].get("code", 0)
        raise FsError(code, data["error"].get("message", "Firestore query error"))
    items = []
    for item in (data if isinstance(data, list) else []):
        doc = item.get("document")
        if not doc:
            continue
        fields = _fs_to_py({"mapValue": {"fields": doc.get("fields", {})}})
        fields["id"] = doc["name"].rsplit("/", 1)[-1]
        items.append(fields)
    return items


# ── Firebase token verification ───────────────────────────────────────────────

def verify_token(id_token: str) -> Optional[Dict]:
    """Verify a Firebase ID token and return the decoded payload, or None."""
    url = (f"https://identitytoolkit.googleapis.com/v1/accounts:lookup"
           f"?key={FIREBASE_API_KEY}")
    body = json.dumps({"idToken": id_token}).encode()
    data = _http_json("POST", url, body=body, timeout=10)
    if "error" in data or "users" not in data:
        return None
    return data["users"][0]


def require_auth(f):
    """Decorator: extracts Bearer token, verifies it, injects (token, user) into handler."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return jsonify({"error": "Missing auth token"}), 401
        token = auth[7:]
        user  = verify_token(token)
        if user is None:
            return jsonify({"error": "Invalid or expired token"}), 401
        # Email verification: required for accounts created after the cutoff
        # (identitytoolkit lookup returns createdAt in epoch ms and emailVerified).
        try:
            created_ms = int(user.get("createdAt", "0"))
        except (TypeError, ValueError):
            created_ms = 0
        if created_ms >= VERIFICATION_CUTOFF_MS and not user.get("emailVerified", False):
            return jsonify({"error": "email-not-verified"}), 403
        return f(token, user, *args, **kwargs)
    return wrapper


# ── Data deserialization helpers ──────────────────────────────────────────────

def _docs_from(raw: List) -> List[Doctor]:
    return [Doctor(id=int(d["id"]), name=d["name"], spec=d["spec"], team=d["team"],
                   initials=d.get("initials",""), first_duty_day=int(d.get("first_duty_day",1)))
            for d in raw]

def _leaves_from(raw: List) -> List[LeaveBlock]:
    return [LeaveBlock(id=int(x["id"]),pid=int(x["pid"]),f=x["f"],t=x["t"]) for x in raw]

def _spec_blocks_from(raw: List) -> List[SpecialtyBlock]:
    return [SpecialtyBlock(id=int(x["id"]),code=x["code"],f=x["f"],t=x["t"]) for x in raw]

def _manual_from(raw: List) -> List[ManualAssignment]:
    return [ManualAssignment(id=int(x["id"]),pid=int(x["pid"]),code=x["code"],day=int(x["day"]))
            for x in raw]

def _shift_config_from(raw: Optional[Dict]) -> ShiftConfig:
    """Parse a shift-config dict from the request body, falling back to defaults."""
    if not raw:
        return DEFAULT_SHIFT_CONFIG
    d = DEFAULT_SHIFT_CONFIG

    def parse_entries(lst, shift_type):
        result = []
        for item in (lst or []):
            result.append(ShiftEntry(
                code       = str(item.get("code", "")).strip().upper(),
                label      = str(item.get("label", "")),
                short      = str(item.get("short", "")),
                color      = str(item.get("color", "DBEAFE")),
                hours      = int(item.get("hours", 8)),
                shift_type = shift_type,
                enabled    = bool(item.get("enabled", True)),
            ))
        return result

    teams       = parse_entries(raw.get("teams"),       "team")
    specialties = parse_entries(raw.get("specialties"), "specialty")
    duties      = parse_entries(raw.get("duties"),      "duty")

    # Enforce minimums so the engine never crashes
    if len(teams) < 3:
        teams = d.teams
    if len(duties) < 1:
        duties = d.duties

    # Parse clinics — support new format (clinics list) and old format (dc_label/dc_short/dc_color)
    clinics_raw = raw.get("clinics")
    if clinics_raw is not None:
        clinics = parse_entries(clinics_raw, "clinic")
        # Allow 0 clinics (disables clinic rotation)
    else:
        # Backward compat: old saved data used dc_label/dc_short/dc_color
        default_clinic = d.clinics[0] if d.clinics else ShiftEntry("DC","GP Clinic","DC","FEF3C7",8,"clinic")
        clinics = [ShiftEntry(
            code       = "DC",
            label      = str(raw.get("dc_label", default_clinic.label)),
            short      = str(raw.get("dc_short", default_clinic.short)),
            color      = str(raw.get("dc_color", default_clinic.color)),
            hours      = 8,
            shift_type = "clinic",
        )]

    return ShiftConfig(
        teams=teams, specialties=specialties, duties=duties, clinics=clinics,
    )


def _rules_from(raw: Optional[Dict]) -> ScheduleRules:
    """Parse a rules dict from the request body, falling back to defaults for missing keys."""
    if not raw:
        return ScheduleRules()
    d = DEFAULT_RULES
    return ScheduleRules(
        max_consecutive_days = int(raw.get("max_consecutive_days", d.max_consecutive_days)),
        post_call_days       = int(raw.get("post_call_days",       d.post_call_days)),
        min_duties           = int(raw.get("min_duties",           d.min_duties)),
        max_duties           = int(raw.get("max_duties",           d.max_duties)),
        min_hours            = int(raw.get("min_hours",            d.min_hours)),
        max_hours            = int(raw.get("max_hours",            d.max_hours)),
        duty_shift_hours     = int(raw.get("duty_shift_hours",     d.duty_shift_hours)),
        morning_shift_hours  = int(raw.get("morning_shift_hours",  d.morning_shift_hours)),
        enforce_weekend_off  = bool(raw.get("enforce_weekend_off", d.enforce_weekend_off)),
    )


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# ── PWA static files (must be served from root for correct SW scope) ──────────
@app.route("/manifest.json")
def pwa_manifest():
    return send_from_directory(app.static_folder, "manifest.json")


@app.route("/sw.js")
def pwa_sw():
    resp = send_from_directory(app.static_folder, "sw.js")
    resp.headers["Service-Worker-Allowed"] = "/"
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


@app.route("/api/constants")
def api_constants():
    """Return all static constants the frontend needs (based on default shift config)."""
    return jsonify({
        "shifts":       SHIFTS,
        "color_map":    COLOR_MAP,
        "months":       MONTHS,
        "dn":           DN,
        "teams":        TEAMS,
        "subs":         SUBS,
        "morning_k":    MORNING_K,
        "duty_set":     list(DUTY_SET),
        "off_set":      list(OFF_SET),
        "spec_options": SPEC_OPTIONS,
        "manual_codes": MANUAL_ASSIGN_CODES,
        "blockable":    BLOCKABLE_SPECIALTIES,
        # Default shift config for the frontend to initialise S.shiftConfig
        "default_shift_config": {
            "teams":      [{"code":e.code,"label":e.label,"short":e.short,"color":e.color,"hours":e.hours,"enabled":e.enabled} for e in DEFAULT_SHIFT_CONFIG.teams],
            "specialties":[{"code":e.code,"label":e.label,"short":e.short,"color":e.color,"hours":e.hours,"enabled":e.enabled} for e in DEFAULT_SHIFT_CONFIG.specialties],
            "duties":     [{"code":e.code,"label":e.label,"short":e.short,"color":e.color,"hours":e.hours,"enabled":e.enabled} for e in DEFAULT_SHIFT_CONFIG.duties],
            "clinics":    [{"code":e.code,"label":e.label,"short":e.short,"color":e.color,"hours":e.hours,"enabled":e.enabled} for e in DEFAULT_SHIFT_CONFIG.clinics],
        },
    })


@app.route("/api/rules/defaults")
def api_rules_defaults():
    """Return the default scheduling rules so the frontend can pre-populate the form."""
    r = DEFAULT_RULES
    return jsonify({
        "max_consecutive_days": r.max_consecutive_days,
        "post_call_days":       r.post_call_days,
        "min_duties":           r.min_duties,
        "max_duties":           r.max_duties,
        "min_hours":            r.min_hours,
        "max_hours":            r.max_hours,
        "duty_shift_hours":     r.duty_shift_hours,
        "morning_shift_hours":  r.morning_shift_hours,
        "enforce_weekend_off":  r.enforce_weekend_off,
    })


@app.route("/api/shift-config/defaults")
def api_shift_config_defaults():
    """Return the default shift config so the frontend can pre-populate the Shifts modal."""
    cfg = DEFAULT_SHIFT_CONFIG
    return jsonify({
        "teams":      [{"code":e.code,"label":e.label,"short":e.short,"color":e.color,"hours":e.hours,"enabled":e.enabled} for e in cfg.teams],
        "specialties":[{"code":e.code,"label":e.label,"short":e.short,"color":e.color,"hours":e.hours,"enabled":e.enabled} for e in cfg.specialties],
        "duties":     [{"code":e.code,"label":e.label,"short":e.short,"color":e.color,"hours":e.hours,"enabled":e.enabled} for e in cfg.duties],
        "clinics":    [{"code":e.code,"label":e.label,"short":e.short,"color":e.color,"hours":e.hours,"enabled":e.enabled} for e in cfg.clinics],
    })


# ── Workspaces ────────────────────────────────────────────────────────────────

def _uid(user: Dict) -> str:
    return user["localId"]


def _email(user: Dict) -> str:
    return user.get("email", "").strip().lower()


def _ws_param(user: Dict) -> Optional[str]:
    """Workspace id from ?ws= query arg; defaults to the caller's own uid.
    Returns None if the supplied id is malformed."""
    ws = request.args.get("ws", "").strip()
    if not ws:
        return _uid(user)
    if not ws.isalnum() or len(ws) > 128:
        return None
    return ws


def _ensure_own_meta(token: str, user: Dict) -> Dict:
    """Load the caller's workspace meta doc, creating it on first login."""
    meta = fs_load(token, ws_meta_url(_uid(user)))
    if meta is None:
        meta = {"owner_uid": _uid(user), "owner_email": _email(user),
                "members": [], "invites": []}
        fs_save(token, ws_meta_url(_uid(user)), meta)
    return meta


@app.route("/api/workspaces", methods=["GET"])
@require_auth
def api_workspaces(token, user):
    try:
        meta = _ensure_own_meta(token, user)
        shared, pending = [], []
        email = _email(user)
        if email:
            for ws in fs_query_shared_workspaces(token, email):
                if ws.get("_id") and ws["_id"] != _uid(user):
                    shared.append({"id": ws["_id"],
                                   "owner_email": ws.get("owner_email", "")})
            for ws in fs_query_shared_workspaces(token, email, field="invites"):
                if ws.get("_id") and ws["_id"] != _uid(user):
                    pending.append({"id": ws["_id"],
                                    "owner_email": ws.get("owner_email", "")})
    except FsError as exc:
        return jsonify({"error": str(exc)}), exc.http_status()
    return jsonify({
        "own": {"id": _uid(user), "owner_email": _email(user),
                "members": meta.get("members", []),
                "invites": meta.get("invites", [])},
        "shared": shared,
        "invites": pending,
    })


@app.route("/api/workspaces/members", methods=["POST"])
@require_auth
def api_workspace_members(token, user):
    """Owner-only: invite an email (pending until accepted) or remove an
    email (cancels a pending invite and/or revokes membership). Ownership is
    also enforced by Firestore rules (writes target the caller's own meta doc
    and rules reject non-owner updates)."""
    body   = request.get_json(force=True)
    action = str(body.get("action", "")).strip().lower()
    member = str(body.get("email", "")).strip().lower()
    if action not in ("add", "remove"):
        return jsonify({"error": "action must be 'add' or 'remove'"}), 400
    if not member or "@" not in member or len(member) > 254:
        return jsonify({"error": "A valid email address is required"}), 400
    if member == _email(user):
        return jsonify({"error": "You already own this workspace"}), 400
    try:
        meta = _ensure_own_meta(token, user)
        members = list(meta.get("members", []))
        invites = list(meta.get("invites", []))
        if action == "add":
            if member in members:
                return jsonify({"error": "Already a member"}), 400
            if member not in invites:
                invites.append(member)
        else:  # remove
            members = [m for m in members if m != member]
            invites = [m for m in invites if m != member]
        meta["members"] = members
        meta["invites"] = invites
        fs_save(token, ws_meta_url(_uid(user)), meta)
    except FsError as exc:
        return jsonify({"error": str(exc)}), exc.http_status()
    return jsonify({"ok": True, "members": members, "invites": invites})


@app.route("/api/data", methods=["GET"])
@require_auth
def api_load(token, user):
    ws = _ws_param(user)
    if ws is None:
        return jsonify({"error": "Invalid workspace id"}), 400
    try:
        data = fs_load(token, ws_data_url(ws))
    except FsError as exc:
        return jsonify({"error": str(exc)}), exc.http_status()
    return jsonify({"data": data, "ws": ws})


@app.route("/api/data", methods=["POST"])
@require_auth
def api_save(token, user):
    ws = _ws_param(user)
    if ws is None:
        return jsonify({"error": "Invalid workspace id"}), 400
    payload = request.get_json(force=True)
    try:
        fs_save(token, ws_data_url(ws), payload)
    except FsError as exc:
        return jsonify({"error": str(exc)}), exc.http_status()
    return jsonify({"ok": True, "ws": ws})


def _notify(token: str, recipient_uid: str, ntype: str, actor_email: str,
            ws_id: str) -> bool:
    """Write a notification into the recipient's feed. Best-effort: returns
    False instead of raising so a failed notification never rolls back the
    membership change it reports on."""
    try:
        fs_create(token, f"{FIRESTORE_BASE}/notifications/{recipient_uid}/items", {
            "type":        ntype,
            "actor_email": actor_email,
            "ws_id":       ws_id,
            "created":     datetime.now(timezone.utc).isoformat(),
            "read":        False,
        })
        return True
    except FsError:
        return False


def _ws_from_body(body: Dict, user: Dict) -> Optional[str]:
    """Validated foreign workspace id from a JSON body (never the caller's own)."""
    ws = str(body.get("ws_id", "")).strip()
    if not ws or not ws.isalnum() or len(ws) > 128 or ws == _uid(user):
        return None
    return ws


@app.route("/api/invitations/respond", methods=["POST"])
@require_auth
def api_invitation_respond(token, user):
    """Invited user accepts or declines a pending invitation. The meta-doc
    transition is enforced server-side here and by Firestore rules
    (self-service diff: only the caller's own email may move)."""
    body   = request.get_json(force=True)
    action = str(body.get("action", "")).strip().lower()
    ws     = _ws_from_body(body, user)
    if action not in ("accept", "decline"):
        return jsonify({"error": "action must be 'accept' or 'decline'"}), 400
    if ws is None:
        return jsonify({"error": "Invalid workspace id"}), 400
    email = _email(user)
    try:
        meta = fs_load(token, ws_meta_url(ws))
        if meta is None:
            return jsonify({"error": "Workspace not found"}), 404
        invites = list(meta.get("invites", []))
        members = list(meta.get("members", []))
        if email not in invites:
            return jsonify({"error": "No pending invitation for this workspace"}), 404
        invites = [m for m in invites if m != email]
        if action == "accept" and email not in members:
            members.append(email)
        meta["invites"] = invites
        meta["members"] = members
        fs_save(token, ws_meta_url(ws), meta)
    except FsError as exc:
        return jsonify({"error": str(exc)}), exc.http_status()
    ntype = "invite_accepted" if action == "accept" else "invite_declined"
    notified = _notify(token, ws, ntype, email, ws)
    return jsonify({"ok": True, "action": action, "notified": notified})


@app.route("/api/workspaces/leave", methods=["POST"])
@require_auth
def api_workspace_leave(token, user):
    """Accepted member leaves a workspace they do not own."""
    body = request.get_json(force=True)
    ws   = _ws_from_body(body, user)
    if ws is None:
        return jsonify({"error": "Invalid workspace id"}), 400
    email = _email(user)
    try:
        meta = fs_load(token, ws_meta_url(ws))
        if meta is None:
            return jsonify({"error": "Workspace not found"}), 404
        members = list(meta.get("members", []))
        if email not in members:
            return jsonify({"error": "You are not a member of this workspace"}), 404
        meta["members"] = [m for m in members if m != email]
        meta["invites"] = list(meta.get("invites", []))
        fs_save(token, ws_meta_url(ws), meta)
    except FsError as exc:
        return jsonify({"error": str(exc)}), exc.http_status()
    notified = _notify(token, ws, "member_left", email, ws)
    return jsonify({"ok": True, "notified": notified})


@app.route("/api/notifications", methods=["GET"])
@require_auth
def api_notifications(token, user):
    """The caller's notifications, newest first, plus an unread count."""
    try:
        items = fs_query_notifications(token, _uid(user))
    except FsError as exc:
        return jsonify({"error": str(exc)}), exc.http_status()
    unread = sum(1 for i in items if not i.get("read"))
    return jsonify({"items": items, "unread": unread})


@app.route("/api/notifications/read", methods=["POST"])
@require_auth
def api_notifications_read(token, user):
    """Mark the given notification ids as read (best-effort per id)."""
    body = request.get_json(force=True)
    ids  = body.get("ids", [])
    if not isinstance(ids, list) or not ids or len(ids) > 100:
        return jsonify({"error": "ids must be a non-empty list (max 100)"}), 400
    updated = 0
    for nid in ids:
        nid = str(nid)
        if not nid.isalnum() or len(nid) > 128:
            continue
        try:
            fs_patch_fields(
                token,
                f"{FIRESTORE_BASE}/notifications/{_uid(user)}/items/{nid}",
                {"read": True})
            updated += 1
        except FsError:
            continue
    return jsonify({"ok": True, "updated": updated})


@app.route("/api/schedule", methods=["POST"])
@require_auth
def api_schedule(token, user):
    body  = request.get_json(force=True)
    docs  = _docs_from(body.get("docs", []))
    asgn  = body.get("asgn", {})
    yr    = int(body["yr"])
    mo    = int(body["mo"])
    leaves       = _leaves_from(body.get("leaves", []))
    spec_blocks  = _spec_blocks_from(body.get("spec_blocks", []))
    manual_asgns = _manual_from(body.get("manual_asgns", []))
    rules        = _rules_from(body.get("rules"))
    shift_config = _shift_config_from(body.get("shiftConfig"))

    # Stamp manual assignments into base asgn before scheduling
    for ma in manual_asgns:
        asgn[f"{ma.pid}|{yr}|{mo}|{ma.day}"] = ma.code

    result = auto_schedule(docs, asgn, leaves, spec_blocks, yr, mo,
                           rules=rules, shift_config=shift_config)
    if "err" in result:
        return jsonify({"error": result["err"]}), 400

    # Re-stamp manual assignments after scheduling (prevent overwrite)
    for ma in manual_asgns:
        result["a"][f"{ma.pid}|{yr}|{mo}|{ma.day}"] = ma.code

    return jsonify({"asgn": result["a"], "pairs": result["pairs"]})


@app.route("/api/summary", methods=["POST"])
@require_auth
def api_summary(token, user):
    body         = request.get_json(force=True)
    docs         = _docs_from(body.get("docs", []))
    asgn         = body.get("asgn", {})
    yr           = int(body["yr"])
    mo           = int(body["mo"])
    rules        = _rules_from(body.get("rules"))
    shift_config = _shift_config_from(body.get("shiftConfig"))
    rows         = compute_summary(docs, asgn, yr, mo, rules=rules, shift_config=shift_config)
    return jsonify({"rows": rows})


@app.route("/api/export/rota", methods=["POST"])
@require_auth
def api_export_rota(token, user):
    body         = request.get_json(force=True)
    docs         = _docs_from(body.get("docs", []))
    asgn         = body.get("asgn", {})
    yr           = int(body["yr"])
    mo           = int(body["mo"])
    shift_config = _shift_config_from(body.get("shiftConfig"))

    def get(pid, d): return asgn.get(f"{pid}|{yr}|{mo}|{d}", "_")
    def p2(n): return str(n).zfill(2)
    def gen_initials(docs):
        used, out = set(), {}
        for ph in docs:
            if ph.initials:
                out[ph.id] = ph.initials; used.add(ph.initials.upper()); continue
            words = ph.name.upper().split()
            words = [w for w in words if w not in ("DR","DR.")]
            cands = []
            if len(words) >= 2: cands.append(words[0][0]+words[1][0])
            if words: cands += [words[0][:i] for i in (3,4,2)]
            init = next((c for c in cands if c not in used), words[0][:3] if words else "?")
            used.add(init); out[ph.id] = init
        return out

    td = dim(yr, mo)
    initials_map = gen_initials(docs)
    color_map = shift_config.to_color_map()

    # Build ROTA_COLS dynamically from shift config
    sub_codes = shift_config.specialty_codes()
    ROTA_COLS = (
        [(e.code, e.short) for e in shift_config.teams] +
        [(e.code, e.short) for e in shift_config.specialties] +
        ([("NENP", "NE/NP")] if "NE" in sub_codes and "NP" in sub_codes else []) +
        [(e.code, e.short) for e in shift_config.clinics] +
        [(e.code, e.short) for e in shift_config.duties]
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Rota"

    title = f"First On Call Internal Medicine ROTA / {MONTHS[mo].upper()} {yr}"
    ws.merge_cells(f"A1:{get_column_letter(3+len(ROTA_COLS))}1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, (_, lbl) in enumerate(ROTA_COLS, start=3):
        cell = ws.cell(row=2, column=ci, value=lbl)
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=2,column=1,value="Day").font = Font(bold=True)
    ws.cell(row=2,column=2,value="Date").font = Font(bold=True)
    ws.cell(row=2,column=1).fill = PatternFill("solid",fgColor="D9EAF7")
    ws.cell(row=2,column=2).fill = PatternFill("solid",fgColor="D9EAF7")
    ws.row_dimensions[2].height = 32

    AMBER = "FFF3CD"
    for d in range(1, td+1):
        row = d + 2
        dw  = day_of_week(yr, mo, d)
        is_weekend = is_we(yr, mo, d)
        day_lbl = f"{p2(d)} {DN[dw]}"
        date_lbl = f"{yr}-{p2(mo+1)}-{p2(d)}"
        ws.cell(row=row, column=1, value=day_lbl).alignment  = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=date_lbl).alignment = Alignment(horizontal="center")
        if is_weekend:
            for c in range(1, 3+len(ROTA_COLS)+1):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=AMBER)

        for ci, (code, _) in enumerate(ROTA_COLS, start=3):
            ph = next((p for p in docs if get(p.id, d) == code), None)
            if code in ("DM","DF") and ph is None:
                ph = next((p for p in docs if get(p.id, d)==("DM" if code=="DM" else "DF")), None)
            val = initials_map.get(ph.id, "?") if ph else ""
            cell = ws.cell(row=row, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if not is_weekend and val:
                cell.fill = PatternFill("solid", fgColor=color_map.get(code,"FFFFFF"))

    # Legend
    leg_col = 3 + len(ROTA_COLS) + 1
    ws.cell(row=2, column=leg_col, value="Legend").font = Font(bold=True)
    for pi, ph in enumerate(docs, start=3):
        ws.cell(row=pi, column=leg_col,   value=initials_map.get(ph.id,""))
        ws.cell(row=pi, column=leg_col+1, value=ph.name)

    for ci, _ in enumerate(ROTA_COLS, start=3):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions[get_column_letter(leg_col)].width = 8
    ws.column_dimensions[get_column_letter(leg_col+1)].width = 22
    ws.freeze_panes = f"C3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Rota_{MONTHS[mo]}_{yr}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/export/full", methods=["POST"])
@require_auth
def api_export_full(token, user):
    body         = request.get_json(force=True)
    docs         = _docs_from(body.get("docs", []))
    asgn         = body.get("asgn", {})
    yr           = int(body["yr"])
    mo           = int(body["mo"])
    shift_config = _shift_config_from(body.get("shiftConfig"))

    def get(pid, d): return asgn.get(f"{pid}|{yr}|{mo}|{d}", "_")
    def p2(n): return str(n).zfill(2)

    td        = dim(yr, mo)
    color_map = shift_config.to_color_map()
    wb   = Workbook()
    ws   = wb.active
    ws.title = "Schedule"

    ws["A1"] = f"Physician Schedule — {MONTHS[mo]} {yr}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.cell(row=3,column=1,value="Day").font  = Font(bold=True)
    ws.cell(row=3,column=2,value="Date").font = Font(bold=True)
    for col in (1,2):
        ws.cell(row=3,column=col).fill = PatternFill("solid",fgColor="D9EAF7")
        ws.cell(row=3,column=col).alignment = Alignment(horizontal="center",vertical="center")

    doc_start = 3
    for ci, ph in enumerate(docs, start=doc_start):
        lbl = ph.initials if ph.initials else ph.name
        ws.cell(row=2,column=ci,value=lbl).font = Font(bold=True)
        ws.cell(row=2,column=ci).fill = PatternFill("solid",fgColor="D9EAF7")
        ws.cell(row=3,column=ci,value=f"{ph.spec} • {ph.team}").font = Font(bold=True)
        ws.cell(row=3,column=ci).fill = PatternFill("solid",fgColor="EEF5FB")
        for r in (2,3):
            ws.cell(row=r,column=ci).alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    for d in range(1, td+1):
        row = 3 + d
        dw  = day_of_week(yr, mo, d)
        ws.cell(row=row,column=1,value=f"{p2(d)} {DN[dw]}").alignment = Alignment(horizontal="center",vertical="center")
        ws.cell(row=row,column=2,value=f"{yr}-{p2(mo+1)}-{p2(d)}").alignment = Alignment(horizontal="center",vertical="center")
        for ci, ph in enumerate(docs, start=doc_start):
            code = get(ph.id, d)
            cell = ws.cell(row=row,column=ci,value=SHIFTS[code]["short"] or "")
            cell.alignment = Alignment(horizontal="center",vertical="center")
            cell.fill = PatternFill("solid",fgColor=color_map.get(code,"FFFFFF"))

    # Summary rows
    summary_rows = compute_summary(docs, asgn, yr, mo, shift_config=shift_config)
    sum_start = td + 6
    ws.cell(row=sum_start,column=1,value="Summary").font = Font(bold=True,size=12)
    metrics = [("Total Hrs","total"),("On-Calls","calls"),("Daycare","daycare"),
               ("Post-Call","postcall"),("Off","off"),("Leave","leave"),("Random","random")]
    for ro,(lbl,_) in enumerate(metrics,start=1):
        c = ws.cell(row=sum_start+ro,column=1,value=lbl)
        c.font = Font(bold=True); c.fill = PatternFill("solid",fgColor="D9EAF7")
    stats_map = {r["name"]: r for r in summary_rows}
    for ci, ph in enumerate(docs, start=doc_start):
        stats = stats_map.get(ph.name, {})
        for ro,(_,key) in enumerate(metrics,start=1):
            ws.cell(row=sum_start+ro,column=ci,value=stats.get(key,0)).alignment = Alignment(horizontal="center",vertical="center")

    for ci_,col_cells in enumerate(ws.columns,start=1):
        maxl = max((len(str(c.value)) for c in col_cells if c.value),default=0)
        ws.column_dimensions[get_column_letter(ci_)].width = min(max(maxl+2,8),26)
    ws.freeze_panes="C4"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"Schedule_{MONTHS[mo]}_{yr}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
